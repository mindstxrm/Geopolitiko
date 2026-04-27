"""Flask routes for the geopolitical news site."""
from typing import Optional

import json
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from urllib.parse import quote
from flask import Blueprint, current_app, render_template, request, redirect, url_for, session, Response, flash
from flask_login import current_user, login_required, login_user, logout_user

from app.auth import get_effective_user_id, verify_password
from app.models import (
    CHOKEPOINT_COORDS,
    get_articles,
    get_articles_trade_supply_chain_relevant,
    get_articles_count,
    get_feed_event_type_counts,
    get_article,
    article_confidence_score,
    article_signal_score,
    get_feed_event_types,
    get_feed_country_options,
    get_sources,
    get_cluster_label,
    search_articles,
    get_trending_topics,
    get_digests,
    get_digest,
    get_clusters_with_counts,
    get_articles_by_cluster,
    get_all_topic_counts,
    get_watchlist,
    get_watchlists,
    get_articles_for_watchlist,
    watchlist_add_topic,
    add_watchlist,
    update_watchlist,
    get_alerts,
    get_alert,
    get_alert_suggestions,
    get_alert_matches,
    get_alert_match_counts,
    get_impact_summary,
    get_domain_counts,
    get_daily_counts,
    add_alert,
    update_alert,
    delete_watchlist,
    delete_alert,
    get_last_scrape_time,
    get_articles_total_count,
    get_saved_views,
    get_saved_view,
    add_saved_view,
    delete_saved_view,
    get_saved_briefings,
    get_saved_briefing,
    add_saved_briefing,
    update_saved_briefing,
    duplicate_saved_briefing,
    delete_saved_briefing,
    create_user,
    get_user_by_username,
    create_api_key,
    update_user_profile,
    get_annotations_for_article,
    get_annotation,
    add_annotation,
    delete_annotation,
    messaging_create_channel,
    messaging_get_channel_by_id,
    messaging_get_channel_by_slug,
    messaging_get_channels_for_user,
    messaging_is_member,
    messaging_get_member_role,
    messaging_add_member,
    messaging_create_invite,
    messaging_get_invite_by_token,
    messaging_use_invite,
    messaging_get_channel_members,
    messaging_set_member_role,
    messaging_get_messages,
    messaging_get_pinned,
    messaging_get_thread_replies,
    messaging_edit_message,
    messaging_delete_message,
    messaging_pin_message,
    messaging_unpin_message,
    messaging_leave_channel,
    messaging_mute_channel,
    messaging_unmute_channel,
    messaging_is_muted,
    messaging_invite_request_create,
    messaging_invite_request_list,
    messaging_invite_request_resolve,
    messaging_report_message,
    messaging_audit_log,
    messaging_get_audit_log,
    messaging_get_all_channels,
    messaging_set_last_read,
    messaging_get_channels_for_user_with_unread,
    messaging_reaction_add,
    messaging_reaction_remove,
    messaging_get_reactions,
    messaging_channel_pref_set_starred,
    messaging_channel_pref_get,
    messaging_get_reports_for_admin,
    messaging_report_resolve,
    messaging_update_channel,
    messaging_get_reply_counts,
    messaging_get_notifications,
    messaging_notification_mark_read,
    messaging_create_notification,
    messaging_add_mention,
    get_user_by_id,
    get_article,
    MESSAGING_CHANNEL_TYPES,
    get_articles_by_entity,
    get_article_counts_by_country,
    get_spike_topics,
    get_declining_topics,
    get_sources_for_window,
    get_country_risk_snapshots,
    get_risk_index,
    get_treaties,
    get_treaties_count,
    get_treaty,
    get_treaty_counts_by_type,
    get_treaty_distinct_regions,
    get_treaty_distinct_coverages,
    get_treaties_by_year,
    get_agreements_page_stats,
    parse_treaty_summary,
    get_diplomacy_summary,
    get_related_treaties,
    add_treaty,
    update_treaty,
    update_treaty_escalation,
    get_sanctions,
    get_sanctions_total_count,
    get_sanctions_global,
    get_sanctions_global_count,
    get_sanction_global,
    get_sanctions_watch_meta,
    add_sanction,
    get_un_votes,
    get_un_resolutions,
    compute_voting_alignment,
    get_voting_alignment,
    TREATY_TYPES,
    get_chokepoints,
    get_chokepoints_with_geo,
    get_chokepoint,
    get_flows_for_chokepoint,
    get_flows_for_country,
    get_chokepoint_countries,
    run_chokepoint_scenario,
    get_election_calendar,
    get_election_calendar_count,
    get_election,
    add_election,
    update_election,
    delete_election,
    get_election_calendar_last_updated,
    get_election_regions,
    get_approval_ratings,
    get_approval_ratings_count,
    get_approval_rating,
    add_approval_rating,
    update_approval_rating,
    delete_approval_rating,
    get_approval_last_updated,
    approval_duplicate_exists,
    get_approval_timeseries,
    get_approval_latest_by_country,
    get_protest_tracking,
    get_protest_tracking_count,
    get_protest,
    get_protest_trigger_topics,
    add_protest_event,
    update_protest,
    delete_protest,
    get_protest_last_updated,
    get_protest_counts_by_country,
    get_protest_counts_by_trigger,
    get_articles,
    get_currency_stress,
    get_food_inflation_alerts,
    get_youth_unemployment,
    get_social_sentiment,
    get_fragility_overview,
    FRAGILITY_LOW_APPROVAL_THRESHOLD,
    get_defense_spending,
    get_defense_spending_count,
    get_military_exercises,
    get_military_exercises_count,
    get_border_incidents,
    get_border_incidents_count,
    get_military_movement,
    get_military_movement_count,
    get_world_monitor_points,
    get_naval_deployments,
    get_naval_deployments_count,
    get_naval_deployment_heat,
    get_arms_trade,
    get_arms_trade_count,
    get_arms_trade_summary,
    get_arms_trade_by_year,
    get_conflict_summary,
    get_escalation_tree,
    get_escalation_tree_by_region,
    get_conflict_alert_rules,
    add_conflict_alert_rule,
    delete_conflict_alert_rule,
    get_defense_spending_with_yoy,
    get_entity_list_alerts,
    get_export_restrictions,
    check_supply_chain_export_rules,
    get_scenarios,
    get_scenario,
    get_scenario_run,
    get_scenario_runs,
    run_scenario_simulation,
    generate_risk_outlook,
    run_scenario_engine,
    SCENARIO_ENGINE_EVENT_TYPES,
    generate_scenario_engine_export,
    add_scenario_engine_run,
    get_scenario_engine_runs,
    get_scenario_engine_runs_filtered,
    get_scenario_engine_run,
    add_scenario_from_engine_run,
    delete_scenario_engine_run,
    update_scenario_engine_run_name_notes,
    get_integration_countries,
    get_integration_country,
    get_macroeconomic_stress,
    get_macroeconomic_stress_history,
    get_energy_commodity_exposure,
    get_military_capability_snapshot,
    get_trade_flow_partners,
    get_multilateral_participation,
    get_capital_flows,
    get_elite_institutional,
    get_climate_resource_vulnerability,
    get_technology_semiconductor,
    get_conflict_event_imports,
    get_geospatial_infrastructure,
    get_legislative_policy_tracker,
    get_energy_commodity_summary,
    get_geospatial_infrastructure_summary,
    get_technology_semiconductor_summary,
    get_military_capability_summary,
    get_multilateral_summary_by_org,
    get_macroeconomic_stress_alerts,
    get_climate_vulnerability_summary,
    get_elite_institutional_summary,
    get_capital_flows_summary,
    list_users_for_assignment,
    policy_task_create,
    policy_tasks_for_user,
    policy_task_set_status,
    object_comment_add,
    object_comments_list,
    touch_user_entity_visit,
    get_user_entity_visit,
    get_entity_change_log,
    get_desk_terminal_intel_for_country,
)
from app.country_data import ISO3_TO_2

try:
    from app.models import get_airspace_restrictions
except ImportError:
    def get_airspace_restrictions(limit: int = 20):
        from datetime import datetime
        now = datetime.utcnow().strftime("%Y-%m-%d")
        return [
            {"region": "Ukraine", "status": "restricted", "description": "Conflict zone; commercial overflights rerouted.", "lat": 48.5, "lon": 31.2, "source": "ICAO/EASA", "updated_at": now},
            {"region": "Israel / Gaza", "status": "restricted", "description": "Heightened risk; airlines avoiding airspace.", "lat": 31.5, "lon": 34.8, "source": "FAA/ICAO", "updated_at": now},
            {"region": "Russia (western)", "status": "restricted", "description": "Sanctions; many carriers avoiding Russian airspace.", "lat": 55.8, "lon": 37.6, "source": "EASA", "updated_at": now},
            {"region": "Red Sea / Yemen", "status": "advisory", "description": "Drone/missile risk; flights advised to reroute.", "lat": 15.0, "lon": 43.0, "source": "ICAO", "updated_at": now},
            {"region": "Taiwan Strait", "status": "advisory", "description": "Tension; some carriers adjusting routes.", "lat": 24.5, "lon": 119.5, "source": "Regional", "updated_at": now},
        ]

bp = Blueprint("main", __name__)


def _policy_entity_page_context(user_id: Optional[int], entity_type: str, entity_ref: str, content_timestamp: Optional[str] = None):
    """Visit tracking, 'updated since' banner, change log, thread comments; records visit when user_id set."""
    ref = (entity_ref or "").strip()
    change_log = get_entity_change_log(entity_type, ref, limit=10) if ref else []
    thread = object_comments_list(entity_type, ref) if (user_id and ref) else []
    last_visit = get_user_entity_visit(user_id, entity_type, ref) if (user_id and ref) else None
    updated_since = False
    if last_visit and content_timestamp and ref:
        try:
            updated_since = (content_timestamp or "")[:19] > (last_visit or "")[:19]
        except Exception:
            updated_since = False
    if user_id and ref:
        touch_user_entity_visit(user_id, entity_type, ref)
    return {
        "entity_change_log": change_log,
        "object_thread_comments": thread,
        "last_entity_visit": last_visit,
        "content_updated_since_visit": updated_since,
    }


def _parse_sensitivity_tier(raw) -> str:
    t = (raw or "internal").strip().lower()
    return t if t in ("public", "internal", "restricted") else "internal"


def _parse_impact_param(value):
    """Convert ?impact=high|med|low or 0-10 to int. Returns 7 (high), 5 (med), 2 (low), or 0-10."""
    if not value:
        return None
    v = value.strip().lower()
    if v in ("high", "3"):
        return 7
    if v in ("med", "medium", "2"):
        return 5
    if v in ("low", "1"):
        return 2
    try:
        return max(0, min(10, int(value)))
    except (TypeError, ValueError):
        return None


@bp.route("/help")
def help_page():
    """Getting started, navigation, keyboard shortcuts, and glossary."""
    return render_template("help.html")


@bp.route("/workspace/tasks", methods=["GET", "POST"])
@login_required
def workspace_tasks():
    """Policy tasks: assign, due dates, link to entities."""
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()
        if action == "create":
            title = (request.form.get("title") or "").strip()
            body = (request.form.get("body") or "").strip()
            et = (request.form.get("entity_type") or "").strip() or None
            er = (request.form.get("entity_ref") or "").strip() or None
            due = (request.form.get("due_date") or "").strip() or None
            assign_raw = request.form.get("assignee_user_id", type=int)
            assignee = assign_raw if assign_raw else None
            if title:
                policy_task_create(
                    current_user.id,
                    title,
                    body=body,
                    entity_type=et,
                    entity_ref=er,
                    assignee_user_id=assignee,
                    due_date=due,
                )
                flash("Task created.", "success")
            else:
                flash("Title required.", "error")
        return redirect(url_for("main.workspace_tasks"))
    tasks_open = policy_tasks_for_user(current_user.id, include_done=False)
    tasks_all = policy_tasks_for_user(current_user.id, include_done=True)
    assign_users = list_users_for_assignment()
    return render_template(
        "workspace_tasks.html",
        tasks_open=tasks_open,
        tasks_all=tasks_all,
        assign_users=assign_users,
    )


@bp.route("/workspace/tasks/<int:task_id>/status", methods=["POST"])
@login_required
def workspace_task_status(task_id):
    status = (request.form.get("status") or "done").strip()
    if policy_task_set_status(task_id, current_user.id, status):
        flash("Task updated.", "success")
    else:
        flash("Could not update task.", "error")
    return redirect(request.referrer or url_for("main.workspace_tasks"))


@bp.route("/workspace/object-thread", methods=["POST"])
@login_required
def workspace_object_thread_post():
    """Threaded discussion on article / country / digest / briefing (separate from channel chat)."""
    entity_type = (request.form.get("entity_type") or "").strip()
    entity_ref = (request.form.get("entity_ref") or "").strip()
    body = (request.form.get("body") or "").strip()
    parent_id = request.form.get("parent_id", type=int)
    next_url = (request.form.get("next") or "").strip() or url_for("main.situation_room")
    try:
        object_comment_add(current_user.id, entity_type, entity_ref, body, parent_id=parent_id)
        flash("Comment posted.", "success")
    except ValueError as e:
        flash(str(e), "error")
    return redirect(next_url)


@bp.route("/indicators")
def indicators_hub():
    """Redirect to Help #indicators (indicator layers integrated into existing pages)."""
    return redirect(url_for("main.help_page", _anchor="indicators"))


@bp.route("/indicators/legislative")
def indicators_legislative_redirect():
    """Redirect legacy URL to diplomacy/legislative."""
    return redirect(url_for("main.diplomacy_legislative", **request.args))


@bp.route("/live")
def live_streams():
    """Live news streams from broadcasters offering free feeds."""
    from app.live_streams import LIVE_STREAMS
    return render_template("live_streams.html", streams=LIVE_STREAMS)


@bp.route("/login", methods=["GET", "POST"])
def login():
    """Log in with username and password."""
    if current_user.is_authenticated:
        return redirect(url_for("main.situation_room"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if not username or not password:
            flash("Username and password required.", "error")
            return render_template("login.html")
        user_row = get_user_by_username(username)
        if not user_row or not verify_password(user_row, password):
            flash("Invalid username or password.", "error")
            return render_template("login.html")
        if not user_row.get("is_active", 1):
            flash("Account is disabled.", "error")
            return render_template("login.html")
        from app.auth import User
        user = User(user_row["id"], user_row["username"], user_row.get("email"), user_row.get("is_active", 1))
        login_user(user, remember=bool(request.form.get("remember")))
        next_url = request.args.get("next") or request.referrer or url_for("main.situation_room")
        if next_url and next_url.startswith("/") and "//" not in next_url:
            return redirect(next_url)
        return redirect(url_for("main.situation_room"))
    return render_template("login.html")


@bp.route("/logout", methods=["GET", "POST"])
def logout():
    """Log out the current user."""
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("main.situation_room"))


@bp.route("/register", methods=["GET", "POST"])
def register():
    """Create a new account. Email, username, name, title, and organization are required. Sends welcome email if SMTP configured."""
    if current_user.is_authenticated:
        return redirect(url_for("main.situation_room"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().lower()
        password = request.form.get("password") or ""
        email = (request.form.get("email") or "").strip() or None
        name = (request.form.get("name") or "").strip() or None
        title = (request.form.get("title") or "").strip() or None
        organization = (request.form.get("organization") or "").strip() or None
        if not username or not password:
            flash("Username and password are required.", "error")
            return render_template("register.html")
        if not email:
            flash("Email is required.", "error")
            return render_template("register.html")
        if not name:
            flash("Name is required.", "error")
            return render_template("register.html")
        if not title:
            flash("Title is required.", "error")
            return render_template("register.html")
        if not organization:
            flash("Organization is required.", "error")
            return render_template("register.html")
        if len(password) < 8:
            flash("Password must be at least 8 characters.", "error")
            return render_template("register.html")
        if get_user_by_username(username):
            flash("That username is already taken.", "error")
            return render_template("register.html")
        try:
            create_user(username, password, email=email, name=name, title=title, organization=organization)
            from app.email_sender import send_welcome_email
            if send_welcome_email(email, name or username):
                flash("Account created. A welcome email has been sent to your address. You can log in now.", "success")
            else:
                flash("Account created. You can log in now.", "success")
            return redirect(url_for("main.login"))
        except ValueError as e:
            flash(str(e), "error")
            return render_template("register.html")
    return render_template("register.html")


@bp.route("/account", methods=["GET", "POST"])
@login_required
def account():
    """Account: profile (name, title, organization) and create API key (shown once)."""
    api_key_created = None
    if request.method == "POST":
        if request.form.get("action") == "update_profile":
            name = (request.form.get("name") or "").strip()
            title = (request.form.get("title") or "").strip()
            organization = (request.form.get("organization") or "").strip()
            try:
                update_user_profile(current_user.id, name=name or None, title=title or None, organization=organization or None)
                flash("Profile updated.", "success")
            except Exception:
                flash("Failed to update profile.", "error")
            return redirect(url_for("main.account"))
        if request.form.get("action") == "update_policy_prefs":
            st = _parse_sensitivity_tier(request.form.get("default_sensitivity_tier"))
            lr = 1 if request.form.get("default_legal_review") == "1" else 0
            try:
                update_user_profile(current_user.id, default_sensitivity_tier=st, default_legal_review=lr)
                flash("Policy defaults saved.", "success")
            except Exception:
                flash("Failed to save policy defaults.", "error")
            return redirect(url_for("main.account"))
        if request.form.get("action") == "create_api_key":
            name = (request.form.get("key_name") or "").strip() or "Default"
            try:
                key_id, plain_key = create_api_key(current_user.id, name=name)
                api_key_created = plain_key
                flash("API key created. Copy it now — it won't be shown again.", "success")
            except Exception:
                flash("Failed to create API key.", "error")
    user_prefs = get_user_by_id(current_user.id) or {}
    return render_template("account.html", api_key_created=api_key_created, user_prefs=user_prefs)


# --- Intelligence Messaging (private, invite-only, encrypted) ---
@bp.route("/messaging/search")
@login_required
def messaging_search():
    """Global search across all user's channels."""
    from app.messaging import search_messages_global, channel_type_label
    q = (request.args.get("q") or "").strip()
    results = []
    if q:
        secret = current_app.config.get("SECRET_KEY", "")
        results = search_messages_global(current_user.id, secret, q, limit=50)
    return render_template("messaging_search.html", q=q, results=results, channel_type_label=channel_type_label)


@bp.route("/messaging")
@login_required
def messaging_hub():
    """List channels with unread counts. Sort by unread/activity, filter unread only."""
    from app.messaging import channel_type_label
    channels = messaging_get_channels_for_user_with_unread(current_user.id)
    sort = (request.args.get("sort") or "activity").strip()
    unread_only = request.args.get("unread") == "1"
    if unread_only:
        channels = [c for c in channels if (c.get("unread_count") or 0) > 0]
    if sort == "unread":
        channels.sort(key=lambda c: (-(c.get("unread_count") or 0), -(c.get("last_message_id") or 0)))
    else:
        channels.sort(key=lambda c: (-(c.get("last_message_id") or 0), c.get("name", "")))
    for c in channels:
        c["starred"] = messaging_channel_pref_get(c["id"], current_user.id).get("starred", 0)
    notifications = messaging_get_notifications(current_user.id, limit=10)
    all_channels = messaging_get_all_channels(include_archived=False)
    featured_channels = [c for c in all_channels if c.get("featured")]
    return render_template(
        "messaging_hub.html",
        channels=channels,
        channel_type_label=channel_type_label,
        sort=sort,
        unread_only=unread_only,
        notifications=notifications,
        featured_channels=featured_channels,
    )


@bp.route("/messaging/channel/<slug>")
@login_required
def messaging_channel(slug):
    """Channel view: messages and composer. Members only."""
    channel = messaging_get_channel_by_slug(slug)
    if not channel:
        return "Channel not found", 404
    if not messaging_is_member(channel["id"], current_user.id):
        flash("You are not a member of this channel. Use an invite link to join.", "error")
        return redirect(url_for("main.messaging_hub"))
    from app.messaging import get_messages_for_channel, get_thread_replies_for_channel, get_pinned_for_channel, channel_type_label, format_user_display
    secret = current_app.config.get("SECRET_KEY", "")
    search_q = (request.args.get("q") or "").strip()
    before_id = request.args.get("before_id", type=int)
    messages = get_messages_for_channel(channel["id"], secret, limit=50, before_id=before_id)
    if search_q:
        search_lower = search_q.lower()
        messages = [m for m in messages if search_lower in (m.get("content") or "").lower()]
    pinned = get_pinned_for_channel(channel["id"], secret)
    reply_counts = messaging_get_reply_counts(channel["id"])
    thread_replies = {}  # lazy-loaded via API when user expands thread
    members = messaging_get_channel_members(channel["id"])
    for m in members:
        m["author_display"] = format_user_display(m)
    my_role = messaging_get_member_role(channel["id"], current_user.id)
    can_invite = my_role in ("admin", "verified_analyst")
    can_manage_roles = my_role == "admin"
    invite_requests = messaging_invite_request_list(channel["id"], "pending") if can_manage_roles else []
    is_muted = messaging_is_muted(channel["id"], current_user.id)
    verified_members = [m for m in members if m.get("role") == "verified_analyst"]
    from_user = request.args.get("from_user", type=int)
    if from_user:
        messages = [m for m in messages if m.get("user_id") == from_user]
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    if date_from or date_to:
        def in_range(m):
            created = (m.get("created_at") or "")[:10]
            if date_from and created < date_from:
                return False
            if date_to and created > date_to:
                return False
            return True
        messages = [m for m in messages if in_range(m)]
    max_msg_id = max([m["id"] for m in messages], default=None)
    if max_msg_id:
        messaging_set_last_read(channel["id"], current_user.id, max_msg_id)
    jump_to_message_id = request.args.get("message_id", type=int)
    star_pref = messaging_channel_pref_get(channel["id"], current_user.id)
    search_date_from = request.args.get("date_from") or ""
    search_date_to = request.args.get("date_to") or ""
    search_from_user = request.args.get("from_user", type=int)
    message_reactions = {}
    for m in messages:
        if m.get("id"):
            message_reactions[m["id"]] = messaging_get_reactions(m["id"])
    if pinned and pinned.get("id"):
        message_reactions[pinned["id"]] = messaging_get_reactions(pinned["id"])
    return render_template(
        "messaging_channel.html",
        channel=channel,
        messages=messages,
        thread_replies=thread_replies,
        reply_counts=reply_counts,
        pinned=pinned,
        members=members,
        verified_members=verified_members,
        can_invite=can_invite,
        can_manage_roles=can_manage_roles,
        invite_requests=invite_requests,
        is_muted=is_muted,
        search_q=search_q,
        channel_type_label=channel_type_label,
        jump_to_message_id=jump_to_message_id,
        starred=star_pref.get("starred", 0),
        message_reactions=message_reactions,
        search_date_from=search_date_from,
        search_date_to=search_date_to,
        search_from_user=search_from_user,
    )


@bp.route("/messaging/channel/<int:channel_id>/message", methods=["POST"])
@login_required
def messaging_send_message(channel_id):
    """Post a message (encrypted at rest). Members only. Rate limited."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel or not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    from app.models import _connection
    with _connection() as conn:
        cur = conn.execute(
            "SELECT COUNT(*) FROM messaging_messages WHERE user_id = ? AND created_at > datetime('now', '-1 minute')",
            (current_user.id,),
        )
        if cur.fetchone()[0] >= 30:
            flash("Rate limit: max 30 messages per minute. Please wait.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    content = (request.form.get("content") or "").strip()
    if not content or len(content) > 10000:
        flash("Message is required and must be under 10,000 characters.", "error")
        return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    from app.messaging import add_message_encrypted, fire_channel_webhook, format_user_display
    secret = current_app.config.get("SECRET_KEY", "")
    parent_id = request.form.get("parent_id", type=int)
    attachment_type = (request.form.get("attachment_type") or "").strip() or None
    attachment_id = request.form.get("attachment_id", type=int)
    attachment_extra = (request.form.get("attachment_extra") or "").strip() or None
    if attachment_type == "risk":
        attachment_id = 0
        attachment_extra = None
    elif attachment_type == "country":
        attachment_extra = (request.form.get("attachment_country_code") or "").strip().upper()[:3] or None
        attachment_id = None
        if not attachment_extra:
            attachment_type = None
    if attachment_type and attachment_type not in ("article", "digest", "watchlist", "alert", "briefing", "country", "risk", "saved_view"):
        attachment_type = None
        attachment_id = None
        attachment_extra = None
    if attachment_type in ("article", "digest", "watchlist", "alert", "briefing", "saved_view") and not attachment_id:
        attachment_type = None
        attachment_extra = None
    if attachment_type == "article" and attachment_id:
        if not get_article(attachment_id):
            flash("Article not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if attachment_type == "digest" and attachment_id:
        if not get_digest(attachment_id):
            flash("Digest not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if attachment_type == "watchlist" and attachment_id:
        if not get_watchlist(attachment_id):
            flash("Watchlist not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if attachment_type == "alert" and attachment_id:
        if not get_alert(attachment_id, user_id=current_user.id):
            flash("Alert not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if attachment_type == "briefing" and attachment_id:
        if not get_saved_briefing(attachment_id, user_id=current_user.id):
            flash("Saved briefing not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if attachment_type == "saved_view" and attachment_id:
        if not get_saved_view(attachment_id, user_id=current_user.id):
            flash("Saved view not found.", "error")
            return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    msg_id = add_message_encrypted(channel_id, current_user.id, content, secret, parent_id=parent_id, attachment_type=attachment_type, attachment_id=attachment_id or None, attachment_extra=attachment_extra)
    user = get_user_by_id(current_user.id)
    messaging_audit_log(channel_id, current_user.id, "message_sent", {"message_id": msg_id})
    fire_channel_webhook(channel_id, msg_id, format_user_display(user) or user.get("username", ""), content)
    import re
    for m in re.findall(r"@(\w+)", content):
        u = get_user_by_username(m)
        if u and u.get("id") != current_user.id:
            messaging_add_mention(msg_id, channel_id, u["id"])
            messaging_create_notification(u["id"], channel_id, msg_id, "mention", from_user_id=current_user.id)
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/forward", methods=["POST"])
@login_required
def messaging_forward():
    """Forward an attachment to a channel (from article/digest/country/etc pages). Form: channel_id, attachment_type, attachment_id?, attachment_country_code?, content (optional)."""
    channel_id = request.form.get("channel_id", type=int)
    if not channel_id:
        flash("Select a channel.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    channel = messaging_get_channel_by_id(channel_id)
    if not channel or not messaging_is_member(channel_id, current_user.id):
        flash("Channel not found or you are not a member.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    from app.messaging import add_message_encrypted, fire_channel_webhook, format_user_display
    secret = current_app.config.get("SECRET_KEY", "")
    content = (request.form.get("content") or "").strip() or "—"
    if len(content) > 10000:
        content = content[:10000]
    attachment_type = (request.form.get("attachment_type") or "").strip() or None
    attachment_id = request.form.get("attachment_id", type=int)
    attachment_extra = (request.form.get("attachment_country_code") or "").strip().upper()[:3] or None
    if attachment_type == "risk":
        attachment_id = 0
        attachment_extra = None
    elif attachment_type == "country":
        attachment_id = None
        if not attachment_extra:
            attachment_type = None
    if attachment_type and attachment_type not in ("article", "digest", "watchlist", "alert", "briefing", "country", "risk", "saved_view"):
        attachment_type = None
        attachment_id = None
        attachment_extra = None
    if attachment_type in ("article", "digest", "watchlist", "alert", "briefing", "saved_view") and not attachment_id:
        attachment_type = None
        attachment_extra = None
    if attachment_type == "article" and attachment_id and not get_article(attachment_id):
        flash("Article not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    if attachment_type == "digest" and attachment_id and not get_digest(attachment_id):
        flash("Digest not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    if attachment_type == "watchlist" and attachment_id and not get_watchlist(attachment_id):
        flash("Watchlist not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    if attachment_type == "alert" and attachment_id and not get_alert(attachment_id, user_id=current_user.id):
        flash("Alert not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    if attachment_type == "briefing" and attachment_id and not get_saved_briefing(attachment_id, user_id=current_user.id):
        flash("Saved briefing not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    if attachment_type == "saved_view" and attachment_id and not get_saved_view(attachment_id, user_id=current_user.id):
        flash("Saved view not found.", "error")
        return redirect(request.referrer or url_for("main.messaging_hub"))
    msg_id = add_message_encrypted(channel_id, current_user.id, content, secret, parent_id=None, attachment_type=attachment_type, attachment_id=attachment_id or None, attachment_extra=attachment_extra)
    user = get_user_by_id(current_user.id)
    messaging_audit_log(channel_id, current_user.id, "message_sent", {"message_id": msg_id})
    fire_channel_webhook(channel_id, msg_id, format_user_display(user) or user.get("username", ""), content)
    flash("Forwarded to channel.", "success")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/create", methods=["GET", "POST"])
@login_required
def messaging_create_channel_route():
    """Create a new channel. Creator becomes admin. Invite-only by default."""
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        slug = (request.form.get("slug") or "").strip().lower().replace(" ", "-")
        channel_type = request.form.get("channel_type") or "thematic"
        description = (request.form.get("description") or "").strip()
        webhook_url = (request.form.get("webhook_url") or "").strip() or None
        invite_only = request.form.get("invite_only") != "0"
        if not name or not slug:
            flash("Name and slug are required.", "error")
            return redirect(url_for("main.messaging_create_channel_route"))
        if channel_type not in MESSAGING_CHANNEL_TYPES:
            channel_type = "thematic"
        existing = messaging_get_channel_by_slug(slug)
        if existing:
            flash("That slug is already taken.", "error")
            return redirect(url_for("main.messaging_create_channel_route"))
        try:
            cid = messaging_create_channel(name, slug, channel_type, description=description, invite_only=invite_only, created_by_user_id=current_user.id, webhook_url=webhook_url)
            messaging_add_member(cid, current_user.id, role="admin")
            flash("Channel created. You can now invite others via the channel page.", "success")
            return redirect(url_for("main.messaging_channel", slug=slug))
        except Exception:
            flash("Failed to create channel.", "error")
            return redirect(url_for("main.messaging_create_channel_route"))
    from app.messaging import channel_type_label
    return render_template("messaging_create_channel.html", channel_types=MESSAGING_CHANNEL_TYPES, channel_type_label=channel_type_label)


@bp.route("/messaging/channel/<int:channel_id>/invite", methods=["POST"])
@login_required
def messaging_create_invite_route(channel_id):
    """Create an invite link. Admin/verified_analyst only."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    role = messaging_get_member_role(channel_id, current_user.id)
    if role not in ("admin", "verified_analyst"):
        return "Forbidden", 403
    token = messaging_create_invite(channel_id, current_user.id)
    if not token:
        flash("Failed to create invite.", "error")
    else:
        invite_url = request.host_url.rstrip("/") + url_for("main.messaging_accept_invite", token=token)
        flash(f"Invite created. Share this link (valid 7 days): {invite_url}", "success")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/member/<int:user_id>/role", methods=["POST"])
@login_required
def messaging_set_member_role_route(channel_id, user_id):
    """Set a member's role (admin/verified_analyst only)."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    my_role = messaging_get_member_role(channel_id, current_user.id)
    if my_role != "admin":
        return "Forbidden", 403
    role = (request.form.get("role") or "").strip()
    if role not in ("member", "admin", "verified_analyst"):
        flash("Invalid role.", "error")
        return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if messaging_set_member_role(channel_id, user_id, role):
        flash("Role updated.", "success")
    else:
        flash("Could not update role.", "error")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/invite/<token>")
@login_required
def messaging_accept_invite(token):
    """Accept an invite: add user to channel and redirect to channel."""
    inv = messaging_get_invite_by_token(token)
    if not inv:
        flash("Invalid or expired invite.", "error")
        return redirect(url_for("main.messaging_hub"))
    if messaging_use_invite(token, current_user.id):
        channel = messaging_get_channel_by_id(inv["channel_id"])
        messaging_audit_log(inv["channel_id"], current_user.id, "member_joined", {"via": "invite"})
        flash("You have joined the channel.", "success")
        return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    flash("Could not join (already a member or invalid invite).", "error")
    return redirect(url_for("main.messaging_hub"))


@bp.route("/messaging/directory")
@login_required
def messaging_directory():
    """Channel directory: list all channels with Request invite."""
    from app.messaging import channel_type_label
    all_channels = messaging_get_all_channels()
    my_slugs = {c["slug"] for c in messaging_get_channels_for_user(current_user.id)}
    return render_template("messaging_directory.html", channels=all_channels, my_slugs=my_slugs, channel_type_label=channel_type_label)


@bp.route("/messaging/channel/<int:channel_id>/request-invite", methods=["POST"])
@login_required
def messaging_request_invite(channel_id):
    """Request access to a channel (invite-only)."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    if messaging_is_member(channel_id, current_user.id):
        flash("You are already a member.", "info")
        return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    if messaging_invite_request_create(channel_id, current_user.id):
        flash("Request sent. An admin will review it.", "success")
    else:
        flash("You already have a pending request.", "info")
    return redirect(url_for("main.messaging_directory"))


@bp.route("/messaging/channel/<int:channel_id>/invite-user", methods=["POST"])
@login_required
def messaging_invite_by_username(channel_id):
    """Create invite and show link; optionally for a specific username."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    role = messaging_get_member_role(channel_id, current_user.id)
    if role not in ("admin", "verified_analyst"):
        return "Forbidden", 403
    username = (request.form.get("username") or "").strip().lower()
    token = messaging_create_invite(channel_id, current_user.id, invited_email=username or None)
    invite_url = request.host_url.rstrip("/") + url_for("main.messaging_accept_invite", token=token)
    flash(f"Invite link created. Share: {invite_url}", "success")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/leave", methods=["POST"])
@login_required
def messaging_leave(channel_id):
    """Leave the channel."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    messaging_leave_channel(channel_id, current_user.id)
    messaging_audit_log(channel_id, current_user.id, "member_left", {})
    flash("You left the channel.", "info")
    return redirect(url_for("main.messaging_hub"))


@bp.route("/messaging/channel/<int:channel_id>/mute", methods=["POST"])
@login_required
def messaging_mute(channel_id):
    messaging_mute_channel(channel_id, current_user.id)
    flash("Channel muted.", "info")
    return redirect(request.referrer or url_for("main.messaging_hub"))


@bp.route("/messaging/channel/<int:channel_id>/unmute", methods=["POST"])
@login_required
def messaging_unmute(channel_id):
    messaging_unmute_channel(channel_id, current_user.id)
    flash("Channel unmuted.", "info")
    return redirect(request.referrer or url_for("main.messaging_hub"))


@bp.route("/messaging/channel/<int:channel_id>/message/<int:message_id>/report", methods=["POST"])
@login_required
def messaging_report(channel_id, message_id):
    channel = messaging_get_channel_by_id(channel_id)
    if not channel or not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    reason = (request.form.get("reason") or "").strip()
    if messaging_report_message(message_id, channel_id, current_user.id, reason=reason):
        flash("Report submitted. Channel admins will be notified.", "success")
    else:
        flash("Could not submit report.", "error")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/message/<int:message_id>/pin", methods=["POST"])
@login_required
def messaging_pin(channel_id, message_id):
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    if messaging_pin_message(message_id, channel_id, current_user.id):
        messaging_audit_log(channel_id, current_user.id, "message_pinned", {"message_id": message_id})
        flash("Message pinned.", "success")
    else:
        flash("Only admins can pin.", "error")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/unpin", methods=["POST"])
@login_required
def messaging_unpin(channel_id):
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    if messaging_unpin_message(channel_id, current_user.id):
        flash("Pinned message removed.", "info")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/message/<int:message_id>/edit", methods=["GET", "POST"])
@login_required
def messaging_edit(channel_id, message_id):
    channel = messaging_get_channel_by_id(channel_id)
    if not channel or not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    from app.models import messaging_get_message_by_id
    from app.messaging import decrypt_message
    msg = messaging_get_message_by_id(message_id)
    if not msg or msg["channel_id"] != channel_id or msg["user_id"] != current_user.id:
        return "Forbidden", 403
    if request.method == "POST":
        content = (request.form.get("content") or "").strip()
        if content and len(content) <= 10000:
            from app.messaging import encrypt_message
            secret = current_app.config.get("SECRET_KEY", "")
            if messaging_edit_message(message_id, current_user.id, encrypt_message(content, secret)):
                flash("Message updated.", "success")
        return redirect(url_for("main.messaging_channel", slug=channel["slug"]))
    secret = current_app.config.get("SECRET_KEY", "")
    msg["content"] = decrypt_message(msg.get("content_encrypted") or b"", secret)
    return render_template("messaging_edit_message.html", channel=channel, message=msg)


@bp.route("/messaging/channel/<int:channel_id>/message/<int:message_id>/delete", methods=["POST"])
@login_required
def messaging_delete(channel_id, message_id):
    channel = messaging_get_channel_by_id(channel_id)
    if not channel or not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    if messaging_delete_message(message_id, current_user.id, soft=True):
        messaging_audit_log(channel_id, current_user.id, "message_deleted", {"message_id": message_id})
        flash("Message deleted.", "info")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<slug>/audit")
@login_required
def messaging_audit(slug):
    """Audit log for channel (admin only)."""
    channel = messaging_get_channel_by_slug(slug)
    if not channel or not messaging_is_member(channel["id"], current_user.id):
        return "Forbidden", 403
    my_role = messaging_get_member_role(channel["id"], current_user.id)
    if my_role != "admin":
        return "Forbidden", 403
    log = messaging_get_audit_log(channel["id"], limit=100)
    from app.messaging import format_user_display
    for e in log:
        e["author_display"] = format_user_display(e)
    return render_template("messaging_audit.html", channel=channel, log=log)


@bp.route("/messaging/channel/<int:channel_id>/invite-requests", methods=["POST"])
@login_required
def messaging_resolve_invite_request(channel_id):
    """Resolve a pending invite request (grant/deny). Admin only."""
    channel = messaging_get_channel_by_id(channel_id)
    if not channel:
        return "Channel not found", 404
    if messaging_get_member_role(channel_id, current_user.id) != "admin":
        return "Forbidden", 403
    request_id = request.form.get("request_id", type=int)
    grant = request.form.get("grant") == "1"
    if request_id and messaging_invite_request_resolve(request_id, channel_id, current_user.id, grant=grant):
        messaging_audit_log(channel_id, current_user.id, "invite_request_resolved", {"request_id": request_id, "granted": grant})
        flash("Request " + ("granted" if grant else "denied") + ".", "success")
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/channel/<int:channel_id>/mark-read", methods=["POST"])
@login_required
def messaging_mark_read(channel_id):
    """Mark channel as read up to latest message or given message_id."""
    if not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    message_id = request.form.get("message_id", type=int)
    messaging_set_last_read(channel_id, current_user.id, last_read_message_id=message_id)
    return redirect(request.referrer or url_for("main.messaging_hub"))


@bp.route("/messaging/channel/<int:channel_id>/star", methods=["POST"])
@login_required
def messaging_star(channel_id):
    """Toggle starred for channel."""
    if not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    starred = request.form.get("starred") == "1"
    messaging_channel_pref_set_starred(channel_id, current_user.id, starred)
    return redirect(request.referrer or url_for("main.messaging_hub"))


@bp.route("/messaging/channel/<int:channel_id>/message/<int:message_id>/reaction", methods=["POST"])
@login_required
def messaging_reaction(channel_id, message_id):
    """Add or remove reaction. Form: emoji=👍&action=add|remove."""
    if not messaging_is_member(channel_id, current_user.id):
        return "Forbidden", 403
    emoji = (request.form.get("emoji") or "👍").strip()[:32]
    action = request.form.get("action") or "add"
    if action == "remove":
        messaging_reaction_remove(message_id, current_user.id, emoji)
    else:
        messaging_reaction_add(message_id, current_user.id, emoji)
    channel = messaging_get_channel_by_id(channel_id)
    return redirect(url_for("main.messaging_channel", slug=channel["slug"]))


@bp.route("/messaging/settings/<slug>", methods=["GET", "POST"])
@login_required
def messaging_channel_settings(slug):
    """Channel settings (name, description, webhook, archived, featured). Admin only."""
    channel = messaging_get_channel_by_slug(slug)
    if not channel or not messaging_is_member(channel["id"], current_user.id):
        return "Forbidden", 403
    if messaging_get_member_role(channel["id"], current_user.id) != "admin":
        return "Forbidden", 403
    if request.method == "POST":
        name = (request.form.get("name") or "").strip() or None
        description = (request.form.get("description") or "").strip() or None
        webhook_url = (request.form.get("webhook_url") or "").strip() or None
        archived = request.form.get("archived") == "1"
        featured = request.form.get("featured") == "1"
        messaging_update_channel(channel["id"], current_user.id, name=name, description=description, webhook_url=webhook_url, archived=archived, featured=featured)
        flash("Channel settings updated.", "success")
        return redirect(url_for("main.messaging_channel", slug=slug))
    return render_template("messaging_channel_settings.html", channel=channel)


@bp.route("/messaging/reports")
@login_required
def messaging_reports_inbox():
    """Report inbox for admins: list reports in channels they admin."""
    reports = messaging_get_reports_for_admin(current_user.id)
    from app.messaging import format_user_display
    for r in reports:
        r["reporter_display"] = format_user_display({
            "username": r.get("reporter_username"),
            "name": r.get("reporter_name"),
            "title": r.get("reporter_title"),
            "organization": r.get("reporter_organization"),
        })
    return render_template("messaging_reports.html", reports=reports)


@bp.route("/messaging/reports/<int:report_id>/resolve", methods=["POST"])
@login_required
def messaging_report_resolve_route(report_id):
    delete_message = request.form.get("delete_message") == "1"
    if messaging_report_resolve(report_id, current_user.id, delete_message=delete_message):
        flash("Report resolved.", "success")
    else:
        flash("Could not resolve report.", "error")
    return redirect(url_for("main.messaging_reports_inbox"))


@bp.route("/messaging/notifications")
@login_required
def messaging_notifications_page():
    """List notifications and mark as read."""
    notifications = messaging_get_notifications(current_user.id, limit=50, unread_only=False)
    return render_template("messaging_notifications.html", notifications=notifications)


@bp.route("/messaging/notifications/<int:nid>/read", methods=["POST"])
@login_required
def messaging_notification_mark_read_route(nid):
    messaging_notification_mark_read(nid, current_user.id)
    return redirect(request.referrer or url_for("main.messaging_notifications_page"))


@bp.route("/")
def situation_room():
    """Global Situation Room: heatmap + breaking alerts + escalation + top risk movers + headlines + quick stats."""
    # Time window: 24h (1) or 7d (7)
    window = request.args.get("window", "24h").strip().lower()
    if window not in ("24h", "7d"):
        window = "24h"
    days = 7 if window == "7d" else 1

    heat_map = get_country_risk_snapshots()
    risk_index_list = get_risk_index()
    integration_countries = get_integration_countries(limit=200)
    uid = get_effective_user_id()
    alerts = get_alerts(user_id=uid)
    alert_counts = get_alert_match_counts(days=days, user_id=uid)
    alert_count_by_id = {c["alert_id"]: c["count"] for c in alert_counts}
    # Sort alerts by match count (hottest first)
    alerts_sorted = sorted(alerts, key=lambda a: -(alert_count_by_id.get(a["id"]) or 0)) if alerts else []

    escalation_treaties = get_treaties(escalation_only=True, limit=10)
    top_5_risk = (heat_map or [])[:5]
    last_scrape = get_last_scrape_time()

    # Headline intel: high-signal articles from last N days
    headline_candidates = get_articles(limit=30, days=min(7, days + 2))
    for a in headline_candidates:
        a["signal_score"] = round(article_signal_score(a), 1)
        a["confidence_score"] = article_confidence_score(a)
    headline_candidates.sort(key=lambda x: (-(x.get("signal_score") or 0), -(x.get("impact_score") or 0)))
    headline_articles = headline_candidates[:5]

    from datetime import datetime as _dt
    _today = _dt.utcnow().strftime("%Y-%m")
    upcoming_elections = get_election_calendar(date_from=_today, status="upcoming", limit=3)
    if not upcoming_elections:
        upcoming_elections = get_election_calendar(date_from=_today, limit=3)

    # Recent incidents (border + military movement)
    border_incidents = get_border_incidents(limit=5)
    military_movements = get_military_movement(limit=3)
    recent_incidents = []
    for b in border_incidents or []:
        recent_incidents.append({
            "type": "border",
            "date": (b.get("incident_date") or "")[:10],
            "summary": b.get("summary") or "Border incident",
            "link": None,
            "label": f"{b.get('country_a_code') or ''}–{b.get('country_b_code') or ''}",
        })
    for m in military_movements or []:
        recent_incidents.append({
            "type": "movement",
            "date": (m.get("observed_date") or m.get("created_at") or "")[:10],
            "summary": (m.get("summary") or m.get("detection_type") or "Military movement")[:80],
            "link": None,
            "label": m.get("country_code") or m.get("region") or "",
        })
    recent_incidents.sort(key=lambda x: x["date"] or "", reverse=True)
    recent_incidents = recent_incidents[:5]

    # Highest-risk region (aggregate heat_map by region via integration_countries)
    code2_to_region = {}
    for c in integration_countries or []:
        code2 = ISO3_TO_2.get((c.get("country_code") or "").upper())
        if code2:
            code2_to_region[code2] = c.get("region") or "Unknown"
    region_scores = defaultdict(float)
    for r in heat_map or []:
        code2 = (r.get("country_code") or "").upper()
        if len(code2) == 2:
            reg = code2_to_region.get(code2)
            if reg:
                region_scores[reg] += float(r.get("risk_score") or 0)
    highest_risk_region = max(region_scores, key=region_scores.get) if region_scores else None

    # GEPI (Escalation Pressure Index) - institutional model
    gepi_latest = None
    try:
        from app.institutional_models.readers import get_gepi_latest
        gepi_latest = get_gepi_latest()
    except Exception:
        pass

    # UN vote analytics: polarization, shocks, bloc cohesion
    un_polarization = []
    un_shocks = []
    un_bloc_cohesion = []
    try:
        from app.un_votes.readers import get_global_polarization, get_alignment_shocks, get_bloc_cohesion
        un_polarization = get_global_polarization(limit=24)
        un_shocks = get_alignment_shocks(limit=10, shock_only=True)
        un_bloc_cohesion = get_bloc_cohesion(limit=12)
    except Exception:
        pass

    # Sanctions: count added in last 7 days (created_at)
    sanctions_all = get_sanctions(limit=200)
    cutoff_7d = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
    new_sanctions_7d = sum(1 for s in sanctions_all or [] if (s.get("created_at") or "")[:10] >= cutoff_7d)

    # Quick stats
    high_risk_count = len([r for r in (heat_map or []) if (r.get("risk_score") or 0) >= 60])
    alerts_firing = sum(alert_count_by_id.values()) or 0
    daily = get_daily_counts(days=days)
    articles_24h = sum(d["count"] for d in daily) if daily else 0

    # Data freshness: stale if last scrape > 24h ago
    hours_since_scrape = None
    if last_scrape:
        try:
            # last_scrape is ISO string e.g. 2025-02-09T12:00:00Z — parse as naive UTC
            s = last_scrape.replace("Z", "").strip()
            if "T" in s:
                scraped = datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
            else:
                scraped = datetime.strptime(s[:10], "%Y-%m-%d")
            hours_since_scrape = (datetime.utcnow() - scraped).total_seconds() / 3600
        except Exception:
            pass
    data_stale = hours_since_scrape is not None and hours_since_scrape > 24
    system_status = "Data refresh delayed" if data_stale else "All systems nominal"

    # Synthesis: trajectory + divergences (lightweight, no LLM for fallback narrative)
    trajectory_data = None
    divergences = []
    try:
        from app.synthesis import get_escalation_trajectory, get_signal_divergences
        trajectory_data = get_escalation_trajectory(days=30)
        divergences = get_signal_divergences()
    except Exception:
        pass

    # World Monitor (integrated): load a comprehensive default marker set for the home map overlay
    wm_points = []
    wm_last_updated = last_scrape
    try:
        from app.world_monitor_config import (
            get_world_monitor_time_days,
        )
        wm_time_days = get_world_monitor_time_days("7d")
        wm_layers = [
            "iranAttacks", "hotspots", "conflicts", "bases", "nuclear",
            "gamma_irradiators", "spaceports", "undersea_cables", "pipelines", "ai_datacenters",
            "military", "waterways", "sanctions", "natural", "weather", "economic",
        ]
        wm_points = get_world_monitor_points(time_range_days=wm_time_days, layers=wm_layers)
    except Exception:
        pass

    try:
        from app.live_streams import LIVE_STREAMS
        live_streams = LIVE_STREAMS
    except Exception:
        live_streams = []

    return render_template(
        "situation_room.html",
        heat_map=heat_map,
        risk_index=risk_index_list,
        live_streams=live_streams,
        wm_points=wm_points,
        wm_last_updated=wm_last_updated,
        integration_countries=integration_countries,
        alerts=alerts_sorted,
        alert_count_by_id=alert_count_by_id,
        escalation_treaties=escalation_treaties,
        top_5_risk=top_5_risk,
        last_scrape_time=last_scrape,
        headline_articles=headline_articles,
        upcoming_elections=upcoming_elections,
        high_risk_count=high_risk_count,
        alerts_firing=alerts_firing,
        articles_24h=articles_24h,
        recent_incidents=recent_incidents,
        highest_risk_region=highest_risk_region,
        new_sanctions_7d=new_sanctions_7d,
        data_stale=data_stale,
        hours_since_scrape=hours_since_scrape,
        system_status=system_status,
        stats_days=days,
        window=window,
        gepi_latest=gepi_latest,
        un_polarization=un_polarization,
        un_shocks=un_shocks,
        un_bloc_cohesion=un_bloc_cohesion,
        trajectory_data=trajectory_data,
        divergences=divergences or [],
    )


def _article_matches_filters(a, days, date_from, date_to, country, countries_list, risk_category, risk_categories_list, impact):
    """Filter a single article by date/country/risk/impact for in-feed search."""
    pub = (a.get("published_utc") or a.get("scraped_at") or "")[:10]
    if date_from and pub < date_from:
        return False
    if date_to and pub > date_to:
        return False
    if days and not date_from and not date_to:
        since = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")
        if pub < since:
            return False
    if country or (countries_list and len(countries_list) > 0):
        pats = [country] if country else list(countries_list)
        combined = " ".join([
            str(a.get("entities") or ""),
            str(a.get("topics") or ""),
            str(a.get("title") or ""),
            str(a.get("summary") or ""),
        ]).lower()
        if not any(p.lower() in combined for p in pats):
            return False
    if risk_category or (risk_categories_list and len(risk_categories_list) > 0):
        et = (a.get("event_type") or "").lower()
        doms = str(a.get("impact_domains") or "").lower()
        checks = [risk_category] if risk_category else list(risk_categories_list)
        if not any(c.lower() in et or c.lower() in doms for c in checks):
            return False
    if impact is not None and (a.get("impact_score") or 0) < impact:
        return False
    return True


@bp.route("/feed")
def index():
    """Intelligence feed: structured intelligence with country/risk filters, confidence score, signal vs noise. Supports pagination, hidden articles, RSS, export."""
    source = request.args.get("source", "").strip() or None
    topic = request.args.get("topic", "").strip() or None
    topics_list = [t.strip() for t in request.args.getlist("topics") if t and t.strip()]
    domain = request.args.get("domain", "").strip() or None
    country = request.args.get("country", "").strip() or None
    countries_list = [c.strip() for c in request.args.getlist("countries") if c and c.strip()]
    risk_category = request.args.get("risk_category", "").strip() or None
    risk_categories_list = [r.strip() for r in request.args.getlist("risk_categories") if r and r.strip()]
    impact = _parse_impact_param(request.args.get("impact"))
    sort = (request.args.get("sort") or "date").strip().lower()
    if sort not in ("date", "signal"):
        sort = "date"
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "3d": 3, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    if date_from and len(date_from) < 10:
        date_from = None
    if date_to and len(date_to) < 10:
        date_to = None
    watchlist_id = request.args.get("watchlist_id", "").strip() or None
    if watchlist_id:
        try:
            wl = get_watchlist(int(watchlist_id))
            if wl and wl.get("topics"):
                topics_list = topics_list or (wl["topics"] if isinstance(wl["topics"], list) else [])
        except (ValueError, TypeError):
            watchlist_id = None
    if not topics_list and topic:
        topics_list = [topic]
        topic = None
    if not countries_list and country:
        countries_list = [country]
        country = None
    if not risk_categories_list and risk_category:
        risk_categories_list = [risk_category]
        risk_category = None
    show_hidden = request.args.get("show_hidden", "").strip() == "1"
    per_page = min(max(10, int(request.args.get("per_page", 25) or 25)), 100)
    page = max(1, int(request.args.get("page", 1) or 1))
    offset = (page - 1) * per_page
    q = request.args.get("q", "").strip() or None
    layout = request.args.get("layout", "card").strip().lower()
    if layout not in ("card", "compact", "grid"):
        layout = "card"

    hidden_ids = list(session.get("feed_hidden_ids") or [])
    hidden_set = set(hidden_ids)
    exclude_ids = None if show_hidden else (hidden_ids if hidden_ids else None)

    if q:
        search_results = search_articles(q, limit=800)
        articles = []
        for a in search_results:
            if exclude_ids and a.get("id") in hidden_set:
                continue
            if not _article_matches_filters(a, days, date_from, date_to, country, countries_list, risk_category, risk_categories_list, impact):
                continue
            if topic:
                if not (a.get("topics_list") or []) and a.get("topics"):
                    try:
                        a["topics_list"] = json.loads(a["topics"]) if isinstance(a["topics"], str) else a["topics"]
                    except (json.JSONDecodeError, TypeError):
                        a["topics_list"] = []
                if topic not in (a.get("topics_list") or []):
                    continue
            if topics_list:
                if not (a.get("topics_list") or []) and a.get("topics"):
                    try:
                        a["topics_list"] = json.loads(a["topics"]) if isinstance(a["topics"], str) else a["topics"]
                    except (json.JSONDecodeError, TypeError):
                        a["topics_list"] = []
                if not any(t in (a.get("topics_list") or []) for t in topics_list):
                    continue
            if source and (a.get("source_name") or "") != source:
                continue
            if domain and not (a.get("impact_domains") or domain not in str(a.get("impact_domains") or "")):
                continue
            articles.append(a)
        total_count = len(articles)
        for a in articles:
            a["confidence_score"] = article_confidence_score(a)
            a["signal_score"] = round(article_signal_score(a), 1)
        if sort == "signal":
            articles.sort(key=lambda x: -x["signal_score"])
        else:
            articles.sort(key=lambda x: (x.get("published_utc") or x.get("scraped_at") or ""), reverse=True)
        articles = articles[offset : offset + per_page]
    else:
        total_count = get_articles_count(
            source=source, topic=topic, topics_list=topics_list or None,
            domain=domain, min_impact=impact, days=days,
            date_from=date_from, date_to=date_to,
            country=country, countries_list=countries_list or None,
            risk_category=risk_category, risk_categories_list=risk_categories_list or None,
            exclude_ids=exclude_ids,
        )
        articles = get_articles(
            limit=per_page,
            offset=offset,
            source=source,
            topic=topic,
            topics_list=topics_list or None,
            domain=domain,
            min_impact=impact,
            days=days,
            date_from=date_from,
            date_to=date_to,
            country=country,
            countries_list=countries_list or None,
            risk_category=risk_category,
            risk_categories_list=risk_categories_list or None,
            exclude_ids=exclude_ids,
        )
        for a in articles:
            a["confidence_score"] = article_confidence_score(a)
            a["signal_score"] = round(article_signal_score(a), 1)
        if sort == "signal":
            articles.sort(key=lambda x: -x["signal_score"])

    total_pages = max(1, (total_count + per_page - 1) // per_page)
    showing_start = offset + 1 if total_count else 0
    showing_end = min(offset + per_page, total_count)

    sources = get_sources()
    topics = get_all_topic_counts()
    impact_summary = get_impact_summary(days=days or 7)
    domain_counts = get_domain_counts(days=days or 7, limit=8)
    event_types = get_feed_event_types(days=days or 7)
    event_type_counts = get_feed_event_type_counts(days=days or 7, limit=10)
    country_options = get_feed_country_options(days=days or 7, limit=40)
    uid = get_effective_user_id()
    alerts = get_alerts(user_id=uid)
    alert_counts = get_alert_match_counts(days=days or 7, user_id=uid)
    alert_count_by_id = {c["alert_id"]: c["count"] for c in alert_counts}
    last_scrape = get_last_scrape_time()
    saved_views = get_saved_views(user_id=uid)
    total_in_db = get_articles_total_count()
    topic_counts_sidebar = get_trending_topics(days=days or 7, limit=15)
    clusters = get_clusters_with_counts(limit=20)
    cluster_count_by_id = {c["cluster_id"]: c["count"] for c in get_clusters_with_counts(limit=300)}
    watchlists = get_watchlists() or []

    daily_counts = get_daily_counts(days=3)
    daily_counts_sparkline = get_daily_counts(days=7)
    daily_sparkline_max = max([d.get("count", 0) for d in daily_counts_sparkline] or [1])
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    yesterday_str = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    today_count = next((d["count"] for d in daily_counts if d.get("day") == today_str), 0)
    yesterday_count = next((d["count"] for d in daily_counts if d.get("day") == yesterday_str), 0)

    feed_last_visit = session.get("feed_last_visit")
    session["feed_last_visit"] = datetime.utcnow().isoformat()

    feed_query = dict(request.args)
    feed_query_no_show_hidden = {k: v for k, v in feed_query.items() if k != "show_hidden"}
    feed_query_no_search = {k: v for k, v in feed_query.items() if k != "q"}
    feed_query_layout_card = dict(feed_query_no_show_hidden)
    feed_query_layout_card["layout"] = "card"
    feed_query_layout_compact = dict(feed_query_no_show_hidden)
    feed_query_layout_compact["layout"] = "compact"
    feed_query_layout_grid = dict(feed_query_no_show_hidden)
    feed_query_layout_grid["layout"] = "grid"
    feed_query_prev = dict(feed_query)
    feed_query_next = dict(feed_query)
    feed_query_prev["page"] = page - 1
    feed_query_next["page"] = page + 1
    if page <= 1:
        feed_query_prev = None
    if page >= total_pages:
        feed_query_next = None
    return render_template(
        "index.html",
        articles=articles,
        sources=sources,
        topics=topics,
        impact_summary=impact_summary,
        domain_counts=domain_counts,
        event_types=event_types,
        event_type_counts=event_type_counts,
        country_options=country_options,
        alerts=alerts,
        alert_count_by_id=alert_count_by_id,
        window=window,
        impact_filter=request.args.get("impact"),
        domain_filter=domain,
        country_filter=country,
        country_filter_list=countries_list,
        risk_category_filter=risk_category,
        risk_category_filter_list=risk_categories_list,
        topic_filter=topic,
        topic_filter_list=topics_list,
        sort=sort,
        last_scrape_time=last_scrape,
        saved_views=saved_views,
        total_articles_in_db=total_in_db,
        total_count=total_count,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        showing_start=showing_start,
        showing_end=showing_end,
        topic_counts_sidebar=topic_counts_sidebar,
        feed_query=feed_query,
        feed_query_no_show_hidden=feed_query_no_show_hidden,
        feed_query_no_search=feed_query_no_search,
        feed_query_layout_card=feed_query_layout_card,
        feed_query_layout_compact=feed_query_layout_compact,
        feed_query_layout_grid=feed_query_layout_grid,
        feed_query_prev=feed_query_prev,
        feed_query_next=feed_query_next,
        show_hidden=show_hidden,
        feed_hidden_count=len(hidden_set),
        today_count=today_count,
        yesterday_count=yesterday_count,
        search_query=q,
        layout=layout,
        clusters=clusters,
        cluster_count_by_id=cluster_count_by_id,
        watchlists=watchlists,
        watchlist_id=watchlist_id,
        daily_counts_sparkline=daily_counts_sparkline,
        daily_sparkline_max=daily_sparkline_max,
        feed_last_visit=feed_last_visit,
        feed_seen_ids=set(session.get("feed_seen_ids") or []),
    )


@bp.route("/feed/hide/<int:article_id>", methods=["POST"])
def feed_hide_article(article_id):
    """Add article to session hidden list and redirect back to feed."""
    hidden = list(session.get("feed_hidden_ids") or [])
    if article_id not in hidden:
        hidden.append(article_id)
    session["feed_hidden_ids"] = hidden
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/feed/unhide", methods=["POST"])
def feed_unhide_all():
    """Clear session hidden list and redirect back to feed."""
    session.pop("feed_hidden_ids", None)
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/feed/seen/<int:article_id>", methods=["POST"])
def feed_mark_seen(article_id):
    """Toggle article in session 'seen' set; redirect back."""
    seen = set(session.get("feed_seen_ids") or [])
    if article_id in seen:
        seen.discard(article_id)
    else:
        seen.add(article_id)
    session["feed_seen_ids"] = list(seen)
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/watchlist/create", methods=["POST"])
def watchlist_create():
    """Create a new watchlist; redirect to feed (optionally filtered by new watchlist)."""
    name = (request.form.get("name") or "").strip()
    if not name:
        return redirect(request.referrer or url_for("main.index"))
    topic = (request.form.get("topic") or "").strip()
    topics = [topic] if topic else []
    try:
        wl_id = add_watchlist(name, topics)
        return redirect(url_for("main.index", watchlist_id=wl_id))
    except Exception:
        return redirect(request.referrer or url_for("main.index"))


@bp.route("/watchlist/<int:watchlist_id>/add_topic", methods=["POST"])
def watchlist_add_topic_route(watchlist_id):
    """Add a topic to a watchlist; redirect back. Expects form field 'topic'."""
    topic = (request.form.get("topic") or request.args.get("topic") or "").strip()
    if topic:
        watchlist_add_topic(watchlist_id, topic)
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/feed.rss")
def feed_rss():
    """RSS feed for current filter view (same params as /feed)."""
    source = request.args.get("source", "").strip() or None
    topic = request.args.get("topic", "").strip() or None
    topics_list = [t.strip() for t in request.args.getlist("topics") if t and t.strip()]
    if not topics_list and topic:
        topics_list = [topic]
        topic = None
    domain = request.args.get("domain", "").strip() or None
    country = request.args.get("country", "").strip() or None
    countries_list = [c.strip() for c in request.args.getlist("countries") if c and c.strip()]
    if not countries_list and country:
        countries_list = [country]
        country = None
    risk_category = request.args.get("risk_category", "").strip() or None
    risk_categories_list = [r.strip() for r in request.args.getlist("risk_categories") if r and r.strip()]
    if not risk_categories_list and risk_category:
        risk_categories_list = [risk_category]
        risk_category = None
    impact = _parse_impact_param(request.args.get("impact"))
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "3d": 3, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    limit = 50
    articles = get_articles(
        limit=limit,
        offset=0,
        source=source,
        topic=topic,
        topics_list=topics_list or None,
        domain=domain,
        min_impact=impact,
        days=days,
        date_from=date_from,
        date_to=date_to,
        country=country,
        countries_list=countries_list or None,
        risk_category=risk_category,
        risk_categories_list=risk_categories_list or None,
    )
    from flask import make_response
    from email.utils import formatdate
    import time
    now_rfc = formatdate(timeval=time.time(), localtime=False)
    base_url = request.host_url.rstrip("/")
    qs = request.query_string
    feed_url = (base_url + url_for("main.index") + "?" + qs.decode("utf-8")) if qs else (base_url + url_for("main.index"))
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        "<rss version=\"2.0\">",
        "<channel>",
        "<title>Intelligence Feed – Geopolitical Terminal</title>",
        "<link>" + feed_url.replace("&", "&amp;") + "</link>",
        "<description>Structured intelligence feed (current filters).</description>",
        "<lastBuildDate>" + now_rfc + "</lastBuildDate>",
    ]
    for a in articles:
        title = (a.get("title") or "Untitled").replace("&", "&amp;").replace("<", "&lt;")[:200]
        link = (a.get("url") or "").replace("&", "&amp;")
        desc = (a.get("summary") or "")[:500].replace("&", "&amp;").replace("<", "&lt;")
        pub = (a.get("published_utc") or a.get("scraped_at") or "")[:19]
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(pub.replace("Z", "+00:00"))
            pub_rfc = formatdate(timeval=dt.timestamp(), localtime=False)
        except Exception:
            pub_rfc = now_rfc
        xml.append("<item>")
        xml.append("<title>" + title + "</title>")
        xml.append("<description>" + desc + "</description>")
        xml.append("<link>" + link + "</link>")
        xml.append("<pubDate>" + pub_rfc + "</pubDate>")
        xml.append("<guid isPermaLink=\"true\">" + link + "</guid>")
        if a.get("event_type"):
            xml.append("<category>" + (a.get("event_type") or "").replace("&", "&amp;").replace("<", "&lt;") + "</category>")
        for t in (a.get("topics_list") or [])[:3]:
            xml.append("<category>" + str(t).replace("&", "&amp;").replace("<", "&lt;")[:100] + "</category>")
        if a.get("image_url"):
            img = (a.get("image_url") or "").replace("&", "&amp;")
            xml.append("<enclosure url=\"" + img + "\" type=\"image/jpeg\" length=\"0\" />")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


@bp.route("/feed/export")
def feed_export_csv():
    """Export current feed view as CSV (same filters, up to 1000 articles; ?limit=500 default)."""
    source = request.args.get("source", "").strip() or None
    topic = request.args.get("topic", "").strip() or None
    topics_list = [t.strip() for t in request.args.getlist("topics") if t and t.strip()]
    if not topics_list and topic:
        topics_list = [topic]
        topic = None
    domain = request.args.get("domain", "").strip() or None
    country = request.args.get("country", "").strip() or None
    countries_list = [c.strip() for c in request.args.getlist("countries") if c and c.strip()]
    if not countries_list and country:
        countries_list = [country]
        country = None
    risk_category = request.args.get("risk_category", "").strip() or None
    risk_categories_list = [r.strip() for r in request.args.getlist("risk_categories") if r and r.strip()]
    if not risk_categories_list and risk_category:
        risk_categories_list = [risk_category]
        risk_category = None
    impact = _parse_impact_param(request.args.get("impact"))
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "3d": 3, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    limit = min(max(1, int(request.args.get("limit", 500) or 500)), 1000)
    articles = get_articles(
        limit=limit,
        offset=0,
        source=source,
        topic=topic,
        topics_list=topics_list or None,
        domain=domain,
        min_impact=impact,
        days=days,
        date_from=date_from,
        date_to=date_to,
        country=country,
        countries_list=countries_list or None,
        risk_category=risk_category,
        risk_categories_list=risk_categories_list or None,
    )
    for a in articles:
        a["confidence_score"] = article_confidence_score(a)
        a["signal_score"] = round(article_signal_score(a), 1)
    out = StringIO()
    w = __import__("csv").writer(out)
    w.writerow(["Title", "URL", "Source", "Published", "Impact", "Event type", "Topics", "Confidence", "Signal", "Key takeaways"])
    for a in articles:
        topics_str = ", ".join(a.get("topics_list") or [])[:200]
        takeaways = (a.get("key_takeaways") or "")[:300].replace("\n", " ")
        w.writerow([
            (a.get("title") or "")[:300],
            a.get("url") or "",
            a.get("source_name") or "",
            (a.get("published_utc") or "")[:10],
            a.get("impact_score") or "",
            a.get("event_type") or "",
            topics_str,
            a.get("confidence_score") or "",
            a.get("signal_score") or "",
            takeaways,
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=intelligence_feed.csv"},
    )


@bp.route("/search")
def search():
    """Cross-platform search: articles + countries + scenarios + treaties + sanctions."""
    q = (request.args.get("q") or "").strip()
    limit = min(int(request.args.get("limit", 50) or 50), 100)
    q_lower = q.lower()

    # Articles (existing behavior)
    if q:
        articles = search_articles(q, limit=limit)
    else:
        articles = get_articles(limit=limit)

    # Countries (from canonical list)
    countries = []
    if q:
        try:
            from app.country_data import ALL_COUNTRIES, ISO3_TO_2
        except Exception:
            ALL_COUNTRIES, ISO3_TO_2 = [], {}
        iso_match = q_upper = q.upper()
        code2_match = q_upper if len(q_upper) == 2 else None
        for iso3, name, region, *_ in ALL_COUNTRIES or []:
            name_l = (name or "").lower()
            if not name_l:
                continue
            if (
                q_lower in name_l
                or iso3 == iso_match
                or (code2_match and ISO3_TO_2.get(iso3, "").upper() == code2_match)
            ):
                countries.append(
                    {
                        "code3": iso3,
                        "name": name,
                        "region": region,
                    }
                )
        countries = sorted(countries, key=lambda c: c["name"].lower())[:15]

    # Scenarios
    scenarios = []
    if q:
        try:
            all_scenarios = get_scenarios(limit=50)
        except Exception:
            all_scenarios = []
        for s in all_scenarios or []:
            name = (s.get("name") or "").strip()
            desc = (s.get("description") or "").strip()
            if q_lower in name.lower() or (desc and q_lower in desc.lower()):
                scenarios.append(s)
        scenarios = scenarios[:15]

    # Treaties
    treaties = []
    if q:
        try:
            treaties = get_treaties(search=q, limit=15)
        except Exception:
            treaties = []

    # Sanctions
    sanctions = []
    if q:
        try:
            sanctions = get_sanctions(search=q, limit=15)
        except Exception:
            sanctions = []

    sources = get_sources()
    return render_template(
        "search.html",
        query=q,
        articles=articles,
        sources=sources,
        countries=countries,
        scenarios=scenarios,
        treaties=treaties,
        sanctions=sanctions,
    )


@bp.route("/insights")
def insights():
    """Insights: digests, trending topics, story clusters, metrics, charts."""
    digest_type = request.args.get("type", "").strip() or None
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "3d": 3, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    spike_window = request.args.get("spike", "24h").strip().lower()
    if spike_window == "7d":
        spike_recent, spike_prior = 7, 7
    else:
        spike_recent, spike_prior = 1, 1

    digests = get_digests(limit=15, digest_type=digest_type)
    for d in digests:
        preview = ""
        if d.get("content"):
            try:
                data = json.loads(d["content"])
                items = data.get("items", [])
                if items and isinstance(items[0], dict):
                    preview = (items[0].get("title") or items[0].get("summary") or "")[:120]
                else:
                    preview = (d["content"][:120] if isinstance(d["content"], str) else "")
            except (json.JSONDecodeError, TypeError):
                preview = (str(d["content"])[:120] if d.get("content") else "")
        d["preview"] = (preview + "…") if preview and len(preview) >= 80 else preview

    trending = get_trending_topics(days=days, limit=15)
    clusters = get_clusters_with_counts(limit=20)
    impact_summary = get_impact_summary(days=days)
    domain_counts = get_domain_counts(days=days, limit=8)
    daily_counts = get_daily_counts(days=days)
    event_type_counts = get_feed_event_type_counts(days=days, limit=10)
    spike_topics = get_spike_topics(days_recent=spike_recent, days_prior=spike_prior, limit=10)
    declining_topics = get_declining_topics(days_recent=spike_recent, days_prior=spike_prior, limit=8)
    top_sources = get_sources_for_window(days=days, limit=10)
    uid_insights = get_effective_user_id()
    alerts = get_alerts(user_id=uid_insights)
    alert_counts = get_alert_match_counts(days=days, user_id=uid_insights)
    alert_count_by_id = {c["alert_id"]: c["count"] for c in alert_counts}
    last_scrape = get_last_scrape_time()
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    yesterday_str = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    daily_for_today = get_daily_counts(days=3)
    today_count = next((d["count"] for d in daily_for_today if d.get("day") == today_str), 0)
    yesterday_count = next((d["count"] for d in daily_for_today if d.get("day") == yesterday_str), 0)
    insights_query = dict(request.args)

    thematic_briefing = None
    thematic_topic = request.args.get("thematic", "").strip()
    if thematic_topic:
        try:
            from app.synthesis import THEMATIC_TOPICS, generate_thematic_briefing
            if thematic_topic in THEMATIC_TOPICS:
                thematic_briefing = generate_thematic_briefing(thematic_topic)
        except Exception:
            pass

    divergences = []
    try:
        from app.synthesis import get_signal_divergences
        divergences = get_signal_divergences()[:3]
    except Exception:
        pass

    return render_template(
        "insights.html",
        digests=digests,
        trending=trending,
        clusters=clusters,
        impact_summary=impact_summary,
        domain_counts=domain_counts,
        window=window,
        daily_counts=daily_counts,
        spike_topics=spike_topics,
        declining_topics=declining_topics,
        event_type_counts=event_type_counts,
        top_sources=top_sources,
        alerts=alerts,
        alert_count_by_id=alert_count_by_id,
        last_scrape_time=last_scrape,
        today_count=today_count,
        yesterday_count=yesterday_count,
        spike_window=spike_window,
        insights_query=insights_query,
        thematic_briefing=thematic_briefing,
        thematic_topic=thematic_topic or None,
        divergences=divergences,
    )


@bp.route("/intelligence/macro")
def macro_dashboard():
    """Live macroeconomic indicators dashboard (time-series)."""
    try:
        # Ensure defaults exist even before the background scheduler runs
        from app.models import macro_seed_defaults
        macro_seed_defaults()
    except Exception:
        pass
    return render_template("macro_dashboard.html")


@bp.route("/export/macro/pdf")
def export_macro_pdf():
    """Print-friendly macro report page; user can Print to PDF."""
    from datetime import datetime, timezone
    group = (request.args.get("group") or "").strip().lower() or None
    if group not in (None, "", "asean", "g20", "major"):
        group = None
    indicator = (request.args.get("indicator") or "").strip().lower() or None
    try:
        from app.models import macro_get_latest
        latest = macro_get_latest(group=group, indicator=indicator, limit=200)
    except Exception:
        latest = []
    try:
        from app.api_routes import api_macro_alerts
        # Reuse the API logic by calling the underlying models again (kept simple here)
        from app.models import macro_get_latest as _ml
        infl = _ml(group=group, indicator="inflation_cpi", limit=500)
        gdp = _ml(group=group, indicator="gdp_growth_yoy", limit=500)
        debt = _ml(group=group, indicator="gov_debt_pct_gdp", limit=500)
        # Inline rule evaluation (same thresholds as /api/macro/alerts)
        def idx(rows):
            return {r["country_code"]: r for r in (rows or []) if r.get("country_code")}
        infl_i, gdp_i, debt_i = idx(infl), idx(gdp), idx(debt)
        alerts = []
        for code, row in infl_i.items():
            try:
                v = float(row.get("value"))
            except (TypeError, ValueError):
                continue
            if v >= 8:
                alerts.append({"severity": "high" if v >= 12 else "med", "country_code": code, "country_name": row.get("country_name"), "value": v, "unit": row.get("unit") or "%", "date": row.get("date"), "message": f"High inflation ({v:.1f}%)."})
        for code, row in gdp_i.items():
            try:
                v = float(row.get("value"))
            except (TypeError, ValueError):
                continue
            if v <= 0:
                alerts.append({"severity": "high" if v <= -2 else "med", "country_code": code, "country_name": row.get("country_name"), "value": v, "unit": row.get("unit") or "%", "date": row.get("date"), "message": f"Negative growth ({v:.1f}%)."})
        for code, row in debt_i.items():
            try:
                v = float(row.get("value"))
            except (TypeError, ValueError):
                continue
            if v >= 90:
                alerts.append({"severity": "high" if v >= 120 else "med", "country_code": code, "country_name": row.get("country_name"), "value": v, "unit": row.get("unit") or "%", "date": row.get("date"), "message": f"High government debt ({v:.0f}% of GDP)."})
        sev_rank = {"high": 0, "med": 1, "low": 2}
        alerts.sort(key=lambda a: (sev_rank.get(a.get("severity") or "low", 9), a.get("country_name") or a.get("country_code") or ""))
    except Exception:
        alerts = []

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return render_template(
        "export_macro_pdf.html",
        group=group,
        indicator=indicator,
        latest=latest,
        alerts=alerts,
        generated_at=generated_at,
    )


@bp.route("/alerts", methods=["GET", "POST"])
def alerts_list():
    """Watchlists: list and create watchlists (topics/countries, min impact, webhook). Create requires login."""
    uid = get_effective_user_id()
    if request.method == "POST" and request.form.get("form_type") == "alert":
        if not current_user.is_authenticated and not uid:
            flash("Log in to create watchlists.", "error")
            return redirect(url_for("main.login", next=url_for("main.alerts_list")))
        name = (request.form.get("name") or "").strip()
        topics_raw = request.form.get("topics") or ""
        topics = [t.strip() for t in topics_raw.split(",") if t.strip()]
        try:
            min_score = int(request.form.get("min_impact_score") or 5)
        except ValueError:
            min_score = 2
        webhook_url = (request.form.get("webhook_url") or "").strip() or None
        if name:
            add_alert(name, topics, min_score, webhook_url=webhook_url, user_id=uid)
        return redirect(url_for("main.alerts_list", _anchor="create-watchlist"))

    alerts = get_alerts(user_id=uid)
    alert_topic_suggestions, alert_country_suggestions = get_alert_suggestions()
    alert_counts = get_alert_match_counts(days=1, user_id=uid)
    alert_count_by_id = {c["alert_id"]: c["count"] for c in alert_counts}
    return render_template(
        "alerts_list.html",
        alerts=alerts,
        alert_topic_suggestions=alert_topic_suggestions,
        alert_country_suggestions=alert_country_suggestions,
        alert_count_by_id=alert_count_by_id,
    )


@bp.route("/insights/digests.rss")
def insights_digests_rss():
    """RSS feed of latest digests (same type filter as insights page)."""
    digest_type = request.args.get("type", "").strip() or None
    limit = 20
    digests = get_digests(limit=limit, digest_type=digest_type)
    from flask import make_response
    from email.utils import formatdate
    import time
    now_rfc = formatdate(timeval=time.time(), localtime=False)
    base_url = request.host_url.rstrip("/")
    feed_url = base_url + url_for("main.insights") + ("?type=" + digest_type if digest_type else "")
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        "<rss version=\"2.0\">",
        "<channel>",
        "<title>Digests – Geopolitical Terminal</title>",
        "<link>" + (feed_url.replace("&", "&amp;") if feed_url else base_url) + "</link>",
        "<description>Latest digests (daily/weekly).</description>",
        "<lastBuildDate>" + now_rfc + "</lastBuildDate>",
    ]
    for d in digests:
        title = (d.get("title") or "Digest").replace("&", "&amp;").replace("<", "&lt;")[:200]
        link = base_url + url_for("main.digest_detail", digest_id=d["id"])
        pub = (d.get("created_at") or "")[:19]
        try:
            from datetime import datetime as _dt
            dt = _dt.fromisoformat(pub.replace("Z", "+00:00"))
            pub_rfc = formatdate(timeval=dt.timestamp(), localtime=False)
        except Exception:
            pub_rfc = now_rfc
        desc = (d.get("digest_type") or "") + " — " + (pub[:10] or "")
        xml.append("<item>")
        xml.append("<title>" + title + "</title>")
        xml.append("<description>" + desc.replace("&", "&amp;") + "</description>")
        xml.append("<link>" + link + "</link>")
        xml.append("<pubDate>" + pub_rfc + "</pubDate>")
        xml.append("<guid isPermaLink=\"true\">" + link + "</guid>")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


@bp.route("/digests/<int:digest_id>")
def digest_detail(digest_id):
    d = get_digest(digest_id)
    if not d:
        return "Digest not found", 404
    try:
        data = json.loads(d["content"])
        d["items"] = data.get("items", [])
    except (json.JSONDecodeError, TypeError):
        d["items"] = []
    policy_context = _policy_entity_page_context(get_effective_user_id(), "digest", str(digest_id), d.get("created_at"))
    return render_template("digest_detail.html", digest=d, policy_context=policy_context)


@bp.route("/clusters/<int:cluster_id>")
def cluster_detail(cluster_id):
    articles = get_articles_by_cluster(cluster_id)
    if not articles:
        return "Cluster not found", 404
    cluster_label = get_cluster_label(cluster_id)
    # Group by source for "How outlets framed this"
    by_source = defaultdict(list)
    for a in articles:
        by_source[a["source_name"]].append(a)
    framing = [{"source": src, "articles": arts} for src, arts in sorted(by_source.items())]
    # Per-cluster implications (LLM synthesis for clusters with >= 3 articles)
    cluster_implications = None
    if len(articles) >= 3:
        from app.synthesis import get_cluster_implications
        cluster_implications = get_cluster_implications(cluster_id)
    return render_template(
        "cluster_detail.html",
        articles=articles,
        cluster_id=cluster_id,
        cluster_label=cluster_label,
        framing=framing,
        cluster_implications=cluster_implications,
    )


@bp.route("/watchlist/<int:watchlist_id>")
def watchlist_detail(watchlist_id):
    wl = get_watchlist(watchlist_id)
    if not wl:
        return "Watchlist not found", 404
    articles = get_articles_for_watchlist(watchlist_id)
    return render_template("watchlist.html", watchlist=wl, articles=articles)


@bp.route("/watchlist/<int:watchlist_id>/edit", methods=["GET", "POST"])
def watchlist_edit(watchlist_id):
    wl = get_watchlist(watchlist_id)
    if not wl:
        return "Watchlist not found", 404
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        topics_raw = request.form.get("topics") or ""
        topics = [t.strip() for t in topics_raw.split(",") if t.strip()]
        if name and update_watchlist(watchlist_id, name, topics):
            return redirect(url_for("main.watchlist_detail", watchlist_id=watchlist_id))
    return render_template("watchlist_edit.html", watchlist=wl)


@bp.route("/watchlist/<int:watchlist_id>/delete", methods=["POST"])
def watchlist_delete(watchlist_id):
    if delete_watchlist(watchlist_id):
        return redirect(url_for("main.index"))
    return "Watchlist not found", 404


@bp.route("/alerts/<int:alert_id>")
def alert_detail(alert_id):
    alert = get_alert(alert_id, user_id=get_effective_user_id())
    if not alert:
        return "Alert not found", 404
    days_param = request.args.get("days", "7")
    try:
        days = max(1, min(30, int(days_param)))
    except ValueError:
        days = 7
    matches = get_alert_matches(alert_id, days=days, limit=100)
    return render_template("alert_detail.html", alert=alert, matches=matches, days=days)


@bp.route("/alerts/<int:alert_id>/edit", methods=["GET", "POST"])
@login_required
def alert_edit(alert_id):
    alert = get_alert(alert_id, user_id=current_user.id)
    if not alert:
        return "Watchlist not found", 404
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        topics_raw = request.form.get("topics") or ""
        topics = [t.strip() for t in topics_raw.split(",") if t.strip()]
        try:
            min_score = int(request.form.get("min_impact_score") or 5)
        except ValueError:
            min_score = 5
        webhook_url = (request.form.get("webhook_url") or "").strip() or None
        if name and update_alert(alert_id, name, topics, min_impact_score=min_score, webhook_url=webhook_url, user_id=current_user.id):
            return redirect(url_for("main.alert_detail", alert_id=alert_id))
    alert_topic_suggestions, alert_country_suggestions = get_alert_suggestions()
    return render_template(
        "alert_edit.html",
        alert=alert,
        alert_topic_suggestions=alert_topic_suggestions,
        alert_country_suggestions=alert_country_suggestions,
    )


@bp.route("/alerts/<int:alert_id>/delete", methods=["POST"])
@login_required
def alert_delete(alert_id):
    if delete_alert(alert_id, user_id=current_user.id):
        return redirect(url_for("main.alerts_list"))
    return "Alert not found", 404


# --- Saved views (user-scoped) ---
@bp.route("/saved-views/save", methods=["POST"])
@login_required
def save_view():
    name = (request.form.get("name") or "").strip()
    if not name:
        return redirect(request.referrer or url_for("main.index"))
    params = dict(request.args)
    add_saved_view(name, params, user_id=current_user.id)
    return redirect(request.referrer or url_for("main.index"))


@bp.route("/saved-views")
def saved_views_list():
    views = get_saved_views(user_id=get_effective_user_id())
    return render_template("saved_views.html", saved_views=views)


@bp.route("/saved-views/<int:view_id>")
def load_saved_view(view_id):
    v = get_saved_view(view_id, user_id=get_effective_user_id())
    if not v:
        return "View not found", 404
    return redirect(url_for("main.index", **v.get("params", {})))


@bp.route("/saved-views/<int:view_id>/delete", methods=["POST"])
@login_required
def delete_saved_view_route(view_id):
    delete_saved_view(view_id, user_id=current_user.id)
    return redirect(url_for("main.saved_views_list"))


# --- Article detail + annotations ---
@bp.route("/article/<int:article_id>")
def article_detail(article_id):
    a = get_article(article_id)
    if not a:
        return "Article not found", 404
    if a.get("topics"):
        try:
            a["topics_list"] = json.loads(a["topics"]) if isinstance(a["topics"], str) else a["topics"]
        except (json.JSONDecodeError, TypeError):
            a["topics_list"] = []
    else:
        a["topics_list"] = []
    if a.get("entities"):
        try:
            a["entities_list"] = json.loads(a["entities"]) if isinstance(a["entities"], str) else a["entities"]
        except (json.JSONDecodeError, TypeError):
            a["entities_list"] = []
    else:
        a["entities_list"] = []
    annotations = get_annotations_for_article(article_id)
    implications = None
    try:
        from app.institutional_models.article_implications import get_article_implications
        implications = get_article_implications(a)
    except Exception:
        pass
    policy_context = _policy_entity_page_context(get_effective_user_id(), "article", str(article_id), a.get("scraped_at"))
    return render_template(
        "article_detail.html",
        article=a,
        annotations=annotations,
        implications=implications,
        policy_context=policy_context,
    )


@bp.route("/article/<int:article_id>/annotations", methods=["POST"])
def article_add_annotation(article_id):
    body = (request.form.get("body") or "").strip()
    if body:
        add_annotation(article_id, body)
    return redirect(url_for("main.article_detail", article_id=article_id))


@bp.route("/annotation/<int:annotation_id>/delete", methods=["POST"])
def annotation_delete(annotation_id):
    ann = get_annotation(annotation_id)
    article_id = ann["article_id"] if ann else None
    delete_annotation(annotation_id)
    if article_id:
        return redirect(url_for("main.article_detail", article_id=article_id))
    return redirect(url_for("main.index"))


# --- Entity / country page ---
@bp.route("/entity/<path:name>")
def entity_page(name):
    name = name.strip()
    days = request.args.get("days", type=int) or 7
    limit = min(int(request.args.get("limit", 100) or 100), 200)
    articles = get_articles_by_entity(name, limit=limit, days=days)
    return render_template("entity.html", entity_name=name, articles=articles, days=days)


# --- Timeline ---
@bp.route("/timeline")
def timeline():
    topic = request.args.get("topic", "").strip() or None
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    articles = get_articles(limit=200, topic=topic, days=days)
    topics = get_all_topic_counts()
    return render_template("timeline.html", articles=articles, topics=topics, window=window)


# --- Map ---
@bp.route("/map")
def map_view():
    window = request.args.get("window", "7d")
    window_map = {"1d": 1, "7d": 7, "30d": 30}
    days = window_map.get(window, 7)
    country_counts = get_article_counts_by_country(days=days, limit=40)
    return render_template("map.html", country_counts=country_counts, window=window)


@bp.route("/world-monitor")
def world_monitor():
    """Global monitor map. URL params: view | (lat, lon, zoom), timeRange, layers.
    Mirrors koala73/worldmonitor parseMapUrlState/buildMapUrl (urlState.ts)."""
    from app.world_monitor_config import (
        get_world_monitor_layer_ids,
        get_world_monitor_live_layer_ids,
        get_world_monitor_layer_labels,
        get_world_monitor_time_days,
        get_view_preset_coords,
        clamp_lat,
        clamp_lon,
        clamp_zoom,
        WORLD_MONITOR_VIEW_LABELS,
    )
    from app.models import get_last_scrape_time

    # View preset overrides lat/lon/zoom (their parseMapUrlState: view param)
    view_param = (request.args.get("view") or "").strip().lower()
    coords = get_view_preset_coords(view_param) if view_param else None
    if coords:
        lat, lon, zoom = coords[0], coords[1], coords[2]
    else:
        lat = request.args.get("lat", type=float) or 20.0
        lon = request.args.get("lon", type=float) or 0.0
        zoom = request.args.get("zoom", type=float) or 1.5
    lat = clamp_lat(lat)
    lon = clamp_lon(lon)
    zoom = clamp_zoom(zoom)

    time_range = (request.args.get("timeRange") or "1d").strip().lower()
    time_days = get_world_monitor_time_days(time_range)

    # Layers: comma-separated or repeated (their LAYER_KEYS + split(','))
    layers_list = request.args.getlist("layers")
    if layers_list:
        layers = []
        for x in layers_list:
            layers.extend([y.strip() for y in (x or "").split(",") if y.strip()])
    else:
        layers_param = request.args.get("layers") or ""
        layers = [x.strip() for x in layers_param.split(",") if x.strip()]

    # Default layers when none selected: show a dense map (Iran Attacks, Intel Hotspots, Conflict Zones, Bases, Nuclear, Spaceports, Cables, Pipelines, AI Data Centers, Military, Waterways, Sanctions)
    if not layers:
        layers = [
            "iranAttacks", "hotspots", "conflicts", "bases", "nuclear",
            "gamma_irradiators", "spaceports", "undersea_cables", "pipelines", "ai_datacenters",
            "military", "waterways", "sanctions", "natural", "weather", "economic",
        ]

    points = get_world_monitor_points(time_range_days=time_days, layers=layers) if layers else []
    all_layer_ids = get_world_monitor_layer_ids()
    live_layer_ids = get_world_monitor_live_layer_ids()
    layer_labels = get_world_monitor_layer_labels()
    last_updated = get_last_scrape_time()

    # Presets as (view_key, display_label) for toolbar (their VIEW_VALUES)
    view_presets = [(k, WORLD_MONITOR_VIEW_LABELS.get(k, k)) for k in WORLD_MONITOR_VIEW_LABELS]
    current_view = view_param if view_param in WORLD_MONITOR_VIEW_LABELS else "global"

    return render_template(
        "world_monitor.html",
        lat=lat,
        lon=lon,
        zoom=zoom,
        time_range=time_range,
        time_days=time_days,
        layers=layers,
        all_layer_ids=all_layer_ids,
        layer_labels=layer_labels,
        points=points,
        live_layer_ids=live_layer_ids,
        last_updated=last_updated,
        view_presets=view_presets,
        current_view=current_view,
    )


# --- Real-Time Geopolitical Risk Engine ---
@bp.route("/risk")
def risk_dashboard():
    """Risk Engine: heat map data, country scores, sector exposure, Forward Risk Probability Index."""
    heat_map = get_country_risk_snapshots()
    risk_index_list = get_risk_index()  # all regions
    gepi_latest = None
    try:
        from app.institutional_models.readers import get_gepi_latest
        gepi_latest = get_gepi_latest()
    except Exception:
        pass
    macro_stress_alerts = get_macroeconomic_stress_alerts(threshold_debt=80, threshold_inflation=15, limit=15)
    climate_vulnerability_summary = get_climate_vulnerability_summary(limit=12)
    trajectory_data = None
    try:
        from app.synthesis import get_escalation_trajectory
        trajectory_data = get_escalation_trajectory(days=30)
    except Exception:
        pass
    return render_template(
        "risk.html",
        heat_map=heat_map,
        risk_index=risk_index_list,
        gepi_latest=gepi_latest,
        macro_stress_alerts=macro_stress_alerts,
        climate_vulnerability_summary=climate_vulnerability_summary,
        trajectory_data=trajectory_data,
    )


# --- Supply chain & trade flow intelligence ---

@bp.route("/supply-chain")
def supply_chain_dashboard():
    """Trade & Supply Chain: global trade flow map (maritime, LNG, semiconductor, rare earth) + side panel on chokepoint click."""
    chokepoints = get_chokepoints_with_geo()
    selected_slug = request.args.get("chokepoint", "").strip().lower()
    selected = None
    if selected_slug:
        for cp in chokepoints:
            if (cp.get("slug") or "").strip().lower() == selected_slug:
                selected = cp
                break
    if not selected and chokepoints:
        selected = chokepoints[0]
        selected_slug = (selected.get("slug") or "").strip().lower()

    flows = []
    naval_deployments = []
    naval_count = 0
    pct_global = None
    insurance_spike = None
    if selected:
        flows = get_flows_for_chokepoint(selected["id"])
        flows = sorted(flows, key=lambda f: -(float(f.get("exposure_pct") or 0)))
        region = (selected.get("region") or "").strip()
        if region:
            naval_deployments = get_naval_deployments(region=region, limit=20)
            naval_count = len(naval_deployments)
        pct_global = selected.get("pct_global_trade")
        exposure_sum = sum(float(f.get("exposure_pct") or 0) for f in flows)
        insurance_spike = min(80, 15 + (exposure_sum / 100.0))

    flows_by_chokepoint = {}
    naval_by_region = {}
    for cp in chokepoints:
        fl = get_flows_for_chokepoint(cp["id"])
        flows_by_chokepoint[str(cp["id"])] = sorted(fl, key=lambda f: -(float(f.get("exposure_pct") or 0)))
        reg = (cp.get("region") or "").strip()
        if reg and reg not in naval_by_region:
            naval_by_region[reg] = get_naval_deployments(region=reg, limit=20)

    flow_countries = len(set(f.get("country_code") for f in flows)) if flows else 0
    flow_sectors = len(set(f.get("sector") for f in flows)) if flows else 0
    top_exposed_codes = [f.get("country_code") for f in flows[:8] if f.get("country_code")]

    energy_commodity_summary = get_energy_commodity_summary(limit=12)
    geospatial_summary = get_geospatial_infrastructure_summary(limit=20)
    tech_semiconductor_summary = get_technology_semiconductor_summary(limit=12)
    trade_supply_chain_articles = get_articles_trade_supply_chain_relevant(limit=15, days=14)
    airspace_restrictions = get_airspace_restrictions()

    coords_fallback = {k: list(v) for k, v in CHOKEPOINT_COORDS.items()}
    return render_template(
        "trade_supply_chain.html",
        chokepoints=chokepoints,
        selected=selected,
        flows=flows,
        naval_deployments=naval_deployments,
        naval_count=naval_count,
        pct_global_trade=pct_global,
        insurance_spike=round(insurance_spike, 1) if insurance_spike is not None else None,
        flows_by_chokepoint=flows_by_chokepoint,
        naval_by_region=naval_by_region,
        flow_countries=flow_countries,
        flow_sectors=flow_sectors,
        top_exposed_codes=top_exposed_codes,
        scenario_url=url_for("main.supply_chain_scenario"),
        chokepoint_detail_base=url_for("main.supply_chain_chokepoint_detail", chokepoint_id=0).replace("0", ""),
        country_base=url_for("main.country_dashboard", country_code="XXX").replace("XXX", ""),
        sanctions_check_url=url_for("main.sanctions_watch_supply_chain_check"),
        coords_fallback=coords_fallback,
        energy_commodity_summary=energy_commodity_summary,
        geospatial_summary=geospatial_summary,
        tech_semiconductor_summary=tech_semiconductor_summary,
        trade_supply_chain_articles=trade_supply_chain_articles,
        airspace_restrictions=airspace_restrictions,
    )


@bp.route("/supply-chain/scenario", methods=["GET", "POST"])
def supply_chain_scenario():
    """If chokepoint X closes → impact on selected countries (instant scenario tree). Compare two chokepoints optional."""
    chokepoints = get_chokepoints()
    countries = get_chokepoint_countries()
    result = None
    result_b = None
    selected_countries = []
    preselected_chokepoint = request.args.get("chokepoint")
    preselected_countries = [c.strip().upper() for c in request.args.get("countries", "").split(",") if c.strip()]
    flows_by_chokepoint = {}
    for cp in chokepoints:
        fl = get_flows_for_chokepoint(cp["id"])
        flows_by_chokepoint[str(cp["id"])] = sorted(fl, key=lambda f: -(float(f.get("exposure_pct") or 0)))
    if request.method == "POST":
        chokepoint_id = request.form.get("chokepoint_id")
        chokepoint_id_b = request.form.get("chokepoint_id_b", "").strip()
        selected_countries = request.form.getlist("countries") or []
        if chokepoint_id and chokepoint_id.isdigit():
            result = run_chokepoint_scenario(int(chokepoint_id), selected_countries)
        if chokepoint_id_b and chokepoint_id_b.isdigit() and chokepoint_id_b != chokepoint_id:
            result_b = run_chokepoint_scenario(int(chokepoint_id_b), selected_countries)
    else:
        selected_countries = preselected_countries
    return render_template(
        "supply_chain_scenario.html",
        chokepoints=chokepoints,
        countries=countries,
        result=result,
        result_b=result_b,
        selected_countries=selected_countries,
        preselected_chokepoint=preselected_chokepoint,
        flows_by_chokepoint=flows_by_chokepoint,
        conflict_url=url_for("main.conflict_dashboard"),
    )


@bp.route("/supply-chain/chokepoint/<int:chokepoint_id>")
def supply_chain_chokepoint_detail(chokepoint_id):
    """Detail view for one chokepoint: flows by country/sector."""
    chokepoint = get_chokepoint(chokepoint_id)
    if not chokepoint:
        return "Chokepoint not found", 404
    flows = get_flows_for_chokepoint(chokepoint_id)
    return render_template(
        "supply_chain_chokepoint.html",
        chokepoint=chokepoint,
        flows=flows,
        conflict_url=url_for("main.conflict_dashboard"),
    )


# --- Political stability & domestic signals ---
def _stability_fragility_diff(current_codes):
    """Compare current fragility country codes to session; return (added, removed) and update session."""
    prev = set(session.get("stability_prev_fragility") or [])
    current = set(current_codes)
    added = sorted(current - prev)
    removed = sorted(prev - current)
    session["stability_prev_fragility"] = sorted(current)
    return added, removed


@bp.route("/stability")
def stability_dashboard():
    """Political stability & domestic signals: elections, approval, protests, currency, food inflation, youth unemployment, sentiment. Real-time fragility monitoring. Supports region, date, protest_days, low_approval_threshold."""
    from datetime import datetime, timedelta
    region = request.args.get("region", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    try:
        protest_days = max(1, min(365, int(request.args.get("protest_days") or 90)))
    except (TypeError, ValueError):
        protest_days = 90
    try:
        low_approval_threshold = max(1, min(99, float(request.args.get("low_approval_threshold") or FRAGILITY_LOW_APPROVAL_THRESHOLD)))
    except (TypeError, ValueError):
        low_approval_threshold = FRAGILITY_LOW_APPROVAL_THRESHOLD
    today = datetime.utcnow().strftime("%Y-%m-%d")
    last_30 = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    fragility = get_fragility_overview(
        limit=20,
        protest_days=protest_days,
        region=region,
        low_approval_threshold=low_approval_threshold,
    )
    fragility_count = len(fragility.get("countries") or [])
    fragility_added, fragility_removed = _stability_fragility_diff([c["country_code"] for c in (fragility.get("countries") or [])])
    elections = get_election_calendar(region=region, date_from=date_from, date_to=date_to, limit=20)
    approval = get_approval_ratings(region=region, date_from=date_from, date_to=date_to, limit=15)
    protests = get_protest_tracking(region=region, date_from=date_from, date_to=date_to, limit=15)
    election_count = get_election_calendar_count(region=region, date_from=date_from, date_to=date_to)
    approval_count = get_approval_ratings_count(region=region, date_from=date_from, date_to=date_to)
    protest_count = get_protest_tracking_count(region=region, date_from=date_from, date_to=date_to)
    query_params = {}
    if region:
        query_params["region"] = region
    if date_from:
        query_params["date_from"] = date_from
    if date_to:
        query_params["date_to"] = date_to
    if protest_days != 90:
        query_params["protest_days"] = protest_days
    if low_approval_threshold != FRAGILITY_LOW_APPROVAL_THRESHOLD:
        query_params["low_approval_threshold"] = low_approval_threshold
    integration_url = url_for(
        "main.integration_dashboard",
        sort="geopolitical_fragility_score",
        order="desc",
        per_page="all",
    )
    if region:
        integration_url = url_for(
            "main.integration_dashboard",
            region=region,
            sort="geopolitical_fragility_score",
            order="desc",
            per_page="all",
        )
    upcoming_elections_url = url_for("main.approval_protests", election_date_from=today, election_status="upcoming") + "#elections"
    return render_template(
        "stability.html",
        fragility=fragility,
        elections=elections,
        approval=approval,
        protests=protests,
        currency_stress=get_currency_stress(region=region, limit=15),
        food_inflation=get_food_inflation_alerts(region=region, limit=15),
        youth_unemployment=get_youth_unemployment(region=region, limit=15),
        social_sentiment=get_social_sentiment(region=region, limit=15),
        regions=get_election_regions(),
        region=region,
        date_from=date_from,
        date_to=date_to,
        election_count=election_count,
        approval_count=approval_count,
        protest_count=protest_count,
        fragility_count=fragility_count,
        fragility_added=fragility_added,
        fragility_removed=fragility_removed,
        approval_protests_url=url_for("main.approval_protests"),
        fragility_export_url=url_for("main.stability_export_fragility", **query_params),
        stability_export_full_url=url_for("main.stability_export_full", **query_params),
        stability_fragility_rss_url=url_for("main.stability_fragility_rss", **query_params),
        stability_dashboard_url=url_for("main.stability_dashboard"),
        integration_dashboard_url=integration_url,
        low_approval_threshold=int(low_approval_threshold),
        protest_days=protest_days,
        filter_query=query_params,
        last_30_from=last_30,
        last_30_to=today,
        today=today,
        upcoming_elections_url=upcoming_elections_url,
        elite_institutional_summary=get_elite_institutional_summary(region=region, limit=12),
        climate_vulnerability_summary=get_climate_vulnerability_summary(limit=12),
    )


@bp.route("/stability/export/fragility")
def stability_export_fragility():
    """Export fragility overview as CSV (honors region, protest_days, low_approval_threshold)."""
    region = request.args.get("region", "").strip() or None
    try:
        protest_days = max(1, min(365, int(request.args.get("protest_days") or 90)))
    except (TypeError, ValueError):
        protest_days = 90
    try:
        low_approval_threshold = max(1, min(99, float(request.args.get("low_approval_threshold") or FRAGILITY_LOW_APPROVAL_THRESHOLD)))
    except (TypeError, ValueError):
        low_approval_threshold = FRAGILITY_LOW_APPROVAL_THRESHOLD
    fragility = get_fragility_overview(limit=500, protest_days=protest_days, region=region, low_approval_threshold=low_approval_threshold)
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Country code", "Country", "Risk score", "Signals (summary)"])
    for c in fragility.get("countries") or []:
        parts = []
        for s in c.get("signals") or []:
            p = (s.get("type") or "").replace("_", " ")
            if s.get("level"):
                p += " (" + str(s["level"]) + ")"
            if s.get("approval_pct") is not None:
                p += " " + str(s["approval_pct"]) + "%"
            parts.append(p)
        signals_str = "; ".join(parts)
        w.writerow([c.get("country_code") or "", c.get("country_name") or "", c.get("risk_score") or 0, signals_str])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=stability_fragility.csv"},
    )


@bp.route("/stability/export/full")
def stability_export_full():
    """Export full stability view as XLSX (all sections with current filters)."""
    region = request.args.get("region", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    try:
        protest_days = max(1, min(365, int(request.args.get("protest_days") or 90)))
    except (TypeError, ValueError):
        protest_days = 90
    try:
        low_approval_threshold = max(1, min(99, float(request.args.get("low_approval_threshold") or FRAGILITY_LOW_APPROVAL_THRESHOLD)))
    except (TypeError, ValueError):
        low_approval_threshold = FRAGILITY_LOW_APPROVAL_THRESHOLD
    fragility = get_fragility_overview(limit=500, protest_days=protest_days, region=region, low_approval_threshold=low_approval_threshold)
    elections = get_election_calendar(region=region, date_from=date_from, date_to=date_to, limit=500)
    approval = get_approval_ratings(region=region, date_from=date_from, date_to=date_to, limit=500)
    protests = get_protest_tracking(region=region, date_from=date_from, date_to=date_to, limit=500)
    currency_stress = get_currency_stress(region=region, limit=500)
    food_inflation = get_food_inflation_alerts(region=region, limit=500)
    youth_unemployment = get_youth_unemployment(region=region, limit=500)
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        return Response("XLSX export requires openpyxl", status=501, mimetype="text/plain")
    wb = Workbook()
    ws = wb.active
    ws.title = "Fragility"
    ws.append(["Country code", "Country", "Risk score", "Signals"])
    for c in fragility.get("countries") or []:
        parts = []
        for s in c.get("signals") or []:
            p = (s.get("type") or "").replace("_", " ")
            if s.get("level"):
                p += " (" + str(s["level"]) + ")"
            if s.get("approval_pct") is not None:
                p += " " + str(s["approval_pct"]) + "%"
            parts.append(p)
        ws.append([c.get("country_code") or "", c.get("country_name") or "", c.get("risk_score") or 0, "; ".join(parts)])
    ws2 = wb.create_sheet("Elections")
    ws2.append(["Country code", "Country", "Type", "Date planned", "Status", "Notes"])
    for e in elections:
        ws2.append([e.get("country_code") or "", e.get("country_name") or "", e.get("election_type") or "", (e.get("date_planned") or "")[:10], e.get("status") or "", (e.get("notes") or "")[:200]])
    ws3 = wb.create_sheet("Approval")
    ws3.append(["Country code", "Country", "Subject", "Approval %", "Poll date", "Source"])
    for a in approval:
        ws3.append([a.get("country_code") or "", a.get("country_name") or "", a.get("subject") or "", a.get("approval_pct"), (a.get("poll_date") or "")[:10], a.get("source") or ""])
    ws4 = wb.create_sheet("Protests")
    ws4.append(["Country code", "Country", "Date", "Summary", "Size", "Trigger"])
    for p in protests:
        ws4.append([p.get("country_code") or "", p.get("country_name") or "", (p.get("event_date") or "")[:10], (p.get("summary") or "")[:200], p.get("estimated_size") or "", p.get("trigger_topic") or ""])
    ws5 = wb.create_sheet("Currency stress")
    ws5.append(["Country code", "Country", "Indicator", "Level", "As of", "Notes"])
    for c in currency_stress:
        ws5.append([c.get("country_code") or "", c.get("country_name") or "", c.get("indicator_value"), c.get("stress_level") or "", (c.get("as_of_date") or "")[:10], (c.get("notes") or "")[:200]])
    ws6 = wb.create_sheet("Food inflation")
    ws6.append(["Country code", "Country", "Inflation %", "Risk level", "As of", "Notes"])
    for f in food_inflation:
        ws6.append([f.get("country_code") or "", f.get("country_name") or "", f.get("inflation_pct"), f.get("risk_level") or "", (f.get("as_of_date") or "")[:10], (f.get("notes") or "")[:200]])
    ws7 = wb.create_sheet("Youth unemployment")
    ws7.append(["Country code", "Country", "Rate %", "As of", "Source"])
    for y in youth_unemployment:
        ws7.append([y.get("country_code") or "", y.get("country_name") or "", y.get("rate_pct"), (y.get("as_of_date") or "")[:10], y.get("source") or ""])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stability_full.xlsx"},
    )


@bp.route("/stability/fragility.rss")
def stability_fragility_rss():
    """RSS feed of current fragility watchlist (countries on the list)."""
    region = request.args.get("region", "").strip() or None
    try:
        protest_days = max(1, min(365, int(request.args.get("protest_days") or 90)))
    except (TypeError, ValueError):
        protest_days = 90
    try:
        low_approval_threshold = max(1, min(99, float(request.args.get("low_approval_threshold") or FRAGILITY_LOW_APPROVAL_THRESHOLD)))
    except (TypeError, ValueError):
        low_approval_threshold = FRAGILITY_LOW_APPROVAL_THRESHOLD
    fragility = get_fragility_overview(limit=100, protest_days=protest_days, region=region, low_approval_threshold=low_approval_threshold)
    from flask import make_response
    from email.utils import formatdate
    import time
    now_rfc = formatdate(timeval=time.time(), localtime=False)
    base_url = request.host_url.rstrip("/")
    stability_url = base_url + url_for("main.stability_dashboard")
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        "<rss version=\"2.0\">",
        "<channel>",
        "<title>Political Stability – Fragility Watchlist</title>",
        "<link>" + stability_url + "</link>",
        "<description>Countries with elevated domestic fragility signals (currency stress, food inflation, recent protests, low approval).</description>",
        "<lastBuildDate>" + now_rfc + "</lastBuildDate>",
    ]
    for c in fragility.get("countries") or []:
        signals_str = ", ".join((s.get("type") or "").replace("_", " ") + (" " + str(s.get("approval_pct", "")) + "%" if s.get("approval_pct") is not None else "") for s in (c.get("signals") or []))
        title = ((c.get("country_name") or "") + " (" + (c.get("country_code") or "") + ") – risk " + str(c.get("risk_score", 0))).replace("&", "&amp;").replace("<", "&lt;")[:100]
        link = base_url + url_for("main.country_dashboard", country_code=c.get("country_code"))
        xml.append("<item>")
        xml.append("<title>" + title + "</title>")
        xml.append("<description>" + (signals_str or "—").replace("&", "&amp;").replace("<", "&lt;")[:300] + "</description>")
        xml.append("<link>" + link.replace("&", "&amp;") + "</link>")
        xml.append("<guid isPermaLink=\"true\">" + link.replace("&", "&amp;") + "</guid>")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


# Election type and status options for filters/forms
ELECTION_TYPES = ["presidential", "parliamentary", "general", "midterm", "municipal", "regional", "referendum", "other"]
ELECTION_STATUSES = ["upcoming", "completed", "postponed", "cancelled"]
@bp.route("/elections")
def elections():
    """Redirect to combined Elections, approval & protests page (preserves query params with election_ prefix)."""
    args = request.args.to_dict(flat=True)
    redirect_args = {}
    param_map = [
        ("country_code", "election_country_code"),
        ("region", "election_region"),
        ("date_from", "election_date_from"),
        ("date_to", "election_date_to"),
        ("status", "election_status"),
        ("election_type", "election_type"),
        ("search", "election_search"),
        ("sort", "election_sort"),
        ("page", "election_page"),
    ]
    for old_key, new_key in param_map:
        if old_key in args and args[old_key]:
            redirect_args[new_key] = args[old_key]
    return redirect(url_for("main.approval_protests", **redirect_args, _anchor="elections"))


@bp.route("/elections/export")
def elections_export():
    """Export election calendar as CSV (current filters)."""
    country_code = request.args.get("country_code", "").strip() or None
    region = request.args.get("region", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    status = request.args.get("status", "").strip() or None
    election_type = request.args.get("election_type", "").strip() or None
    search = request.args.get("search", "").strip() or None
    order_by = request.args.get("sort", "").strip() or None
    rows = get_election_calendar(
        country_code=country_code,
        region=region,
        date_from=date_from,
        date_to=date_to,
        status=status,
        election_type=election_type,
        search=search,
        limit=5000,
        order_by=order_by,
    )
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Country code", "Country", "Type", "Date planned", "Status", "Notes"])
    for r in rows:
        w.writerow([
            r.get("country_code") or "",
            r.get("country_name") or "",
            r.get("election_type") or "",
            (r.get("date_planned") or "")[:10],
            r.get("status") or "",
            (r.get("notes") or "")[:500],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=elections.csv"},
    )


@bp.route("/elections/add", methods=["GET", "POST"])
def election_add():
    """Add an election to the calendar."""
    if request.method == "POST":
        country_code = (request.form.get("country_code") or "").strip()
        country_name = (request.form.get("country_name") or "").strip()
        election_type = (request.form.get("election_type") or "").strip() or "other"
        date_planned = (request.form.get("date_planned") or "").strip()
        status = (request.form.get("status") or "").strip() or None
        notes = (request.form.get("notes") or "").strip() or None
        if country_code and country_name and date_planned:
            add_election(
                country_code=country_code,
                country_name=country_name,
                election_type=election_type,
                date_planned=date_planned,
                status=status,
                notes=notes,
            )
            flash("Election added to calendar.")
            return redirect(url_for("main.approval_protests", _anchor="elections"))
    countries = get_integration_countries(limit=300)
    return render_template(
        "election_add.html",
        election_types=ELECTION_TYPES,
        election_statuses=ELECTION_STATUSES,
        countries=countries,
    )


@bp.route("/elections/<int:election_id>")
def election_detail(election_id):
    """Election detail with link to country and Scenario Engine."""
    election = get_election(election_id)
    if not election:
        return "Election not found", 404
    scenario_url = url_for("main.scenario_engine") + "?event_type=election_upset&country=" + (election.get("country_name") or election.get("country_code") or "")
    return render_template(
        "election_detail.html",
        election=election,
        scenario_engine_url=scenario_url,
        stability_dashboard_url=url_for("main.stability_dashboard"),
    )


@bp.route("/elections/<int:election_id>/delete", methods=["POST"])
def election_delete(election_id):
    """Delete an election."""
    if delete_election(election_id):
        flash("Election removed from calendar.")
    else:
        flash("Election not found.", "error")
    return redirect(url_for("main.approval_protests", _anchor="elections"))


@bp.route("/elections.rss")
def elections_rss():
    """RSS feed of upcoming elections."""
    from datetime import datetime, timedelta
    from flask import make_response
    from email.utils import formatdate
    import time
    today = datetime.utcnow().strftime("%Y-%m")
    end_12 = (datetime.utcnow() + timedelta(days=365)).strftime("%Y-%m")
    elections_list = get_election_calendar(date_from=today, date_to=end_12, limit=50, order_by=None)
    now_rfc = formatdate(timeval=time.time(), localtime=False)
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        "<rss version=\"2.0\">",
        "<channel>",
        "<title>Upcoming Elections – Geopolitical Terminal</title>",
        "<link>" + request.host_url.rstrip("/") + url_for("main.approval_protests") + "#elections</link>",
        "<description>Upcoming elections by country and date.</description>",
        "<lastBuildDate>" + now_rfc + "</lastBuildDate>",
    ]
    for e in elections_list:
        title = ((e.get("country_name") or "") + " – " + (e.get("election_type") or "") + " " + (e.get("date_planned") or "")[:7]).replace("&", "&amp;").replace("<", "&lt;")[:100]
        desc = (e.get("notes") or "")[:300].replace("&", "&amp;").replace("<", "&lt;")
        date_str = (e.get("date_planned") or "")[:10]
        link = request.host_url.rstrip("/") + url_for("main.election_detail", election_id=e.get("id"))
        try:
            dt = datetime.strptime(date_str[:7], "%Y-%m")
            pub_rfc = formatdate(timeval=dt.timestamp(), localtime=False)
        except (ValueError, TypeError):
            pub_rfc = now_rfc
        xml.append("<item>")
        xml.append("<title>" + title + "</title>")
        xml.append("<description>" + desc + "</description>")
        xml.append("<pubDate>" + pub_rfc + "</pubDate>")
        xml.append("<link>" + link.replace("&", "&amp;") + "</link>")
        xml.append("<guid isPermaLink=\"true\">" + link.replace("&", "&amp;") + "</guid>")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


@bp.route("/elections.ics")
def elections_ical():
    """iCal feed of upcoming elections (next 12 months)."""
    from datetime import datetime, timedelta
    from flask import make_response
    today = datetime.utcnow().strftime("%Y-%m")
    end_12 = (datetime.utcnow() + timedelta(days=365)).strftime("%Y-%m")
    elections_list = get_election_calendar(date_from=today, date_to=end_12, limit=200, order_by=None)
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Geopolitical Terminal//Elections//EN",
        "CALSCALE:GREGORIAN",
        "X-WR-CALNAME:Upcoming Elections",
    ]
    for e in elections_list:
        dt = (e.get("date_planned") or "")[:10].replace("-", "")
        if len(dt) < 8:
            dt = (dt + "01")[:8]
        dt_end = dt[:8]
        if len(dt) >= 8:
            try:
                d = datetime.strptime(dt[:8], "%Y%m%d") + timedelta(days=1)
                dt_end = d.strftime("%Y%m%d")
            except (ValueError, TypeError):
                pass
        summary = ((e.get("country_name") or "") + " " + (e.get("election_type") or "")).strip() or "Election"
        summary = summary.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,")
        desc = (e.get("notes") or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,")
        link = request.host_url.rstrip("/") + url_for("main.election_detail", election_id=e.get("id"))
        uid = "election-" + str(e.get("id")) + "@geopolitical"
        lines.append("BEGIN:VEVENT")
        lines.append("UID:" + uid)
        lines.append("DTSTART;VALUE=DATE:" + dt[:8])
        lines.append("DTEND;VALUE=DATE:" + dt_end)
        lines.append("SUMMARY:" + summary[:200])
        if desc:
            lines.append("DESCRIPTION:" + desc[:500])
        lines.append("URL:" + link)
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    resp = make_response("\r\n".join(lines))
    resp.headers["Content-Type"] = "text/calendar; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=elections.ics"
    return resp


@bp.route("/elections/<int:election_id>/edit", methods=["GET", "POST"])
def election_edit(election_id):
    """Edit an election."""
    election = get_election(election_id)
    if not election:
        return "Election not found", 404
    if request.method == "POST":
        update_election(
            election_id,
            country_code=(request.form.get("country_code") or "").strip() or None,
            country_name=(request.form.get("country_name") or "").strip() or None,
            election_type=(request.form.get("election_type") or "").strip() or None,
            date_planned=(request.form.get("date_planned") or "").strip() or None,
            status=(request.form.get("status") or "").strip() or None,
            notes=(request.form.get("notes") or "").strip() or None,
        )
        flash("Election updated.")
        return redirect(url_for("main.approval_protests", _anchor="elections"))
    countries = get_integration_countries(limit=300)
    return render_template(
        "election_edit.html",
        election=election,
        election_types=ELECTION_TYPES,
        election_statuses=ELECTION_STATUSES,
        countries=countries,
    )


def _approval_protests_approval_params(request):
    """Extract approval filter params from request (prefix approval_)."""
    return {
        "country_code": request.args.get("approval_country_code", "").strip() or None,
        "region": request.args.get("approval_region", "").strip() or None,
        "date_from": request.args.get("approval_date_from", "").strip() or None,
        "date_to": request.args.get("approval_date_to", "").strip() or None,
        "subject": request.args.get("approval_subject", "").strip() or None,
        "min_approval": request.args.get("approval_min", type=float),
        "max_approval": request.args.get("approval_max", type=float),
        "source": request.args.get("approval_source", "").strip() or None,
        "page": max(1, int(request.args.get("approval_page") or 1)),
        "order_by": request.args.get("approval_sort", "").strip() or None,
        "per_page": _approval_protests_per_page(request.args.get("approval_per_page"), 15),
    }


def _approval_protests_protest_params(request):
    """Extract protest filter params from request (prefix protest_)."""
    return {
        "country_code": request.args.get("protest_country_code", "").strip() or None,
        "region": request.args.get("protest_region", "").strip() or None,
        "date_from": request.args.get("protest_date_from", "").strip() or None,
        "date_to": request.args.get("protest_date_to", "").strip() or None,
        "trigger_topic": request.args.get("protest_trigger", "").strip() or None,
        "search": request.args.get("protest_search", "").strip() or None,
        "page": max(1, int(request.args.get("protest_page") or 1)),
        "order_by": request.args.get("protest_sort", "").strip() or None,
        "per_page": _approval_protests_per_page(request.args.get("protest_per_page"), 15),
    }


def _approval_protests_per_page(val, default=15):
    v = int(val) if val is not None else default
    return min(max(v, 15), 50) if v in (15, 30, 50) else default


def _approval_protests_election_params(request):
    """Extract election filter params from request (prefix election_)."""
    per_page = 20
    return {
        "country_code": request.args.get("election_country_code", "").strip() or None,
        "region": request.args.get("election_region", "").strip() or None,
        "date_from": request.args.get("election_date_from", "").strip() or None,
        "date_to": request.args.get("election_date_to", "").strip() or None,
        "status": request.args.get("election_status", "").strip() or None,
        "election_type": request.args.get("election_type", "").strip() or None,
        "search": request.args.get("election_search", "").strip() or None,
        "order_by": request.args.get("election_sort", "").strip() or None,
        "page": max(1, int(request.args.get("election_page") or 1)),
        "per_page": per_page,
    }


@bp.route("/approval-protests")
def approval_protests():
    """Elections, approval ratings and protest tracking — combined screen with filters and pagination."""
    from datetime import datetime as _dt, timedelta
    _election_today_month = _dt.utcnow().strftime("%Y-%m")
    _election_end_12 = (_dt.utcnow() + timedelta(days=365)).strftime("%Y-%m")
    ap = _approval_protests_approval_params(request)
    pp = _approval_protests_protest_params(request)
    ep = _approval_protests_election_params(request)
    all_args = request.args.to_dict(flat=True)
    # Elections data for combined page
    elections_list = get_election_calendar(
        country_code=ep["country_code"],
        region=ep["region"],
        date_from=ep["date_from"],
        date_to=ep["date_to"],
        status=ep["status"],
        election_type=ep["election_type"],
        search=ep["search"],
        limit=ep["per_page"],
        offset=(ep["page"] - 1) * ep["per_page"],
        order_by=ep["order_by"],
    )
    election_total = get_election_calendar_count(
        country_code=ep["country_code"],
        region=ep["region"],
        date_from=ep["date_from"],
        date_to=ep["date_to"],
        status=ep["status"],
        election_type=ep["election_type"],
        search=ep["search"],
    )
    election_total_pages = max(1, (election_total + ep["per_page"] - 1) // ep["per_page"])
    all_args = request.args.to_dict(flat=True)
    election_query_base = {k: v for k, v in all_args.items() if k.startswith("election_")}
    election_query_prev = dict(all_args) if ep["page"] > 1 else None
    if election_query_prev:
        election_query_prev["election_page"] = ep["page"] - 1
    election_query_next = dict(all_args) if ep["page"] < election_total_pages else None
    if election_query_next:
        election_query_next["election_page"] = ep["page"] + 1
    upcoming_count = get_election_calendar_count(date_from=_election_today_month, status="upcoming")
    next_12_count = get_election_calendar_count(date_from=_election_today_month, date_to=_election_end_12)
    election_last_updated = get_election_calendar_last_updated()
    election_presets = [
        ("Upcoming only", {"election_date_from": _election_today_month, "election_status": "upcoming"}),
        ("Next 12 months", {"election_date_from": _election_today_month, "election_date_to": _election_end_12}),
        ("Presidential", {"election_type": "presidential"}),
    ]
    elections_export_query = {
        "country_code": ep["country_code"],
        "region": ep["region"],
        "date_from": ep["date_from"],
        "date_to": ep["date_to"],
        "status": ep["status"],
        "election_type": ep["election_type"],
        "search": ep["search"],
        "sort": ep["order_by"],
    }
    elections_export_url = url_for("main.elections_export", **{k: v for k, v in elections_export_query.items() if v is not None})
    per_approval = ap["per_page"]
    per_protest = pp["per_page"]
    low_threshold = request.args.get("low_approval_threshold", type=float) or 35
    approval_list = get_approval_ratings(
        country_code=ap["country_code"],
        region=ap["region"],
        date_from=ap["date_from"],
        date_to=ap["date_to"],
        subject=ap["subject"],
        min_approval=ap["min_approval"],
        max_approval=ap["max_approval"],
        source=ap["source"],
        limit=per_approval,
        offset=(ap["page"] - 1) * per_approval,
        order_by=ap["order_by"],
    )
    approval_total = get_approval_ratings_count(
        country_code=ap["country_code"],
        region=ap["region"],
        date_from=ap["date_from"],
        date_to=ap["date_to"],
        subject=ap["subject"],
        min_approval=ap["min_approval"],
        max_approval=ap["max_approval"],
        source=ap["source"],
    )
    protest_list = get_protest_tracking(
        country_code=pp["country_code"],
        region=pp["region"],
        date_from=pp["date_from"],
        date_to=pp["date_to"],
        trigger_topic=pp["trigger_topic"],
        search=pp["search"],
        limit=per_protest,
        offset=(pp["page"] - 1) * per_protest,
        order_by=pp["order_by"],
    )
    protest_total = get_protest_tracking_count(
        country_code=pp["country_code"],
        region=pp["region"],
        date_from=pp["date_from"],
        date_to=pp["date_to"],
        trigger_topic=pp["trigger_topic"],
        search=pp["search"],
    )
    approval_pages = max(1, (approval_total + per_approval - 1) // per_approval)
    protest_pages = max(1, (protest_total + per_protest - 1) // per_protest)
    today = datetime.utcnow().strftime("%Y-%m-%d")
    last_30 = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    regions = get_election_regions()
    all_args = request.args.to_dict(flat=True)
    approval_query = {k: v for k, v in all_args.items() if k.startswith("approval_")}
    protest_query = {k: v for k, v in all_args.items() if k.startswith("protest_")}
    approval_prev = dict(all_args) if ap["page"] > 1 else None
    if approval_prev:
        approval_prev["approval_page"] = ap["page"] - 1
    approval_next = dict(all_args) if ap["page"] < approval_pages else None
    if approval_next:
        approval_next["approval_page"] = ap["page"] + 1
    protest_prev = dict(all_args) if pp["page"] > 1 else None
    if protest_prev:
        protest_prev["protest_page"] = pp["page"] - 1
    protest_next = dict(all_args) if pp["page"] < protest_pages else None
    if protest_next:
        protest_next["protest_page"] = pp["page"] + 1
    low_approval_count = get_approval_ratings_count(
        country_code=ap["country_code"], region=ap["region"],
        date_from=ap["date_from"], date_to=ap["date_to"],
        subject=ap["subject"], max_approval=low_threshold, source=ap["source"],
    ) if approval_total else 0
    protests_last_30_count = get_protest_tracking_count(date_from=last_30, date_to=today)
    approval_timeseries = []
    if ap["country_code"]:
        approval_timeseries = get_approval_timeseries(
            country_codes=[ap["country_code"]],
            date_from=ap["date_from"] or last_30,
            date_to=ap["date_to"] or today,
            subject=ap["subject"],
            limit=100,
        )
    elif not ap["region"] and not ap["date_from"] and not ap["date_to"]:
        approval_timeseries = get_approval_timeseries(limit=50)
    approval_latest_bars = get_approval_latest_by_country(
        region=ap["region"], date_from=ap["date_from"], date_to=ap["date_to"],
        max_approval=ap["max_approval"], limit=15,
    )
    protest_counts_country = get_protest_counts_by_country(
        date_from=pp["date_from"] or last_30, date_to=pp["date_to"] or today, limit=15,
    )
    protest_counts_trigger = get_protest_counts_by_trigger(
        date_from=pp["date_from"] or last_30, date_to=pp["date_to"] or today, limit=10,
    )
    trigger_topics = get_protest_trigger_topics()
    subject_presets = ["President", "Government", "Parliament", "Prime Minister"]
    approval_sort_asc = dict(all_args)
    approval_sort_asc["approval_sort"] = "approval_asc"
    approval_sort_desc = dict(all_args)
    approval_sort_desc["approval_sort"] = "approval_desc"
    return render_template(
        "approval_protests.html",
        approval=approval_list,
        approval_total=approval_total,
        approval_page=ap["page"],
        approval_pages=approval_pages,
        approval_per_page=per_approval,
        approval_params=ap,
        approval_query_prev=approval_prev,
        approval_query_next=approval_next,
        approval_export_url=url_for("main.approval_protests_export_approval", **approval_query),
        protests=protest_list,
        protest_total=protest_total,
        protest_page=pp["page"],
        protest_pages=protest_pages,
        protest_per_page=per_protest,
        protest_params=pp,
        protest_query_prev=protest_prev,
        protest_query_next=protest_next,
        protest_export_url=url_for("main.approval_protests_export_protests", **protest_query),
        regions=regions,
        approval_last_updated=get_approval_last_updated(),
        protest_last_updated=get_protest_last_updated(),
        stability_dashboard_url=url_for("main.stability_dashboard"),
        situation_room_url=url_for("main.situation_room"),
        scenario_engine_url=url_for("main.scenario_engine"),
        approval_last_30_from=last_30,
        approval_last_30_to=today,
        protest_last_30_from=last_30,
        protest_last_30_to=today,
        clear_filters_url=url_for("main.approval_protests"),
        low_approval_threshold=low_threshold,
        low_approval_count=low_approval_count,
        protests_last_30_count=protests_last_30_count,
        approval_timeseries=approval_timeseries,
        approval_latest_bars=approval_latest_bars,
        protest_counts_country=protest_counts_country,
        protest_counts_trigger=protest_counts_trigger,
        trigger_topics=trigger_topics,
        subject_presets=subject_presets,
        return_to=request.args.get("return_to", "").strip() or None,
        approval_sort_asc=approval_sort_asc,
        approval_sort_desc=approval_sort_desc,
        elections=elections_list,
        election_total=election_total,
        election_page=ep["page"],
        election_total_pages=election_total_pages,
        election_per_page=ep["per_page"],
        election_query_base=election_query_base,
        election_query_prev=election_query_prev,
        election_query_next=election_query_next,
        election_country_code=ep["country_code"],
        election_region=ep["region"],
        election_date_from=ep["date_from"],
        election_date_to=ep["date_to"],
        election_status=ep["status"],
        election_type=ep["election_type"],
        election_search=ep["search"],
        election_order_by=ep["order_by"],
        election_types=ELECTION_TYPES,
        election_statuses=ELECTION_STATUSES,
        upcoming_count=upcoming_count,
        next_12_count=next_12_count,
        election_last_updated=election_last_updated,
        election_presets=election_presets,
        elections_export_url=elections_export_url,
        elections_rss_url=url_for("main.elections_rss"),
        elections_ical_url=url_for("main.elections_ical"),
    )


@bp.route("/approval-protests/export/approval")
def approval_protests_export_approval():
    """Export approval ratings as CSV or XLSX (current approval filters)."""
    ap = _approval_protests_approval_params(request)
    rows = get_approval_ratings(
        country_code=ap["country_code"],
        region=ap["region"],
        date_from=ap["date_from"],
        date_to=ap["date_to"],
        subject=ap["subject"],
        min_approval=ap["min_approval"],
        max_approval=ap["max_approval"],
        source=ap["source"],
        limit=5000,
        order_by=ap["order_by"],
    )
    date_suffix = ""
    if ap["date_from"] or ap["date_to"]:
        date_suffix = "_" + (ap["date_from"] or "") + "_" + (ap["date_to"] or "")
    fmt = (request.args.get("format") or "csv").strip().lower()
    if fmt == "xlsx":
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Approval"
            headers = ["Country code", "Country", "Subject", "Approval %", "Poll date", "Source", "Poll URL", "Sample size", "Created at"]
            ws.append(headers)
            for r in rows:
                ws.append([
                    r.get("country_code") or "",
                    r.get("country_name") or "",
                    r.get("subject") or "",
                    r.get("approval_pct") or "",
                    (r.get("poll_date") or "")[:10],
                    r.get("source") or "",
                    r.get("poll_url") or "",
                    r.get("sample_size"),
                    (r.get("created_at") or "")[:10],
                ])
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=approval_ratings{date_suffix}.xlsx"},
            )
        except Exception:
            pass
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Country code", "Country", "Subject", "Approval %", "Poll date", "Source", "Poll URL", "Sample size", "Created at"])
    for r in rows:
        w.writerow([
            r.get("country_code") or "",
            r.get("country_name") or "",
            r.get("subject") or "",
            r.get("approval_pct") or "",
            (r.get("poll_date") or "")[:10],
            r.get("source") or "",
            r.get("poll_url") or "",
            r.get("sample_size") or "",
            (r.get("created_at") or "")[:10],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=approval_ratings{date_suffix}.csv"},
    )


@bp.route("/approval-protests/export/protests")
def approval_protests_export_protests():
    """Export protest tracking as CSV or XLSX (current protest filters)."""
    pp = _approval_protests_protest_params(request)
    rows = get_protest_tracking(
        country_code=pp["country_code"],
        region=pp["region"],
        date_from=pp["date_from"],
        date_to=pp["date_to"],
        trigger_topic=pp["trigger_topic"],
        search=pp["search"],
        limit=5000,
        order_by=pp["order_by"],
    )
    date_suffix = ""
    if pp["date_from"] or pp["date_to"]:
        date_suffix = "_" + (pp["date_from"] or "") + "_" + (pp["date_to"] or "")
    fmt = (request.args.get("format") or "csv").strip().lower()
    if fmt == "xlsx":
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Protests"
            headers = ["Country code", "Country", "Event date", "Summary", "Size", "Trigger", "Location", "Severity", "Source URL", "Created at"]
            ws.append(headers)
            for r in rows:
                ws.append([
                    r.get("country_code") or "",
                    r.get("country_name") or "",
                    (r.get("event_date") or "")[:10],
                    (r.get("summary") or "")[:500],
                    r.get("estimated_size") or "",
                    r.get("trigger_topic") or "",
                    r.get("location") or "",
                    r.get("severity") or "",
                    r.get("source_url") or "",
                    (r.get("created_at") or "")[:10],
                ])
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=protests{date_suffix}.xlsx"},
            )
        except Exception:
            pass
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Country code", "Country", "Event date", "Summary", "Size", "Trigger", "Location", "Severity", "Source URL", "Created at"])
    for r in rows:
        w.writerow([
            r.get("country_code") or "",
            r.get("country_name") or "",
            (r.get("event_date") or "")[:10],
            (r.get("summary") or "")[:500],
            r.get("estimated_size") or "",
            r.get("trigger_topic") or "",
            r.get("location") or "",
            r.get("severity") or "",
            r.get("source_url") or "",
            (r.get("created_at") or "")[:10],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=protests{date_suffix}.csv"},
    )


@bp.route("/approval-protests/import", methods=["GET", "POST"])
def approval_protests_import():
    """Bulk import approval ratings or protests from CSV."""
    if request.method != "POST":
        return render_template("approval_protests_import.html")
    import csv
    from io import StringIO
    kind = (request.form.get("type") or "approval").strip().lower()
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please upload a CSV file.", "error")
        return redirect(url_for("main.approval_protests_import"))
    try:
        content = file.stream.read().decode("utf-8", errors="replace")
        reader = csv.DictReader(StringIO(content))
        rows = list(reader)
    except Exception as e:
        flash(f"Could not read CSV: {e}", "error")
        return redirect(url_for("main.approval_protests_import"))
    if kind == "approval":
        added, skipped = 0, 0
        for r in rows:
            cc = (r.get("country_code") or r.get("Country code") or "").strip()
            cn = (r.get("country_name") or r.get("Country") or r.get("Country name") or "").strip()
            subj = (r.get("subject") or r.get("Subject") or "").strip()
            pct = r.get("approval_pct") or r.get("Approval %") or r.get("Approval")
            try:
                pct = float(pct) if pct else None
            except (TypeError, ValueError):
                pct = None
            if not cc or not cn or not subj or pct is None:
                skipped += 1
                continue
            poll_date = (r.get("poll_date") or r.get("Poll date") or "").strip() or None
            source = (r.get("source") or r.get("Source") or "").strip() or None
            poll_url = (r.get("poll_url") or r.get("Poll URL") or "").strip() or None
            sample_size = r.get("sample_size") or r.get("Sample size")
            try:
                sample_size = int(sample_size) if sample_size else None
            except (TypeError, ValueError):
                sample_size = None
            if approval_duplicate_exists(cc, subj, poll_date, exclude_id=None):
                skipped += 1
                continue
            add_approval_rating(
                country_code=cc, country_name=cn, subject=subj, approval_pct=pct,
                poll_date=poll_date, source=source, poll_url=poll_url, sample_size=sample_size,
            )
            added += 1
        flash(f"Approval: {added} added, {skipped} skipped (duplicates or invalid rows).")
    else:
        added, skipped = 0, 0
        for r in rows:
            cc = (r.get("country_code") or r.get("Country code") or "").strip()
            cn = (r.get("country_name") or r.get("Country") or r.get("Country name") or "").strip()
            ed = (r.get("event_date") or r.get("Event date") or "").strip()
            if not cc or not cn or not ed:
                skipped += 1
                continue
            summary = (r.get("summary") or r.get("Summary") or "").strip() or None
            size = (r.get("estimated_size") or r.get("Size") or "").strip() or None
            trigger = (r.get("trigger_topic") or r.get("Trigger") or "").strip() or None
            location = (r.get("location") or r.get("Location") or "").strip() or None
            severity = (r.get("severity") or r.get("Severity") or "").strip() or None
            source_url = (r.get("source_url") or r.get("Source URL") or "").strip() or None
            add_protest_event(
                country_code=cc, country_name=cn, event_date=ed,
                summary=summary, estimated_size=size, trigger_topic=trigger,
                location=location, severity=severity, source_url=source_url,
            )
            added += 1
        flash(f"Protests: {added} added, {skipped} skipped (invalid rows).")
    return redirect(url_for("main.approval_protests"))


@bp.route("/approval-protests/approval/add", methods=["GET", "POST"])
def approval_add():
    if request.method == "POST":
        country_code = (request.form.get("country_code") or "").strip()
        country_name = (request.form.get("country_name") or "").strip()
        subject = (request.form.get("subject") or "").strip()
        approval_pct = request.form.get("approval_pct", type=float)
        poll_date = (request.form.get("poll_date") or "").strip() or None
        source = (request.form.get("source") or "").strip() or None
        poll_url = (request.form.get("poll_url") or "").strip() or None
        sample_size = request.form.get("sample_size", type=int)
        if country_code and country_name and subject and approval_pct is not None:
            if approval_duplicate_exists(country_code, subject, poll_date, exclude_id=None):
                flash("A rating for this country, subject, and poll date already exists. Add anyway was skipped.", "error")
            else:
                add_approval_rating(
                    country_code=country_code,
                    country_name=country_name,
                    subject=subject,
                    approval_pct=approval_pct,
                    poll_date=poll_date,
                    source=source,
                    poll_url=poll_url,
                    sample_size=sample_size,
                )
                flash("Approval rating added.")
                return redirect(url_for("main.approval_protests", approval_country_code=country_code or None))
    countries = get_integration_countries(limit=300)
    return render_template("approval_add.html", countries=countries or [])


@bp.route("/approval-protests/approval/<int:rating_id>/edit", methods=["GET", "POST"])
def approval_edit(rating_id):
    rating = get_approval_rating(rating_id)
    if not rating:
        return "Approval rating not found", 404
    if request.method == "POST":
        update_approval_rating(
            rating_id,
            country_code=(request.form.get("country_code") or "").strip() or None,
            country_name=(request.form.get("country_name") or "").strip() or None,
            subject=(request.form.get("subject") or "").strip() or None,
            approval_pct=request.form.get("approval_pct", type=float),
            poll_date=(request.form.get("poll_date") or "").strip() or None,
            source=(request.form.get("source") or "").strip() or None,
            poll_url=(request.form.get("poll_url") or "").strip() or None,
            sample_size=request.form.get("sample_size", type=int),
        )
        flash("Approval rating updated.")
        return redirect(url_for("main.approval_protests"))
    countries = get_integration_countries(limit=300)
    return render_template("approval_edit.html", rating=rating, countries=countries or [])


@bp.route("/approval-protests/approval/<int:rating_id>/delete", methods=["POST"])
def approval_delete(rating_id):
    if delete_approval_rating(rating_id):
        flash("Approval rating removed.")
    else:
        flash("Approval rating not found.", "error")
    return redirect(url_for("main.approval_protests"))


@bp.route("/approval-protests/protests/add", methods=["GET", "POST"])
def protest_add():
    if request.method == "POST":
        country_code = (request.form.get("country_code") or "").strip()
        country_name = (request.form.get("country_name") or "").strip()
        event_date = (request.form.get("event_date") or "").strip()
        summary = (request.form.get("summary") or "").strip() or None
        estimated_size = (request.form.get("estimated_size") or "").strip() or None
        trigger_topic = (request.form.get("trigger_topic") or "").strip() or None
        location = (request.form.get("location") or "").strip() or None
        severity = (request.form.get("severity") or "").strip() or None
        source_url = (request.form.get("source_url") or "").strip() or None
        if country_code and country_name and event_date:
            add_protest_event(
                country_code=country_code,
                country_name=country_name,
                event_date=event_date,
                summary=summary,
                estimated_size=estimated_size,
                trigger_topic=trigger_topic,
                location=location,
                severity=severity,
                source_url=source_url,
            )
            flash("Protest event added.")
            return redirect(url_for("main.approval_protests", protest_country_code=country_code or None))
    countries = get_integration_countries(limit=300)
    trigger_topics = get_protest_trigger_topics()
    return render_template("protest_add.html", countries=countries or [], trigger_topics=trigger_topics or [])


@bp.route("/approval-protests/protests/<int:protest_id>/edit", methods=["GET", "POST"])
def protest_edit(protest_id):
    protest = get_protest(protest_id)
    if not protest:
        return "Protest not found", 404
    if request.method == "POST":
        update_protest(
            protest_id,
            country_code=(request.form.get("country_code") or "").strip() or None,
            country_name=(request.form.get("country_name") or "").strip() or None,
            event_date=(request.form.get("event_date") or "").strip() or None,
            summary=(request.form.get("summary") or "").strip() or None,
            estimated_size=(request.form.get("estimated_size") or "").strip() or None,
            trigger_topic=(request.form.get("trigger_topic") or "").strip() or None,
            location=(request.form.get("location") or "").strip() or None,
            severity=(request.form.get("severity") or "").strip() or None,
            source_url=(request.form.get("source_url") or "").strip() or None,
        )
        flash("Protest event updated.")
        return redirect(url_for("main.approval_protests"))
    countries = get_integration_countries(limit=300)
    trigger_topics = get_protest_trigger_topics()
    return render_template("protest_edit.html", protest=protest, countries=countries or [], trigger_topics=trigger_topics or [])


@bp.route("/approval-protests/protests/<int:protest_id>/delete", methods=["POST"])
def protest_delete(protest_id):
    if delete_protest(protest_id):
        flash("Protest event removed.")
    else:
        flash("Protest not found.", "error")
    return redirect(url_for("main.approval_protests"))


# --- Conflict & military escalation monitor ---
CONFLICT_PRESETS = [
    ("East Asia", {"exercises_region": "East Asia", "movement_region": "East Asia", "naval_region": "East Asia", "incident_days": "365"}),
    ("NATO / Europe", {"exercises_region": "Europe", "naval_region": "Europe", "incidents_country": "", "incident_days": "365"}),
    ("Middle East / Gulf", {"exercises_region": "Middle East", "movement_region": "Middle East", "naval_region": "Middle East", "incident_days": "365"}),
]
CONFLICT_SOURCES = {
    "defense": "SIPRI / IISS (seed).",
    "exercises": "Seed / manual.",
    "incidents": "Seed / manual.",
    "movement": "Satellite/OSINT (seed).",
    "naval": "Seed / manual.",
    "arms": "SIPRI / UN Register (seed).",
}


def _conflict_filters():
    """Parse conflict dashboard filter and pagination params from request."""
    def _int(val, default=None):
        try:
            return int(val) if val not in (None, "") else default
        except (TypeError, ValueError):
            return default

    incident_days = _int(request.args.get("incident_days"), 365)
    per_page = 20

    return {
        "defense_country": (request.args.get("defense_country") or "").strip() or None,
        "defense_year_from": _int(request.args.get("defense_year_from")),
        "defense_year_to": _int(request.args.get("defense_year_to")),
        "exercises_region": (request.args.get("exercises_region") or "").strip() or None,
        "exercises_date_from": (request.args.get("exercises_date_from") or "").strip() or None,
        "exercises_date_to": (request.args.get("exercises_date_to") or "").strip() or None,
        "exercises_search": (request.args.get("exercises_search") or "").strip() or None,
        "incidents_country": (request.args.get("incidents_country") or "").strip() or None,
        "incidents_date_from": (request.args.get("incidents_date_from") or "").strip() or None,
        "incidents_date_to": (request.args.get("incidents_date_to") or "").strip() or None,
        "incidents_severity": (request.args.get("incidents_severity") or "").strip() or None,
        "movement_country": (request.args.get("movement_country") or "").strip() or None,
        "movement_region": (request.args.get("movement_region") or "").strip() or None,
        "movement_type": (request.args.get("movement_type") or "").strip() or None,
        "movement_date_from": (request.args.get("movement_date_from") or "").strip() or None,
        "movement_date_to": (request.args.get("movement_date_to") or "").strip() or None,
        "naval_region": (request.args.get("naval_region") or "").strip() or None,
        "naval_country": (request.args.get("naval_country") or "").strip() or None,
        "arms_supplier": (request.args.get("arms_supplier") or "").strip() or None,
        "arms_recipient": (request.args.get("arms_recipient") or "").strip() or None,
        "arms_year_from": _int(request.args.get("arms_year_from")),
        "arms_year_to": _int(request.args.get("arms_year_to")),
        "incident_days": incident_days,
        "per_page": per_page,
        "page_defense": max(1, _int(request.args.get("page_defense"), 1)),
        "page_exercises": max(1, _int(request.args.get("page_exercises"), 1)),
        "page_incidents": max(1, _int(request.args.get("page_incidents"), 1)),
        "page_movement": max(1, _int(request.args.get("page_movement"), 1)),
        "page_naval": max(1, _int(request.args.get("page_naval"), 1)),
        "page_arms": max(1, _int(request.args.get("page_arms"), 1)),
    }


@bp.route("/conflict")
def conflict_dashboard():
    """Conflict & military escalation: defense spending, exercises, border incidents, military movement, naval heat map, arms trade. Filters, pagination, export."""
    f = _conflict_filters()
    escalation = get_escalation_tree(incident_days=f["incident_days"])
    escalation_by_region = get_escalation_tree_by_region(incident_days=f["incident_days"])
    summary = get_conflict_summary()
    arms_summary = get_arms_trade_summary(top_n=5)
    country_name_to_code = _country_name_to_code_map()
    arms_chart_suppliers = arms_summary.get("top_suppliers", [])[:8]
    arms_chart_recipients = arms_summary.get("top_recipients", [])[:8]

    def _offset(page, per):
        return (page - 1) * per

    defense_spending = get_defense_spending(
        country_code=f["defense_country"],
        year_from=f["defense_year_from"],
        year_to=f["defense_year_to"],
        limit=f["per_page"],
        offset=_offset(f["page_defense"], f["per_page"]),
    )
    defense_spending = get_defense_spending_with_yoy(defense_spending)
    defense_total = get_defense_spending_count(
        country_code=f["defense_country"],
        year_from=f["defense_year_from"],
        year_to=f["defense_year_to"],
    )
    military_exercises = get_military_exercises(
        region=f["exercises_region"],
        date_from=f["exercises_date_from"],
        date_to=f["exercises_date_to"],
        search=f["exercises_search"],
        limit=f["per_page"],
        offset=_offset(f["page_exercises"], f["per_page"]),
    )
    exercises_total = get_military_exercises_count(
        region=f["exercises_region"],
        date_from=f["exercises_date_from"],
        date_to=f["exercises_date_to"],
        search=f["exercises_search"],
    )
    border_incidents = get_border_incidents(
        country_code=f["incidents_country"],
        date_from=f["incidents_date_from"],
        date_to=f["incidents_date_to"],
        severity=f["incidents_severity"],
        limit=f["per_page"],
        offset=_offset(f["page_incidents"], f["per_page"]),
    )
    incidents_total = get_border_incidents_count(
        country_code=f["incidents_country"],
        date_from=f["incidents_date_from"],
        date_to=f["incidents_date_to"],
        severity=f["incidents_severity"],
    )
    military_movement = get_military_movement(
        country_code=f["movement_country"],
        region=f["movement_region"],
        detection_type=f["movement_type"],
        date_from=f["movement_date_from"],
        date_to=f["movement_date_to"],
        limit=f["per_page"],
        offset=_offset(f["page_movement"], f["per_page"]),
    )
    movement_total = get_military_movement_count(
        country_code=f["movement_country"],
        region=f["movement_region"],
        detection_type=f["movement_type"],
        date_from=f["movement_date_from"],
        date_to=f["movement_date_to"],
    )
    naval_heat = get_naval_deployment_heat(region=f["naval_region"], limit_regions=15)
    naval_deployments = get_naval_deployments(
        region=f["naval_region"],
        country_code=f["naval_country"],
        limit=f["per_page"],
        offset=_offset(f["page_naval"], f["per_page"]),
    )
    naval_total = get_naval_deployments_count(region=f["naval_region"], country_code=f["naval_country"])
    arms_trade = get_arms_trade(
        supplier=f["arms_supplier"],
        recipient=f["arms_recipient"],
        year_from=f["arms_year_from"],
        year_to=f["arms_year_to"],
        limit=f["per_page"],
        offset=_offset(f["page_arms"], f["per_page"]),
    )
    arms_total = get_arms_trade_count(
        supplier=f["arms_supplier"],
        recipient=f["arms_recipient"],
        year_from=f["arms_year_from"],
        year_to=f["arms_year_to"],
    )

    def _total_pages(total, per):
        return max(1, (total + per - 1) // per)

    query_base = request.args.to_dict(flat=True) if request.args else {}
    per = f["per_page"]

    def _page_url(section_key, page_num):
        q = dict(query_base)
        q[section_key] = page_num
        return url_for("main.conflict_dashboard", **q)

    pagination = {
        "defense": {"prev": _page_url("page_defense", f["page_defense"] - 1) if f["page_defense"] > 1 else None, "next": _page_url("page_defense", f["page_defense"] + 1) if f["page_defense"] < _total_pages(defense_total, per) else None},
        "exercises": {"prev": _page_url("page_exercises", f["page_exercises"] - 1) if f["page_exercises"] > 1 else None, "next": _page_url("page_exercises", f["page_exercises"] + 1) if f["page_exercises"] < _total_pages(exercises_total, per) else None},
        "incidents": {"prev": _page_url("page_incidents", f["page_incidents"] - 1) if f["page_incidents"] > 1 else None, "next": _page_url("page_incidents", f["page_incidents"] + 1) if f["page_incidents"] < _total_pages(incidents_total, per) else None},
        "movement": {"prev": _page_url("page_movement", f["page_movement"] - 1) if f["page_movement"] > 1 else None, "next": _page_url("page_movement", f["page_movement"] + 1) if f["page_movement"] < _total_pages(movement_total, per) else None},
        "naval": {"prev": _page_url("page_naval", f["page_naval"] - 1) if f["page_naval"] > 1 else None, "next": _page_url("page_naval", f["page_naval"] + 1) if f["page_naval"] < _total_pages(naval_total, per) else None},
        "arms": {"prev": _page_url("page_arms", f["page_arms"] - 1) if f["page_arms"] > 1 else None, "next": _page_url("page_arms", f["page_arms"] + 1) if f["page_arms"] < _total_pages(arms_total, per) else None},
    }

    arms_chart_years = get_arms_trade_by_year(limit_years=15)
    military_capability_summary = get_military_capability_summary(limit=15)
    conflict_event_imports = get_conflict_event_imports(country_code=None, source=None, limit=30)
    return render_template(
        "conflict.html",
        escalation=escalation,
        escalation_by_region=escalation_by_region,
        defense_spending=defense_spending,
        defense_total=defense_total,
        defense_total_pages=_total_pages(defense_total, f["per_page"]),
        military_exercises=military_exercises,
        exercises_total=exercises_total,
        exercises_total_pages=_total_pages(exercises_total, f["per_page"]),
        border_incidents=border_incidents,
        incidents_total=incidents_total,
        incidents_total_pages=_total_pages(incidents_total, f["per_page"]),
        military_movement=military_movement,
        movement_total=movement_total,
        movement_total_pages=_total_pages(movement_total, f["per_page"]),
        naval_heat=naval_heat,
        naval_deployments=naval_deployments,
        naval_total=naval_total,
        naval_total_pages=_total_pages(naval_total, f["per_page"]),
        arms_trade=arms_trade,
        arms_total=arms_total,
        arms_total_pages=_total_pages(arms_total, f["per_page"]),
        conflict_summary=summary,
        arms_summary=arms_summary,
        arms_chart_suppliers=arms_chart_suppliers,
        arms_chart_recipients=arms_chart_recipients,
        arms_chart_years=arms_chart_years,
        section_sources=CONFLICT_SOURCES,
        presets=CONFLICT_PRESETS,
        country_name_to_code=country_name_to_code,
        filters=f,
        conflict_export_url=url_for("main.conflict_export", export_type="all"),
        conflict_movement_map_url=url_for(
            "main.conflict_movement_map",
            **{k: v for k, v in [
                ("region", f["movement_region"]), ("country_code", f["movement_country"]),
                ("date_from", f["movement_date_from"]), ("date_to", f["movement_date_to"]),
            ] if v}
        ),
        relationship_mapper_url=url_for("main.relationship_mapper") + "?mode=military",
        supply_chain_url=url_for("main.supply_chain_dashboard"),
        alerts_url=url_for("main.alerts_list"),
        scenario_engine_url=url_for("main.scenario_engine"),
        conflict_alerts_url=url_for("main.conflict_alerts"),
        conflict_compare_url=url_for("main.conflict_compare"),
        query_base=query_base,
        pagination=pagination,
        military_capability_summary=military_capability_summary,
        conflict_event_imports=conflict_event_imports,
    )


@bp.route("/conflict/export/<export_type>")
def conflict_export(export_type):
    """Export conflict data as CSV (section) or ZIP (all)."""
    import csv
    from io import StringIO, BytesIO

    f = _conflict_filters()
    limit = 2000

    def _csv_response(filename, rows, headers, row_fn):
        out = StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        for r in rows:
            w.writerow(row_fn(r))
        return Response(
            out.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if export_type == "defense":
        rows = get_defense_spending(
            country_code=f["defense_country"],
            year_from=f["defense_year_from"],
            year_to=f["defense_year_to"],
            limit=limit,
        )
        return _csv_response(
            "conflict_defense_spending.csv",
            rows,
            ["Country", "Country code", "Year", "Spending (USD bn)", "% GDP", "Source"],
            lambda r: [
                r.get("country_name") or "",
                r.get("country_code") or "",
                r.get("year") or "",
                "%.0f" % (r.get("spending_usd_billions") or 0),
                "%.1f" % (r.get("pct_gdp") or 0),
                r.get("source") or "",
            ],
        )
    if export_type == "exercises":
        rows = get_military_exercises(
            region=f["exercises_region"],
            date_from=f["exercises_date_from"],
            date_to=f["exercises_date_to"],
            search=f["exercises_search"],
            limit=limit,
        )
        return _csv_response(
            "conflict_military_exercises.csv",
            rows,
            ["Participants", "Name", "Region", "Start", "End", "Scale", "Description"],
            lambda r: [
                r.get("participants") or "",
                r.get("name") or "",
                r.get("region") or "",
                (r.get("start_date") or "")[:10],
                (r.get("end_date") or "")[:10],
                r.get("scale") or "",
                (r.get("description") or "")[:500],
            ],
        )
    if export_type == "incidents":
        rows = get_border_incidents(
            country_code=f["incidents_country"],
            date_from=f["incidents_date_from"],
            date_to=f["incidents_date_to"],
            severity=f["incidents_severity"],
            limit=limit,
        )
        return _csv_response(
            "conflict_border_incidents.csv",
            rows,
            ["Country A", "Code A", "Country B", "Code B", "Date", "Summary", "Severity", "Status"],
            lambda r: [
                r.get("country_a_name") or "",
                r.get("country_a_code") or "",
                r.get("country_b_name") or "",
                r.get("country_b_code") or "",
                (r.get("incident_date") or "")[:10],
                (r.get("summary") or "")[:500],
                r.get("severity") or "",
                r.get("status") or "",
            ],
        )
    if export_type == "movement":
        rows = get_military_movement(
            country_code=f["movement_country"],
            region=f["movement_region"],
            detection_type=f["movement_type"],
            date_from=f["movement_date_from"],
            date_to=f["movement_date_to"],
            limit=limit,
        )
        return _csv_response(
            "conflict_military_movement.csv",
            rows,
            ["Country", "Code", "Region", "Type", "Summary", "Observed", "Lat", "Lon"],
            lambda r: [
                r.get("country_name") or "",
                r.get("country_code") or "",
                r.get("region") or "",
                r.get("detection_type") or "",
                (r.get("summary") or "")[:500],
                (r.get("observed_date") or "")[:10],
                r.get("lat") or "",
                r.get("lon") or "",
            ],
        )
    if export_type == "naval":
        rows = get_naval_deployments(
            region=f["naval_region"],
            country_code=f["naval_country"],
            limit=limit,
        )
        return _csv_response(
            "conflict_naval_deployments.csv",
            rows,
            ["Country", "Code", "Region", "Vessel/description", "As of"],
            lambda r: [
                r.get("country_name") or "",
                r.get("country_code") or "",
                r.get("region") or "",
                r.get("vessel_description") or "",
                (r.get("as_of_date") or "")[:10],
            ],
        )
    if export_type == "arms":
        rows = get_arms_trade(
            supplier=f["arms_supplier"],
            recipient=f["arms_recipient"],
            year_from=f["arms_year_from"],
            year_to=f["arms_year_to"],
            limit=limit,
        )
        return _csv_response(
            "conflict_arms_trade.csv",
            rows,
            ["Supplier", "Recipient", "Weapon type", "Value (USD m)", "Year", "Status"],
            lambda r: [
                r.get("supplier_country") or "",
                r.get("recipient_country") or "",
                r.get("weapon_type") or "",
                "%.0f" % (r.get("value_usd_millions") or 0),
                r.get("year") or "",
                r.get("deal_status") or "",
            ],
        )
    if export_type == "all":
        import zipfile
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for section, etype in [
                ("defense", "defense"),
                ("exercises", "exercises"),
                ("incidents", "incidents"),
                ("movement", "movement"),
                ("naval", "naval"),
                ("arms", "arms"),
            ]:
                if etype == "defense":
                    rows = get_defense_spending(
                        country_code=f["defense_country"],
                        year_from=f["defense_year_from"],
                        year_to=f["defense_year_to"],
                        limit=limit,
                    )
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Country", "Country code", "Year", "Spending (USD bn)", "% GDP", "Source"])
                    for r in rows:
                        w.writerow([r.get("country_name") or "", r.get("country_code") or "", r.get("year") or "", "%.0f" % (r.get("spending_usd_billions") or 0), "%.1f" % (r.get("pct_gdp") or 0), r.get("source") or ""])
                elif etype == "exercises":
                    rows = get_military_exercises(region=f["exercises_region"], date_from=f["exercises_date_from"], date_to=f["exercises_date_to"], search=f["exercises_search"], limit=limit)
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Participants", "Name", "Region", "Start", "End", "Scale", "Description"])
                    for r in rows:
                        w.writerow([r.get("participants") or "", r.get("name") or "", r.get("region") or "", (r.get("start_date") or "")[:10], (r.get("end_date") or "")[:10], r.get("scale") or "", (r.get("description") or "")[:500]])
                elif etype == "incidents":
                    rows = get_border_incidents(country_code=f["incidents_country"], date_from=f["incidents_date_from"], date_to=f["incidents_date_to"], severity=f["incidents_severity"], limit=limit)
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Country A", "Code A", "Country B", "Code B", "Date", "Summary", "Severity", "Status"])
                    for r in rows:
                        w.writerow([r.get("country_a_name") or "", r.get("country_a_code") or "", r.get("country_b_name") or "", r.get("country_b_code") or "", (r.get("incident_date") or "")[:10], (r.get("summary") or "")[:500], r.get("severity") or "", r.get("status") or ""])
                elif etype == "movement":
                    rows = get_military_movement(country_code=f["movement_country"], region=f["movement_region"], detection_type=f["movement_type"], date_from=f["movement_date_from"], date_to=f["movement_date_to"], limit=limit)
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Country", "Code", "Region", "Type", "Summary", "Observed", "Lat", "Lon"])
                    for r in rows:
                        w.writerow([r.get("country_name") or "", r.get("country_code") or "", r.get("region") or "", r.get("detection_type") or "", (r.get("summary") or "")[:500], (r.get("observed_date") or "")[:10], r.get("lat") or "", r.get("lon") or ""])
                elif etype == "naval":
                    rows = get_naval_deployments(region=f["naval_region"], country_code=f["naval_country"], limit=limit)
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Country", "Code", "Region", "Vessel/description", "As of"])
                    for r in rows:
                        w.writerow([r.get("country_name") or "", r.get("country_code") or "", r.get("region") or "", r.get("vessel_description") or "", (r.get("as_of_date") or "")[:10]])
                else:
                    rows = get_arms_trade(supplier=f["arms_supplier"], recipient=f["arms_recipient"], year_from=f["arms_year_from"], year_to=f["arms_year_to"], limit=limit)
                    so = StringIO()
                    w = csv.writer(so)
                    w.writerow(["Supplier", "Recipient", "Weapon type", "Value (USD m)", "Year", "Status"])
                    for r in rows:
                        w.writerow([r.get("supplier_country") or "", r.get("recipient_country") or "", r.get("weapon_type") or "", "%.0f" % (r.get("value_usd_millions") or 0), r.get("year") or "", r.get("deal_status") or ""])
                zf.writestr(f"conflict_{section}.csv", so.getvalue())
        buf.seek(0)
        return Response(buf.read(), mimetype="application/zip", headers={"Content-Disposition": "attachment; filename=conflict_export.zip"})
    if export_type == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            f = _conflict_filters()
            limit = 2000
            wb = Workbook()
            wb.remove(wb.active)
            def _sheet_defense(ws):
                rows = get_defense_spending(country_code=f["defense_country"], year_from=f["defense_year_from"], year_to=f["defense_year_to"], limit=limit)
                ws.append(["Country", "Country code", "Year", "Spending (USD bn)", "% GDP", "Source"])
                for r in rows:
                    ws.append([r.get("country_name") or "", r.get("country_code") or "", r.get("year") or "", (r.get("spending_usd_billions") or 0), (r.get("pct_gdp") or 0), r.get("source") or ""])
            def _sheet_exercises(ws):
                rows = get_military_exercises(region=f["exercises_region"], date_from=f["exercises_date_from"], date_to=f["exercises_date_to"], search=f["exercises_search"], limit=limit)
                ws.append(["Participants", "Name", "Region", "Start", "End", "Scale", "Description"])
                for r in rows:
                    ws.append([r.get("participants") or "", r.get("name") or "", r.get("region") or "", (r.get("start_date") or "")[:10], (r.get("end_date") or "")[:10], r.get("scale") or "", (r.get("description") or "")[:500]])
            def _sheet_incidents(ws):
                rows = get_border_incidents(country_code=f["incidents_country"], date_from=f["incidents_date_from"], date_to=f["incidents_date_to"], severity=f["incidents_severity"], limit=limit)
                ws.append(["Country A", "Code A", "Country B", "Code B", "Date", "Summary", "Severity", "Status"])
                for r in rows:
                    ws.append([r.get("country_a_name") or "", r.get("country_a_code") or "", r.get("country_b_name") or "", r.get("country_b_code") or "", (r.get("incident_date") or "")[:10], (r.get("summary") or "")[:500], r.get("severity") or "", r.get("status") or ""])
            def _sheet_movement(ws):
                rows = get_military_movement(country_code=f["movement_country"], region=f["movement_region"], detection_type=f["movement_type"], date_from=f["movement_date_from"], date_to=f["movement_date_to"], limit=limit)
                ws.append(["Country", "Code", "Region", "Type", "Summary", "Observed", "Lat", "Lon"])
                for r in rows:
                    ws.append([r.get("country_name") or "", r.get("country_code") or "", r.get("region") or "", r.get("detection_type") or "", (r.get("summary") or "")[:500], (r.get("observed_date") or "")[:10], r.get("lat") or "", r.get("lon") or ""])
            def _sheet_naval(ws):
                rows = get_naval_deployments(region=f["naval_region"], country_code=f["naval_country"], limit=limit)
                ws.append(["Country", "Code", "Region", "Vessel/description", "As of"])
                for r in rows:
                    ws.append([r.get("country_name") or "", r.get("country_code") or "", r.get("region") or "", r.get("vessel_description") or "", (r.get("as_of_date") or "")[:10]])
            def _sheet_arms(ws):
                rows = get_arms_trade(supplier=f["arms_supplier"], recipient=f["arms_recipient"], year_from=f["arms_year_from"], year_to=f["arms_year_to"], limit=limit)
                ws.append(["Supplier", "Recipient", "Weapon type", "Value (USD m)", "Year", "Status"])
                for r in rows:
                    ws.append([r.get("supplier_country") or "", r.get("recipient_country") or "", r.get("weapon_type") or "", (r.get("value_usd_millions") or 0), r.get("year") or "", r.get("deal_status") or ""])
            for title, fn in [("Defense", _sheet_defense), ("Exercises", _sheet_exercises), ("Incidents", _sheet_incidents), ("Movement", _sheet_movement), ("Naval", _sheet_naval), ("Arms", _sheet_arms)]:
                ws = wb.create_sheet(title=title[:31])
                fn(ws)
            from io import BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(buf.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=conflict_export.xlsx"})
        except Exception as e:
            return redirect(url_for("main.conflict_dashboard") + "?export_error=excel")
    return redirect(url_for("main.conflict_dashboard"))


@bp.route("/conflict/movement-map")
def conflict_movement_map():
    """Map view of satellite-detected military movement (lat/lon). Optional date filter and clustering."""
    region = (request.args.get("region") or "").strip() or None
    country_code = (request.args.get("country_code") or "").strip() or None
    date_from = (request.args.get("date_from") or "").strip() or None
    date_to = (request.args.get("date_to") or "").strip() or None
    movements = get_military_movement(region=region, country_code=country_code, date_from=date_from, date_to=date_to, limit=500)
    points = [m for m in movements if m.get("lat") is not None and m.get("lon") is not None]
    return render_template(
        "conflict_movement_map.html",
        movements=points,
        filter_region=region,
        filter_country=country_code,
        filter_date_from=date_from,
        filter_date_to=date_to,
    )


@bp.route("/conflict/alerts", methods=["GET", "POST"])
def conflict_alerts():
    """Conflict-specific alert rules: notify when new border incidents, exercises, or military movement (by region/country)."""
    if request.method == "POST" and request.form.get("form_type") == "conflict_alert":
        name = (request.form.get("name") or "").strip()
        event_types = request.form.getlist("event_types") or [request.form.get("event_types") or ""]
        event_types = [x.strip() for x in event_types if x.strip()]
        if not event_types:
            event_types = ["border_incident", "military_exercise", "military_movement"]
        region = (request.form.get("region") or "").strip() or None
        country_code = (request.form.get("country_code") or "").strip() or None
        webhook_url = (request.form.get("webhook_url") or "").strip() or None
        if name:
            add_conflict_alert_rule(name, event_types, region=region, country_code=country_code, webhook_url=webhook_url)
        return redirect(url_for("main.conflict_alerts"))
    rules = get_conflict_alert_rules()
    return render_template(
        "conflict_alerts.html",
        rules=rules,
        conflict_dashboard_url=url_for("main.conflict_dashboard"),
    )


@bp.route("/conflict/alerts/<int:rule_id>/delete", methods=["POST"])
def conflict_alert_delete(rule_id):
    delete_conflict_alert_rule(rule_id)
    return redirect(url_for("main.conflict_alerts"))


@bp.route("/conflict/compare")
def conflict_compare():
    """Compare two countries on military/security: defense spending, incidents, exercises, arms."""
    country_a = (request.args.get("country_a") or "").strip() or None
    country_b = (request.args.get("country_b") or "").strip() or None
    countries = get_integration_countries(limit=300)
    country_list = [{"code": c.get("country_code"), "name": c.get("country_name")} for c in countries if c.get("country_code")]
    data_a = data_b = None
    name_a = name_b = None
    if country_a:
        c = next((x for x in countries if (x.get("country_code") or "").upper() == (country_a or "").upper()), None)
        name_a = (c.get("country_name") or country_a) if c else country_a
        data_a = {
            "defense": get_defense_spending(country_code=country_a, limit=15),
            "incidents": get_border_incidents(country_code=country_a, limit=15),
            "exercises": get_military_exercises(region=c.get("region"), limit=15) if c and c.get("region") else [],
            "movement": get_military_movement(country_code=country_a, limit=15),
            "naval": get_naval_deployments(country_code=country_a, limit=15),
            "arms_supplier": get_arms_trade(supplier=name_a, limit=15),
            "arms_recipient": get_arms_trade(recipient=name_a, limit=15),
            "region": c.get("region") if c else None,
        }
    if country_b:
        c = next((x for x in countries if (x.get("country_code") or "").upper() == (country_b or "").upper()), None)
        name_b = (c.get("country_name") or country_b) if c else country_b
        data_b = {
            "defense": get_defense_spending(country_code=country_b, limit=15),
            "incidents": get_border_incidents(country_code=country_b, limit=15),
            "exercises": get_military_exercises(region=c.get("region"), limit=15) if c and c.get("region") else [],
            "movement": get_military_movement(country_code=country_b, limit=15),
            "naval": get_naval_deployments(country_code=country_b, limit=15),
            "arms_supplier": get_arms_trade(supplier=name_b, limit=15),
            "arms_recipient": get_arms_trade(recipient=name_b, limit=15),
            "region": c.get("region") if c else None,
        }
    return render_template(
        "conflict_compare.html",
        country_list=country_list,
        country_a=country_a,
        country_b=country_b,
        data_a=data_a,
        data_b=data_b,
        country_name_to_code=_country_name_to_code_map(),
    )


@bp.route("/conflict/exercises.rss")
def conflict_exercises_rss():
    """RSS feed of military exercises (optionally filtered by region)."""
    region = (request.args.get("region") or "").strip() or None
    exercises = get_military_exercises(region=region, limit=50)
    from flask import make_response
    xml = ['<?xml version="1.0" encoding="UTF-8"?>', '<rss version="2.0">', "<channel>", "<title>Military exercises – Geopolitical Terminal</title>", "<link>" + request.host_url.rstrip("/") + url_for("main.conflict_dashboard") + "</link>", "<description>Military exercises (seed/manual data)</description>"]
    for e in exercises:
        title = (e.get("name") or "Exercise")[:100]
        desc = (e.get("description") or "")[:500] or (e.get("participants") or "")
        date_str = (e.get("start_date") or "")[:10]
        xml.append("<item>")
        xml.append("<title>" + title.replace("&", "&amp;").replace("<", "&lt;") + "</title>")
        xml.append("<description>" + (desc or "").replace("&", "&amp;").replace("<", "&lt;") + "</description>")
        xml.append("<pubDate>" + date_str + "</pubDate>")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


@bp.route("/conflict/exercises.ics")
def conflict_exercises_ics():
    """iCal feed of military exercises (optionally filtered by region)."""
    region = (request.args.get("region") or "").strip() or None
    exercises = get_military_exercises(region=region, limit=200)
    lines = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Geopolitical Terminal//Conflict//EN", "CALSCALE:GREGORIAN"]
    for e in exercises:
        uid = "exercise-" + str(e.get("id", "")) + "@geopolitical"
        start = (e.get("start_date") or "")[:10].replace("-", "")
        end = (e.get("end_date") or e.get("start_date") or "")[:10].replace("-", "")
        if not end:
            end = start
        summary = (e.get("name") or "Exercise").replace("\r", "").replace("\n", " ")
        desc = (e.get("description") or e.get("participants") or "").replace("\r", "").replace("\n", " ")
        lines.append("BEGIN:VEVENT")
        lines.append("UID:" + uid)
        lines.append("DTSTART;VALUE=DATE:" + start)
        lines.append("DTEND;VALUE=DATE:" + end)
        lines.append("SUMMARY:" + summary[:250])
        if desc:
            lines.append("DESCRIPTION:" + desc[:500])
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    resp = Response("\r\n".join(lines), mimetype="text/calendar; charset=utf-8")
    resp.headers["Content-Disposition"] = "attachment; filename=conflict_exercises.ics"
    return resp


# --- Sanctions, export controls & regulatory watch ---
def _country_name_to_code_map():
    """Build mapping country_name -> country_code for sanctions table links."""
    name_to_code = {}
    for c in get_integration_countries(limit=300):
        name = (c.get("country_name") or "").strip()
        code = (c.get("country_code") or "").strip()
        if name and code:
            name_to_code[name.upper()] = code
            name_to_code[name] = code
    return name_to_code


@bp.route("/sanctions-watch")
def sanctions_watch_dashboard():
    """Redirect to merged sanctions registry under Diplomacy & Agreements."""
    return redirect(url_for("main.diplomacy_sanctions", **request.args.to_dict(flat=True)))


@bp.route("/sanctions-watch/export/<export_type>")
def sanctions_watch_export(export_type):
    """Export sanctions, entity list, or export restrictions as CSV."""
    import csv
    from io import StringIO
    source = (request.args.get("source") or "").strip() or None
    target = (request.args.get("target") or "").strip() or None
    imposing = (request.args.get("imposing") or "").strip() or None
    date_from = (request.args.get("date_from") or "").strip() or None
    date_to = (request.args.get("date_to") or "").strip() or None
    search = (request.args.get("search") or "").strip() or None

    if export_type == "sanctions":
        rows = get_sanctions(source=source, target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search, limit=1000)
        out = StringIO()
        w = csv.writer(out)
        w.writerow(["Imposing", "Target", "Measure type", "Description", "Start date", "End date", "Source", "Source URL"])
        for r in rows:
            w.writerow([
                r.get("imposing_country") or "",
                r.get("target_country") or "",
                r.get("measure_type") or "",
                (r.get("description") or "")[:500],
                (r.get("start_date") or "")[:10],
                (r.get("end_date") or "")[:10],
                r.get("source") or "",
                r.get("source_url") or "",
            ])
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=sanctions_export.csv"})
    if export_type == "entities":
        entity_search = (request.args.get("entity_search") or "").strip() or None
        source_filter = (request.args.get("source") or "").strip() or None
        rows = get_entity_list_alerts(source=source_filter, search=entity_search, limit=1000)
        out = StringIO()
        w = csv.writer(out)
        w.writerow(["Source", "Entity name", "Entity type", "Country", "List name", "Listed date", "Summary"])
        for r in rows:
            w.writerow([
                r.get("source") or "",
                r.get("entity_name") or "",
                r.get("entity_type") or "",
                r.get("country") or "",
                r.get("list_name") or "",
                (r.get("listed_date") or "")[:10],
                (r.get("summary") or "")[:500],
            ])
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=entity_list_export.csv"})
    if export_type == "restrictions":
        export_search = (request.args.get("export_search") or "").strip() or None
        issuer = (request.args.get("issuer") or "").strip() or None
        restriction_type = (request.args.get("restriction_type") or "").strip() or None
        rows = get_export_restrictions(issuer=issuer, restriction_type=restriction_type, search=export_search, limit=1000)
        out = StringIO()
        w = csv.writer(out)
        w.writerow(["Issuer", "Type", "Title", "Description", "Effective date", "Source URL"])
        for r in rows:
            w.writerow([
                r.get("issuer") or "",
                r.get("restriction_type") or "",
                r.get("title") or "",
                (r.get("description") or "")[:500],
                (r.get("effective_date") or "")[:10],
                r.get("source_url") or "",
            ])
        return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=export_restrictions.csv"})
    if export_type == "all":
        import zipfile
        from io import BytesIO
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            s_rows = get_sanctions(source=source, target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search, limit=2000)
            so = StringIO()
            w = csv.writer(so)
            w.writerow(["Imposing", "Target", "Measure type", "Description", "Start date", "End date", "Source", "Source URL"])
            for r in s_rows:
                w.writerow([r.get("imposing_country") or "", r.get("target_country") or "", r.get("measure_type") or "", (r.get("description") or "")[:500], (r.get("start_date") or "")[:10], (r.get("end_date") or "")[:10], r.get("source") or "", r.get("source_url") or ""])
            zf.writestr("sanctions.csv", so.getvalue())
            e_search = (request.args.get("entity_search") or "").strip() or None
            e_src = (request.args.get("entity_source") or request.args.get("source") or "").strip() or None
            e_rows = get_entity_list_alerts(source=e_src, search=e_search, limit=2000)
            eo = StringIO()
            w = csv.writer(eo)
            w.writerow(["Source", "Entity name", "Entity type", "Country", "List name", "Listed date", "Summary"])
            for r in e_rows:
                w.writerow([r.get("source") or "", r.get("entity_name") or "", r.get("entity_type") or "", r.get("country") or "", r.get("list_name") or "", (r.get("listed_date") or "")[:10], (r.get("summary") or "")[:500]])
            zf.writestr("entity_list.csv", eo.getvalue())
            ex_search = (request.args.get("export_search") or "").strip() or None
            ex_issuer = (request.args.get("export_issuer") or request.args.get("issuer") or "").strip() or None
            ex_type = (request.args.get("export_type") or request.args.get("restriction_type") or "").strip() or None
            ex_rows = get_export_restrictions(issuer=ex_issuer, restriction_type=ex_type, search=ex_search, limit=2000)
            xo = StringIO()
            w = csv.writer(xo)
            w.writerow(["Issuer", "Type", "Title", "Description", "Effective date", "Source URL"])
            for r in ex_rows:
                w.writerow([r.get("issuer") or "", r.get("restriction_type") or "", r.get("title") or "", (r.get("description") or "")[:500], (r.get("effective_date") or "")[:10], r.get("source_url") or ""])
            zf.writestr("export_restrictions.csv", xo.getvalue())
        buf.seek(0)
        return Response(buf.read(), mimetype="application/zip", headers={"Content-Disposition": "attachment; filename=sanctions_watch_export.zip"})
    return redirect(url_for("main.diplomacy_sanctions"))


@bp.route("/sanctions-watch/supply-chain-check", methods=["GET", "POST"])
def sanctions_watch_supply_chain_check():
    """Which of my supply chains violate new export rules? Supports textarea and CSV file upload."""
    result = None
    history = session.get("sanctions_check_history") or []
    if request.method == "POST":
        entities = []
        raw = request.form.get("entities", "") or ""
        entities.extend([line.strip() for line in raw.splitlines() if line.strip()])
        if "entities_file" in request.files:
            f = request.files["entities_file"]
            if f and f.filename and f.filename.lower().endswith(".csv"):
                try:
                    import csv as csv_mod
                    stream = StringIO(f.stream.read().decode("utf-8", errors="replace"))
                    reader = csv_mod.reader(stream)
                    for row in reader:
                        if row and row[0].strip():
                            entities.append(row[0].strip())
                except Exception:
                    pass
        entities = list(dict.fromkeys([e for e in entities if e]))
        result = check_supply_chain_export_rules(entities)
        history = [{"total": result["total_checked"], "matches": len(result["matches"])}] + history[:9]
        session["sanctions_check_history"] = history
    return render_template("sanctions_watch_supply_chain_check.html", result=result, history=history)


# --- Scenario planning engine (MCDA / Delphi) ---
def _scenario_type_to_engine_event(scenario_type):
    """Map library scenario_type to Scenario Engine event_type for Open in Engine prefill."""
    if not scenario_type:
        return "election_upset"
    st = (scenario_type or "").lower()
    if "taiwan" in st or "crisis" in st and "currency" not in st:
        return "military_incursion"
    if "currency" in st or "stress" in st:
        return "financial_crisis"
    if "energy" in st or "gulf" in st:
        return "trade_embargo"
    if "defence" in st or "defense" in st:
        return "election_upset"
    return "election_upset"


@bp.route("/scenarios")
def scenarios_list():
    """Scenario planning: Taiwan crisis 2028, ASEAN currency stress, probability-weighted runs, 10-year outlook."""
    scenarios = get_scenarios(limit=20)
    for s in scenarios:
        s["engine_event_type"] = _scenario_type_to_engine_event(s.get("scenario_type"))
    return render_template("scenarios.html", scenarios=scenarios)


@bp.route("/scenarios/engine", methods=["GET", "POST"])
def scenario_engine():
    """Signature Scenario Engine: define event (left), multi-agent simulation (middle), probability paths + export (right)."""
    event_types = list(SCENARIO_ENGINE_EVENT_TYPES)
    result = None
    if request.method == "POST":
        event_type = (request.form.get("event_type") or "").strip() or "election_upset"
        region = (request.form.get("region") or "").strip()
        country = (request.form.get("country") or "").strip()
        horizon_year = request.form.get("horizon_year", type=int) or None
        run_name = (request.form.get("run_name") or "").strip()
        result = run_scenario_engine(event_type, region=region, country=country)
        if horizon_year:
            result["horizon_year"] = horizon_year
        if run_name:
            result["run_name"] = run_name
        run_id = add_scenario_engine_run(result)
        result["engine_run_id"] = run_id
        session["scenario_engine_last_run"] = result
        return redirect(url_for("main.scenario_engine", run_id=run_id))
    # GET: optional prefill from query (copy-link or Open in Engine)
    result = session.get("scenario_engine_last_run")
    prefill = {
        "event_type": request.args.get("event_type", "").strip() or (result.get("event_type") if result else "election_upset"),
        "region": request.args.get("region", "").strip() or (result.get("region") if result else ""),
        "country": request.args.get("country", "").strip() or (result.get("country") if result else ""),
    }
    if not result and (prefill["region"] or prefill["country"] or prefill["event_type"] != "election_upset"):
        # Build a minimal result for display if we have prefill but no prior run
        result = None
    recent_runs = get_scenario_engine_runs(limit=5)
    # If run_id requested, load that run as result for display and prefill form
    run_id_param = request.args.get("run_id", type=int)
    if run_id_param:
        loaded = get_scenario_engine_run(run_id_param)
        if loaded:
            result = loaded
            result["engine_run_id"] = run_id_param
            prefill["event_type"] = loaded.get("event_type") or prefill["event_type"]
            prefill["region"] = loaded.get("region") or prefill["region"]
            prefill["country"] = loaded.get("country") or prefill["country"]
            prefill["horizon_year"] = loaded.get("horizon_year") or prefill.get("horizon_year")
    prefill.setdefault("horizon_year", result.get("horizon_year") if result else None)
    return render_template(
        "scenario_engine.html",
        event_types=event_types,
        result=result,
        prefill=prefill,
        recent_runs=recent_runs,
    )


@bp.route("/scenarios/engine/export")
def scenario_engine_export():
    """Export Scenario Engine run as policy memo, risk briefing, investor note, PDF, or executive summary (PPT-ready). Optional run_id= to export a past run."""
    fmt = (request.args.get("format") or "policy_memo").strip().lower()
    run_id = request.args.get("run_id", type=int)
    result = None
    if run_id:
        result = get_scenario_engine_run(run_id)
    if not result:
        result = session.get("scenario_engine_last_run")
    if not result:
        flash("No run found. Run a simulation first, or choose a run from Recent runs.")
        return redirect(url_for("main.scenario_engine"))
    content, filename, mimetype = generate_scenario_engine_export(result, fmt)
    return Response(
        content,
        mimetype=mimetype,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@bp.route("/scenarios/engine/summary")
def scenario_engine_summary():
    """Printable summary of a Scenario Engine run (Print → Save as PDF). Optional run_id= for past runs."""
    run_id = request.args.get("run_id", type=int)
    result = None
    if run_id:
        result = get_scenario_engine_run(run_id)
    if not result:
        result = session.get("scenario_engine_last_run")
    if not result:
        flash("No run found. Run a simulation first.")
        return redirect(url_for("main.scenario_engine"))
    return render_template("scenario_engine_summary.html", result=result)


@bp.route("/scenarios/engine/save-as-scenario", methods=["POST"])
def scenario_engine_save_as_scenario():
    """Save the current Scenario Engine run as a new scenario in the library."""
    result = session.get("scenario_engine_last_run")
    if not result:
        flash("No run to save. Run a simulation first.")
        return redirect(url_for("main.scenario_engine"))
    name = (request.form.get("name") or "").strip() or (result.get("event_label") or "Engine run")
    scenario_id = add_scenario_from_engine_run(name, result)
    flash(f"Saved as scenario: {name}. You can run it from the Scenario library.")
    return redirect(url_for("main.scenario_detail", scenario_id=scenario_id))


@bp.route("/scenarios/engine/delete/<int:run_id>", methods=["POST"])
def scenario_engine_delete_run(run_id):
    """Delete a Scenario Engine run."""
    if delete_scenario_engine_run(run_id):
        flash("Run deleted.")
    return redirect(url_for("main.scenario_engine"))


@bp.route("/scenarios/engine/update-name/<int:run_id>", methods=["POST"])
def scenario_engine_update_name(run_id):
    """Update name and/or notes for a run."""
    name = (request.form.get("name") or "").strip()
    notes = (request.form.get("notes") or "").strip()
    update_scenario_engine_run_name_notes(run_id, name=name, notes=notes)
    flash("Run updated.")
    return redirect(url_for("main.scenario_engine", run_id=run_id))


@bp.route("/scenarios/engine/history")
def scenario_engine_history():
    """List all runs with optional filters and pagination."""
    page = max(1, request.args.get("page", type=int) or 1)
    per_page = min(50, max(10, request.args.get("per_page", type=int) or 20))
    event_type = (request.args.get("event_type") or "").strip() or None
    region = (request.args.get("region") or "").strip() or None
    country = (request.args.get("country") or "").strip() or None
    date_from = (request.args.get("date_from") or "").strip() or None
    date_to = (request.args.get("date_to") or "").strip() or None
    offset = (page - 1) * per_page
    runs, total = get_scenario_engine_runs_filtered(
        limit=per_page,
        offset=offset,
        event_type=event_type,
        region=region,
        country=country,
        date_from=date_from,
        date_to=date_to,
    )
    total_pages = (total + per_page - 1) // per_page if total else 0
    event_types = list(SCENARIO_ENGINE_EVENT_TYPES)
    return render_template(
        "scenario_engine_history.html",
        runs=runs,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        event_types=event_types,
        filters={"event_type": event_type, "region": region, "country": country, "date_from": date_from, "date_to": date_to},
    )


@bp.route("/scenarios/engine/compare")
def scenario_engine_compare():
    """Side-by-side comparison of two runs. Query: run_a=, run_b=."""
    run_a_id = request.args.get("run_a", type=int)
    run_b_id = request.args.get("run_b", type=int)
    run_a = get_scenario_engine_run(run_a_id) if run_a_id else None
    run_b = get_scenario_engine_run(run_b_id) if run_b_id else None
    if not run_a and not run_b:
        flash("Select two runs to compare (run_a= and run_b=).")
        return redirect(url_for("main.scenario_engine_history"))
    recent_runs = get_scenario_engine_runs(limit=20)
    return render_template(
        "scenario_engine_compare.html",
        run_a=run_a,
        run_b=run_b,
        run_a_id=run_a_id,
        run_b_id=run_b_id,
        recent_runs=recent_runs,
    )


@bp.route("/scenarios/engine/export-all")
def scenario_engine_export_all():
    """Page with links to download all formats for a run, or ZIP. Optional run_id=."""
    run_id = request.args.get("run_id", type=int)
    result = None
    if run_id:
        result = get_scenario_engine_run(run_id)
    if not result:
        result = session.get("scenario_engine_last_run")
        run_id = result.get("engine_run_id") if result else None
    if not result:
        flash("No run found. Run a simulation first or pass run_id=.")
        return redirect(url_for("main.scenario_engine"))
    if not run_id and result:
        run_id = result.get("engine_run_id")
    return render_template(
        "scenario_engine_export_all.html",
        result=result,
        run_id=run_id,
    )


@bp.route("/scenarios/<int:scenario_id>")
def scenario_detail(scenario_id):
    scenario = get_scenario(scenario_id)
    if not scenario:
        return "Scenario not found", 404
    runs = get_scenario_runs(scenario_id, limit=10)
    scenario["engine_event_type"] = _scenario_type_to_engine_event(scenario.get("scenario_type"))
    return render_template("scenario_detail.html", scenario=scenario, runs=runs)


@bp.route("/scenarios/<int:scenario_id>/run", methods=["POST"])
def scenario_run(scenario_id):
    run_id = run_scenario_simulation(scenario_id)
    if not run_id:
        return "Scenario not found", 404
    return redirect(url_for("main.scenario_run_result", run_id=run_id))


@bp.route("/scenarios/run/<int:run_id>")
def scenario_run_result(run_id):
    run = get_scenario_run(run_id)
    if not run:
        return "Run not found", 404
    scenario = get_scenario(run["scenario_id"])
    return render_template("scenario_run_result.html", run=run, scenario=scenario or {})


def _outlook_markdown_to_html(md: str, toc_slugs: list = None) -> str:
    """Render markdown to HTML for the outlook report. If toc_slugs given, inject id into h2/h3 for ToC."""
    try:
        import markdown
        import re
        html = markdown.markdown(md or "", extensions=["extra", "nl2br"])
        if toc_slugs:
            idx = [0]
            def add_id(m):
                if idx[0] < len(toc_slugs):
                    sid = toc_slugs[idx[0]]
                    idx[0] += 1
                    return m.group(0).replace(">", f' id="{sid}">', 1)
                return m.group(0)
            html = re.sub(r"<h[23]\s*>", add_id, html)
        return html
    except Exception:
        return "<pre>" + (md or "").replace("<", "&lt;").replace(">", "&gt;") + "</pre>"


def _outlook_toc_from_markdown(md: str) -> list:
    """Extract ## and ### heading lines for table of contents. Returns list of (level, text, slug)."""
    toc = []
    for line in (md or "").splitlines():
        s = line.strip()
        if s.startswith("##") and not s.startswith("###"):
            text = s.lstrip("#").strip()
            slug = "".join(c if c.isalnum() or c in " -" else "" for c in text).replace(" ", "-").strip("-").lower() or "h"
            toc.append((2, text, slug))
        elif s.startswith("###"):
            text = s.lstrip("#").strip()
            slug = "".join(c if c.isalnum() or c in " -" else "" for c in text).replace(" ", "-").strip("-").lower() or "h"
            toc.append((3, text, slug))
    return toc


@bp.route("/scenarios/outlook")
def scenarios_outlook():
    """10-year risk outlook report. Optional ?scenario_id= for scenario-specific outlook."""
    scenario_id = request.args.get("scenario_id", type=int) or None
    horizon = request.args.get("years", type=int) or 10
    outlook = generate_risk_outlook(scenario_id=scenario_id, horizon_years=min(30, max(5, horizon)))
    report_md = outlook.get("report_markdown") or ""
    toc_list = _outlook_toc_from_markdown(report_md)
    toc_slugs = [t[2] for t in toc_list]
    outlook["report_html"] = _outlook_markdown_to_html(report_md, toc_slugs=toc_slugs)
    outlook["report_toc"] = toc_list
    outlook["report_markdown"] = report_md
    download_md_url = url_for("main.scenarios_outlook_download_md", scenario_id=scenario_id, years=min(30, max(5, horizon)))
    return render_template("scenarios_outlook.html", outlook=outlook, scenarios=get_scenarios(limit=20), download_md_url=download_md_url)


@bp.route("/scenarios/outlook/download.md")
def scenarios_outlook_download_md():
    """Download the current outlook report as .md (query params: scenario_id, years)."""
    scenario_id = request.args.get("scenario_id", type=int) or None
    horizon = request.args.get("years", type=int) or 10
    outlook = generate_risk_outlook(scenario_id=scenario_id, horizon_years=min(30, max(5, horizon)))
    report_md = outlook.get("report_markdown") or ""
    from flask import Response
    return Response(report_md, mimetype="text/markdown", headers={"Content-Disposition": "attachment; filename=outlook-report.md"})


# --- Economic-geopolitical integration layer ---
@bp.route("/command")
def command_search():
    """Command bar: search by type (country, region, treaty, risk, scenario, sanctions, keyword). Redirect to appropriate screen."""
    q = (request.args.get("q") or "").strip()
    search_type = (request.args.get("type") or "keyword").strip().lower()
    if not q:
        return redirect(url_for("main.situation_room"))
    if search_type == "country":
        countries = get_integration_countries(limit=200)
        q_upper, q_lower = q.upper(), q.lower()
        for c in countries:
            if c.get("country_code", "").upper() == q_upper or (c.get("country_name") or "").lower() == q_lower:
                return redirect(url_for("main.integration_country", country_code=c["country_code"]))
            if q_lower in (c.get("country_name") or "").lower():
                return redirect(url_for("main.integration_country", country_code=c["country_code"]))
        return redirect(url_for("main.integration_dashboard", region=q))
    if search_type == "region":
        return redirect(url_for("main.integration_dashboard", region=q))
    if search_type == "sector":
        return redirect(url_for("main.risk_dashboard"))
    if search_type == "treaty":
        return redirect(url_for("main.diplomacy_treaties", party=q))
    if search_type == "risk":
        return redirect(url_for("main.risk_dashboard"))
    if search_type == "scenario":
        scenarios = get_scenarios(limit=50)
        q_lower = q.lower()
        for s in scenarios:
            if q_lower in (s.get("name") or "").lower():
                return redirect(url_for("main.scenario_detail", scenario_id=s["id"]))
        return redirect(url_for("main.scenarios_list"))
    if search_type == "sanctions":
        return redirect(url_for("main.sanctions_watch_supply_chain_check") + "?entities=" + quote(q, safe=""))
    # Professional shortcuts: open workspace with preset (CHN GOV, SGP TRADE, ASEAN RISK, TWN ESC)
    q_upper = q.upper().replace(" ", "")
    workspace_presets = {
        "CHNGOV": url_for("main.workspace", layout="4", t1="/country/CN", t2="/supply-chain", t3="/risk", t4="/country/TW"),
        "SGPTRADE": url_for("main.workspace", layout="4", t1="/country/SG", t2="/supply-chain", t3="/risk", t4="/country/MY"),
        "ASEANRISK": url_for("main.workspace", layout="4", t1="/risk", t2="/country/ID", t3="/country/TH", t4="/country/VN"),
        "TWNESC": url_for("main.workspace", layout="4", t1="/country/TW", t2="/scenarios/engine", t3="/risk", t4="/supply-chain"),
    }
    if q_upper in workspace_presets:
        return redirect(workspace_presets[q_upper])
    return redirect(url_for("main.search", q=q))


def _workspace_normalize_path(path: str) -> str:
    if not path or not path.strip():
        return ""
    p = path.strip()
    if not p.startswith("http") and not p.startswith("/"):
        return "/" + p
    return p


def _workspace_path_from_query(key: str) -> str:
    """Return panel path only if `key` is present in the query string (no default pages)."""
    if key not in request.args:
        return ""
    return _workspace_normalize_path(request.args.get(key, ""))


def get_workspace_panel_groups():
    """Grouped (path, label) pairs for workspace panel picker — users choose any screen per tile."""
    groups = [
        (
            "Home & intelligence",
            [
                ("/", "Situation room"),
                ("/feed", "Intelligence feed"),
                ("/insights", "Insights"),
                ("/alerts", "Watchlists"),
                ("/saved-views", "Saved views"),
                ("/live", "Live streams"),
                ("/timeline", "Timeline"),
                ("/search", "Search"),
            ],
        ),
        (
            "Risk, trade & supply",
            [
                ("/risk", "Risk index"),
                ("/supply-chain", "Supply chain stress"),
                ("/supply-chain/scenario", "Supply chain scenario"),
            ],
        ),
        (
            "Geography",
            [
                ("/integration", "Countries & regions"),
                ("/relationship-mapper", "Relationship Mapper"),
                ("/map", "Map"),
                ("/world-monitor", "World monitor"),
            ],
        ),
        (
            "Security & conflict",
            [
                ("/conflict", "Military & security"),
                ("/conflict/compare", "Conflict compare"),
                ("/conflict/movement-map", "Movement map"),
            ],
        ),
        (
            "Diplomacy & sanctions",
            [
                ("/diplomacy", "Diplomacy hub"),
                ("/diplomacy/treaties", "Treaties & agreements"),
                ("/diplomacy/sanctions", "Sanctions"),
                ("/diplomacy/sanctions/global", "Sanctions — global"),
                ("/diplomacy/alignment", "UN voting & alignment"),
                ("/diplomacy/compare", "Diplomacy compare"),
                ("/diplomacy/escalation", "Treaty escalation"),
                ("/diplomacy/legislative", "Legislative tracker"),
                ("/sanctions-watch", "Sanctions watch"),
                ("/sanctions-watch/supply-chain-check", "Sanctions supply-chain check"),
            ],
        ),
        (
            "Politics & stability",
            [
                ("/stability", "Political stability"),
                ("/elections", "Elections calendar"),
                ("/approval-protests", "Approval & protests"),
            ],
        ),
        (
            "Scenarios & reports",
            [
                ("/scenarios", "Scenario library"),
                ("/scenarios/engine", "Scenario Engine"),
                ("/scenarios/outlook", "10-year outlook"),
                ("/reports", "Reports"),
                ("/export/briefing", "Briefing builder"),
                ("/workspace", "Workspace (nested)"),
                ("/command", "Command search"),
            ],
        ),
        (
            "Indicators & synthesis",
            [
                ("/indicators", "Indicators"),
                ("/indicators/legislative", "Legislative indicators"),
                ("/synthesis", "Synthesis hub"),
                ("/synthesis/thematic", "Thematic synthesis"),
                ("/synthesis/divergences", "Divergences"),
                ("/synthesis/executive", "Executive synthesis"),
                ("/synthesis/trajectory", "Trajectory"),
                ("/synthesis/meta-briefing", "Meta briefing"),
            ],
        ),
        (
            "Help",
            [
                ("/help", "Help"),
            ],
        ),
    ]
    if getattr(current_user, "is_authenticated", False):
        groups.insert(
            5,
            (
                "Messaging",
                [
                    ("/messaging", "Intelligence messaging"),
                ],
            ),
        )
    countries = get_integration_countries(limit=300) or []
    country_items = []
    for c in countries:
        code = (c.get("country_code") or "").strip()
        if not code:
            continue
        path = f"/country/{code}"
        label = f"{c.get('country_name') or code} — {code}"
        country_items.append((path, label))
    country_items.sort(key=lambda x: x[1].lower())
    groups.append(("Country dashboards", country_items))
    return groups


def _workspace_panel_label(path: str) -> str:
    """Return a short friendly label for a workspace panel URL."""
    if not path:
        return "Panel"
    p = path.strip()
    if p in ("/", ""):
        return "Home"
    if p.startswith("/country/"):
        code = p.replace("/country/", "").strip().upper().split("?")[0]
        from app.country_data import ISO3_TO_2, ALL_COUNTRIES
        rev = {v: k for k, v in (ISO3_TO_2 or {}).items()}
        code3 = code if len(code) == 3 else rev.get(code, code)
        for iso3, name, *_ in (ALL_COUNTRIES or []):
            if iso3 == code3:
                return name
        return code3 or code
    labels = {
        "/": "Situation room",
        "/feed": "Intelligence feed",
        "/supply-chain": "Supply chain",
        "/risk": "Risk",
        "/scenarios/engine": "Scenario Engine",
        "/integration": "Countries & regions",
        "/relationship-mapper": "Relationship Mapper",
        "/live": "Live streams",
        "/scenarios": "Scenario library",
        "/scenarios/outlook": "10-year outlook",
        "/reports": "Reports",
        "/insights": "Insights",
        "/alerts": "Watchlists",
        "/conflict": "Military & security",
        "/diplomacy": "Diplomacy",
        "/diplomacy/sanctions": "Sanctions",
        "/stability": "Political stability",
        "/workspace": "Workspace",
    }
    return labels.get(p.split("?")[0], p[:30] + ("…" if len(p) > 30 else ""))


@bp.route("/workspace")
def workspace():
    """Multi-screen: user picks any page per panel (t1–t4). Optional quick examples; ?name= title. Copy link to share."""
    any_panel_in_query = any(k in request.args for k in ("t1", "t2", "t3", "t4"))
    layout = request.args.get("layout")
    if layout is None:
        # First visit /workspace with no panels: single empty tile; otherwise default to quad layout
        layout = "1" if not any_panel_in_query else "4"
    if layout not in ("1", "2", "3", "4"):
        layout = "4"
    t1 = _workspace_path_from_query("t1")
    t2 = _workspace_path_from_query("t2")
    t3 = _workspace_path_from_query("t3")
    t4 = _workspace_path_from_query("t4")
    workspace_name = (request.args.get("name") or request.args.get("title") or "").strip() or None
    presets = [
        ("Example: CHN quad", url_for("main.workspace", layout="4", t1="/country/CN", t2="/supply-chain", t3="/risk", t4="/country/TW")),
        ("Example: SGP trade", url_for("main.workspace", layout="4", t1="/country/SG", t2="/supply-chain", t3="/risk", t4="/country/MY")),
        ("Example: ASEAN risk", url_for("main.workspace", layout="4", t1="/risk", t2="/country/ID", t3="/country/TH", t4="/country/VN")),
        ("Example: TWN esc.", url_for("main.workspace", layout="4", t1="/country/TW", t2="/scenarios/engine", t3="/risk", t4="/supply-chain")),
        ("1 panel + home", url_for("main.workspace", layout="1", t1="/")),
        ("2 side-by-side", url_for("main.workspace", layout="2", t1="/country/CN", t2="/supply-chain")),
        ("Demo: 4-up 2×2 same", url_for("main.workspace", layout="4", t1="/feed", t2="/risk", t3="/feed", t4="/risk")),
    ]
    panel_groups = get_workspace_panel_groups()
    panel_label_1 = _workspace_panel_label(t1)
    panel_label_2 = _workspace_panel_label(t2)
    panel_label_3 = _workspace_panel_label(t3)
    panel_label_4 = _workspace_panel_label(t4)
    return render_template(
        "workspace.html",
        layout=layout,
        t1=t1,
        t2=t2,
        t3=t3,
        t4=t4,
        workspace_name=workspace_name,
        presets=presets,
        panel_groups=panel_groups,
        panel_label_1=panel_label_1,
        panel_label_2=panel_label_2,
        panel_label_3=panel_label_3,
        panel_label_4=panel_label_4,
    )


@bp.route("/export/excel")
def export_workspace_excel():
    """Export key data to Excel: risk index, integration countries, and current workspace state (from query params)."""
    try:
        import xlsxwriter
    except ImportError:
        return redirect(url_for("main.workspace"))
    risk_rows = get_risk_index()
    countries = get_integration_countries(limit=200)
    layout = request.args.get("layout", "")
    t1 = request.args.get("t1", "")
    t2 = request.args.get("t2", "")
    t3 = request.args.get("t3", "")
    t4 = request.args.get("t4", "")
    ws_name = (request.args.get("name") or request.args.get("title") or "").strip()
    base = request.url_root.rstrip("/")
    workspace_url = base + url_for("main.workspace") + "?layout=" + (layout or "4") + "&t1=" + (t1 or "") + "&t2=" + (t2 or "") + "&t3=" + (t3 or "") + "&t4=" + (t4 or "")
    if ws_name:
        workspace_url += "&name=" + ws_name.replace(" ", "+")
    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws_workspace = wb.add_worksheet("Workspace")
    ws_workspace.write(0, 0, "Layout")
    ws_workspace.write(0, 1, layout or "—")
    ws_workspace.write(1, 0, "Name")
    ws_workspace.write(1, 1, ws_name or "—")
    ws_workspace.write(2, 0, "Panel 1")
    ws_workspace.write(2, 1, t1 or "—")
    ws_workspace.write(3, 0, "Panel 2")
    ws_workspace.write(3, 1, t2 or "—")
    ws_workspace.write(4, 0, "Panel 3")
    ws_workspace.write(4, 1, t3 or "—")
    ws_workspace.write(5, 0, "Panel 4")
    ws_workspace.write(5, 1, t4 or "—")
    ws_workspace.write(6, 0, "Permalink")
    ws_workspace.write(6, 1, workspace_url)
    ws_risk = wb.add_worksheet("Risk Index")
    headers = ["Region", "Coup %", "Sanctions %", "Trade disruption %", "Updated"]
    for col, h in enumerate(headers):
        ws_risk.write(0, col, h)
    for row, r in enumerate(risk_rows or [], 1):
        ws_risk.write(row, 0, r.get("region_code") or "")
        ws_risk.write(row, 1, r.get("coup_likelihood_pct"))
        ws_risk.write(row, 2, r.get("sanctions_probability_pct"))
        ws_risk.write(row, 3, r.get("trade_disruption_pct"))
        ws_risk.write(row, 4, (r.get("updated_at") or "")[:19])
    ws_countries = wb.add_worksheet("Countries")
    for col, h in enumerate(["Country", "Region", "Geo fragility", "Economic fragility"]):
        ws_countries.write(0, col, h)
    for row, c in enumerate(countries or [], 1):
        ws_countries.write(row, 0, c.get("country_name") or c.get("country_code") or "")
        ws_countries.write(row, 1, c.get("region") or "")
        ws_countries.write(row, 2, c.get("geopolitical_fragility_score"))
        ws_countries.write(row, 3, c.get("economic_fragility_score"))
    wb.close()
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="geopolitical-workspace-export.xlsx"'},
    )


@bp.route("/relationship-mapper")
def relationship_mapper():
    """Relationship Mapper: country-to-country alignment, trade, voting blocs, defense pacts. Toggle economic/military/diplomatic/sanctions."""
    return render_template("relationship_mapper.html")


REGION_ORDER = [
    "East Asia", "South Asia", "Southeast Asia", "Central Asia", "Middle East",
    "Africa", "Caribbean", "Latin America", "Europe", "North America", "Oceania",
]


def _build_code_to_blocs(blocs_dict, p5_list):
    """Build mapping: country_code -> list of bloc keys the country belongs to."""
    from collections import defaultdict
    code_to_blocs = defaultdict(list)
    for key, members in list(blocs_dict.items()) + [("P5", p5_list)]:
        for code in members:
            code_to_blocs[code].append(key)
    return dict(code_to_blocs)


@bp.route("/integration")
def integration_dashboard():
    """Trade flows, debt distress, capital flight, reserves, FX, energy exposure. Click country → fragility scores."""
    region = request.args.get("region", "").strip() or None
    sort = request.args.get("sort", "").strip() or None
    order = request.args.get("order", "asc").strip().lower()
    if order not in ("asc", "desc"):
        order = "asc"
    if not sort:
        sort = "combined_systemic_risk_score"
        order = "desc"
    q = request.args.get("q", "").strip() or None
    cols = request.args.get("cols", "").strip().lower()
    show_extra_cols = cols == "extended" or cols == "1"
    risk_filter = request.args.get("risk", "").strip().lower() or "all"
    if risk_filter not in ("all", "high", "medium", "med_high"):
        risk_filter = "all"
    per_page_param = request.args.get("per_page", "").strip().lower()
    per_page = 50
    if per_page_param in ("50", "100", "200"):
        per_page = int(per_page_param)
    elif per_page_param == "all":
        per_page = 0
    page = max(1, int(request.args.get("page", "1") or 1))

    # Single fetch; when no region, use same list for region_counts (before q/risk filter)
    countries_raw = get_integration_countries(region=region, limit=500, sort=sort, order=order)
    if not region:
        region_counts = dict(Counter((c.get("region") or "—") for c in countries_raw))
        region_counts_ordered = [(r, region_counts[r]) for r in REGION_ORDER if r in region_counts]
        for r, cnt in region_counts.items():
            if r not in REGION_ORDER:
                region_counts_ordered.append((r, cnt))
    else:
        region_counts = {}
        region_counts_ordered = []

    countries = countries_raw
    if q:
        q_lower = q.lower()
        q_upper = q.upper()
        countries = [
            c for c in countries
            if q_lower in ((c.get("country_name") or "").lower())
            or q_upper in ((c.get("country_code") or "").upper())
        ]
    if risk_filter == "high":
        countries = [c for c in countries if (c.get("combined_systemic_risk_score") or 0) >= 60]
    elif risk_filter == "medium":
        countries = [c for c in countries if 40 <= (c.get("combined_systemic_risk_score") or 0) < 60]
    elif risk_filter == "med_high":
        countries = [c for c in countries if (c.get("combined_systemic_risk_score") or 0) >= 40]

    total_count = len(countries)
    risk_high = sum(1 for c in countries if (c.get("combined_systemic_risk_score") or 0) >= 60)
    risk_med = sum(1 for c in countries if 40 <= (c.get("combined_systemic_risk_score") or 0) < 60)
    risk_low = sum(1 for c in countries if (c.get("combined_systemic_risk_score") or 0) < 40)

    latest_updated = None
    for c in countries:
        u = c.get("updated_at")
        if u and (not latest_updated or (u > latest_updated)):
            latest_updated = u
    now_cutoff = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
    for c in countries:
        c["_updated_recently"] = bool(c.get("updated_at") and (c.get("updated_at") or "")[:10] >= now_cutoff)

    # Pagination: slice for display
    if per_page > 0:
        total_pages = max(1, (total_count + per_page - 1) // per_page)
        page = min(page, total_pages)
        start = (page - 1) * per_page
        countries_paged = countries[start : start + per_page]
    else:
        total_pages = 1
        countries_paged = countries

    blocs_for_display = []
    blocs_grouped = []
    bloc_cohesion_by_name = {}
    code_to_blocs = {}
    try:
        from app.un_votes.blocs import BLOCS, P5, BLOC_DISPLAY_NAMES, BLOC_CATEGORIES
        from app.un_votes.country_map import get_iso3_to_canonical_name
        iso3_to_name = get_iso3_to_canonical_name()
        try:
            from app.un_votes.readers import get_bloc_cohesion
            cohesion_rows = get_bloc_cohesion(limit=50)
            for r in cohesion_rows:
                bn = (r.get("bloc_name") or "").strip()
                if bn:
                    bloc_cohesion_by_name[bn] = r.get("cohesion_score")
        except Exception:
            pass
        for code, blocs in _build_code_to_blocs(BLOCS, P5).items():
            code_to_blocs[code] = blocs
        for cat_name, keys in BLOC_CATEGORIES.items():
            cat_blocs = []
            for key in keys:
                members = BLOCS.get(key) or (P5 if key == "P5" else [])
                display = BLOC_DISPLAY_NAMES.get(key, key)
                members_with_names = [(m, iso3_to_name.get(m, m)) for m in members]
                cohesion = bloc_cohesion_by_name.get(key.upper()) or bloc_cohesion_by_name.get(key)
                cat_blocs.append({
                    "key": key,
                    "name": display,
                    "member_count": len(members),
                    "members": members_with_names,
                    "cohesion_pct": round(cohesion, 0) if cohesion is not None else None,
                })
                blocs_for_display.append({"key": key, "name": display, "member_count": len(members), "members": members_with_names, "cohesion_pct": round(cohesion, 0) if cohesion is not None else None})
            if cat_blocs:
                blocs_grouped.append({"category": cat_name, "blocs": cat_blocs})
    except ImportError:
        blocs_for_display = []
        blocs_grouped = []

    avg_risk = round(sum(c.get("combined_systemic_risk_score") or 0 for c in countries) / max(len(countries), 1), 1)
    top_regions_by_risk = []
    if not region and region_counts_ordered:
        region_avg = {}
        for c in countries_raw:
            reg = c.get("region") or "—"
            if reg not in region_avg:
                region_avg[reg] = []
            region_avg[reg].append(c.get("combined_systemic_risk_score") or 0)
        for reg, scores in region_avg.items():
            top_regions_by_risk.append((reg, round(sum(scores) / len(scores), 1)))
        top_regions_by_risk.sort(key=lambda x: -x[1])
        top_regions_by_risk = top_regions_by_risk[:5]

    has_active_filters = bool(region or q or risk_filter != "all")
    macro_stress_alerts = get_macroeconomic_stress_alerts(threshold_debt=80, threshold_inflation=15, limit=12)
    capital_flows_summary = get_capital_flows_summary(limit=12)

    return render_template(
        "integration.html",
        countries=countries_paged,
        countries_all=countries,
        selected_region=region,
        sort=sort,
        order=order,
        search_q=q,
        show_extra_cols=show_extra_cols,
        total_count=total_count,
        risk_high=risk_high,
        risk_med=risk_med,
        risk_low=risk_low,
        latest_updated=latest_updated,
        region_counts_ordered=region_counts_ordered,
        risk_filter=risk_filter,
        per_page=per_page,
        page=page,
        total_pages=total_pages,
        blocs=blocs_for_display,
        blocs_grouped=blocs_grouped,
        code_to_blocs=code_to_blocs,
        avg_risk=avg_risk,
        top_regions_by_risk=top_regions_by_risk,
        has_active_filters=has_active_filters,
        macro_stress_alerts=macro_stress_alerts,
        capital_flows_summary=capital_flows_summary,
    )


def _integration_export_countries(region, sort, order, q, risk_filter):
    """Shared helper: fetch and filter countries for export (same logic as dashboard)."""
    countries = get_integration_countries(region=region, limit=500, sort=sort, order=order)
    if q:
        q_upper = q.upper()
        countries = [
            c for c in countries
            if q.lower() in ((c.get("country_name") or "").lower())
            or q_upper in ((c.get("country_code") or "").upper())
        ]
    if risk_filter == "high":
        countries = [c for c in countries if (c.get("combined_systemic_risk_score") or 0) >= 60]
    elif risk_filter == "medium":
        countries = [c for c in countries if 40 <= (c.get("combined_systemic_risk_score") or 0) < 60]
    elif risk_filter == "med_high":
        countries = [c for c in countries if (c.get("combined_systemic_risk_score") or 0) >= 40]
    return countries


@bp.route("/integration/export")
def integration_export():
    """Export current Countries & Regions view as CSV (same filters: region, q, sort, order, risk)."""
    region = request.args.get("region", "").strip() or None
    sort = request.args.get("sort", "").strip() or "combined_systemic_risk_score"
    order = request.args.get("order", "desc").strip().lower()
    if order not in ("asc", "desc"):
        order = "desc"
    q = request.args.get("q", "").strip() or None
    risk_filter = request.args.get("risk", "").strip().lower() or "all"
    if risk_filter not in ("all", "high", "medium", "med_high"):
        risk_filter = "all"
    countries = _integration_export_countries(region, sort, order, q, risk_filter)
    import csv
    buf = BytesIO()
    writer = csv.writer(buf)
    headers = [
        "country_code", "country_code_2", "country_name", "region", "population_2026", "land_area_km2", "density_per_km2",
        "geopolitical_fragility_score", "economic_fragility_score", "combined_systemic_risk_score",
        "trade_flow_pct_gdp", "debt_distress_score", "reserve_months_imports", "energy_import_exposure_pct",
        "updated_at",
    ]
    writer.writerow(headers)
    for c in countries:
        code3 = (c.get("country_code") or "").upper()
        code2 = ISO3_TO_2.get(code3, "")
        writer.writerow([
            c.get("country_code") or "",
            code2,
            c.get("country_name") or "",
            c.get("region") or "",
            c.get("population_2026"),
            c.get("land_area_km2"),
            c.get("density_per_km2"),
            c.get("geopolitical_fragility_score"),
            c.get("economic_fragility_score"),
            c.get("combined_systemic_risk_score"),
            c.get("trade_flow_pct_gdp"),
            c.get("debt_distress_score"),
            c.get("reserve_months_imports"),
            c.get("energy_import_exposure_pct"),
            (c.get("updated_at") or "")[:19],
        ])
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="countries-regions.csv"'},
    )


@bp.route("/integration/export/excel")
def integration_export_excel():
    """Export current Countries & Regions view as Excel (same filters as CSV)."""
    try:
        import xlsxwriter
    except ImportError:
        return redirect(url_for("main.integration_dashboard"))
    region = request.args.get("region", "").strip() or None
    sort = request.args.get("sort", "").strip() or "combined_systemic_risk_score"
    order = request.args.get("order", "desc").strip().lower()
    if order not in ("asc", "desc"):
        order = "desc"
    q = request.args.get("q", "").strip() or None
    risk_filter = request.args.get("risk", "").strip().lower() or "all"
    if risk_filter not in ("all", "high", "medium", "med_high"):
        risk_filter = "all"
    countries = _integration_export_countries(region, sort, order, q, risk_filter)
    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("Countries")
    headers = [
        "country_code", "country_code_2", "country_name", "region", "population_2026", "land_area_km2", "density_per_km2",
        "geopolitical_fragility_score", "economic_fragility_score", "combined_systemic_risk_score",
        "trade_flow_pct_gdp", "debt_distress_score", "reserve_months_imports", "energy_import_exposure_pct",
        "updated_at",
    ]
    for col, h in enumerate(headers):
        ws.write(0, col, h)
    for row, c in enumerate(countries, 1):
        code3 = (c.get("country_code") or "").upper()
        code2 = ISO3_TO_2.get(code3, "")
        ws.write(row, 0, c.get("country_code") or "")
        ws.write(row, 1, code2)
        ws.write(row, 2, c.get("country_name") or "")
        ws.write(row, 3, c.get("region") or "")
        ws.write(row, 4, c.get("population_2026"))
        ws.write(row, 5, c.get("land_area_km2"))
        ws.write(row, 6, c.get("density_per_km2"))
        ws.write(row, 7, c.get("geopolitical_fragility_score"))
        ws.write(row, 8, c.get("economic_fragility_score"))
        ws.write(row, 9, c.get("combined_systemic_risk_score"))
        ws.write(row, 10, c.get("trade_flow_pct_gdp"))
        ws.write(row, 11, c.get("debt_distress_score"))
        ws.write(row, 12, c.get("reserve_months_imports"))
        ws.write(row, 13, c.get("energy_import_exposure_pct"))
        ws.write(row, 14, (c.get("updated_at") or "")[:19])
    wb.close()
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="countries-regions.xlsx"'},
    )


def _get_gpi_un_country_summary(country_code: str):
    """Return GPI UN alignment summary for a country; None if unavailable."""
    try:
        from app.un_votes.readers import get_country_alignment_summary
        return get_country_alignment_summary(country_code or "")
    except Exception:
        return None


def _resolve_country_code(code: str):
    """Return 3-letter country code for integration lookup; accept 2- or 3-letter input or full name."""
    raw = (code or "").strip()
    if not raw:
        return None
    code = raw.upper()
    if len(code) == 3 and code.isalpha():
        return code
    if len(code) == 2:
        rev = {v: k for k, v in ISO3_TO_2.items()}
        return rev.get(code)
    try:
        from app.un_votes.country_map import normalize_country_to_iso3
        iso3 = normalize_country_to_iso3(raw)
        return iso3
    except ImportError:
        pass
    return None


@bp.route("/country/<country_code>")
def country_dashboard(country_code):
    """Country operating sheet: header (risk, economic, escalation) + 5 tabs (Stability, Economic & Trade, Security, Treaties, Scenario Sensitivity)."""
    code_3 = _resolve_country_code(country_code)
    if not code_3:
        return "Country not found", 404
    country = get_integration_country(code_3)
    if not country:
        return "Country not found", 404
    country_code = code_3
    code_2 = ISO3_TO_2.get(country_code.upper())
    region_risk = get_risk_index(region_code=code_2) if code_2 else None
    escalation_pct = None
    if region_risk:
        s = region_risk.get("sanctions_probability_pct") or 0
        t = region_risk.get("trade_disruption_pct") or 0
        escalation_pct = round((s + t) / 2, 1)
    name = (country.get("country_name") or "").strip() or country_code
    cdei_latest = sfi_latest = fragility_latest = None
    try:
        from app.institutional_models.readers import (
            get_cdei_by_country, get_sfi_by_country, get_fragility_by_country,
            get_fragility_history, get_cdei_history, get_sfi_history,
            get_gepi_history_by_country,
        )
        cdei_rows = get_cdei_by_country(country_code=country_code, as_of=None)
        sfi_rows = get_sfi_by_country(country_code=country_code, as_of=None)
        fragility_rows = get_fragility_by_country(country_code=country_code, as_of=None)
        cdei_latest = cdei_rows[0] if cdei_rows else None
        sfi_latest = sfi_rows[0] if sfi_rows else None
        fragility_latest = fragility_rows[0] if fragility_rows else None
        fragility_history = get_fragility_history(country_code=country_code, days=90)
        cdei_history = get_cdei_history(country_code=country_code, days=90)
        sfi_history = get_sfi_history(country_code=country_code, days=90)
        gepi_history = get_gepi_history_by_country(country_code=country_code, days=90)
    except Exception:
        fragility_history = cdei_history = sfi_history = gepi_history = []
    try:
        macro_history = get_macroeconomic_stress_history(country_code=country_code, days=365)
    except Exception:
        macro_history = []
    try:
        from app.country_history import get_country_history_from_wikipedia, get_country_activity_timeline
        wiki_history = get_country_history_from_wikipedia(name)
        activity_timeline = get_country_activity_timeline(country_code=country_code, country_name=name, days=90)
    except Exception:
        wiki_history = {"extract": None, "url": None, "title": None, "error": None}
        activity_timeline = []
    # Synthesis: operating picture (country briefing) + executive one-pager
    country_briefing_result = {}
    executive_result = {}
    try:
        from app.synthesis import generate_country_operating_picture, generate_executive_one_pager
        country_briefing_result = generate_country_operating_picture(country_code)
        exec_data = generate_executive_one_pager(country_code=country_code) or {}
        if exec_data and not exec_data.get("error"):
            exec_data["links"] = {
                "Risk Engine": url_for("main.risk_dashboard"),
                "Supply Chain": url_for("main.supply_chain_dashboard"),
                "Integration": url_for("main.integration_dashboard"),
            }
            executive_result = exec_data
    except Exception:
        pass
    sanctions_target = get_sanctions(target=name, limit=30)
    return render_template(
        "country_dashboard.html",
        country=country,
        escalation_pct=escalation_pct,
        region_risk=region_risk,
        cdei_latest=cdei_latest,
        sfi_latest=sfi_latest,
        fragility_latest=fragility_latest,
        # Tab 1: Political Stability
        approval=get_approval_ratings(country_code=country_code, limit=20),
        protests=get_protest_tracking(country_code=country_code, limit=20),
        elections=get_election_calendar(country_code=country_code, limit=20),
        youth_unemployment=get_youth_unemployment(country_code=country_code, limit=15),
        food_inflation=get_food_inflation_alerts(country_code=country_code, limit=15),
        # Tab 2: Economic & Trade
        currency_stress=get_currency_stress(country_code=country_code, limit=15),
        # Tab 3: Security & Military
        defense_spending=get_defense_spending(country_code=country_code, limit=15),
        border_incidents=get_border_incidents(country_code=country_code, limit=15),
        military_exercises=get_military_exercises(region=country.get("region"), limit=15),
        naval_deployments=get_naval_deployments(region=country.get("region"), limit=20),
        arms_trade_supplier=get_arms_trade(supplier=name, limit=15),
        arms_trade_recipient=get_arms_trade(recipient=name, limit=15),
        # Tab 4: Treaties & Diplomacy
        treaties=get_treaties(party=name, limit=50),
        sanctions_target=sanctions_target,
        voting_alignment=get_voting_alignment(country=country_code, limit=30),
        un_votes=get_un_votes(country_code=country_code, limit=50) if country_code else [],
        gpi_un_summary=_get_gpi_un_country_summary(country_code) if country_code else None,
        chokepoint_exposure=get_flows_for_country(country_code=country_code, limit=50),
        # Tab 5: Scenario Sensitivity (no backend yet; template uses JS sliders)
        # Extended indicators (roadmap 1–13)
        macroeconomic_stress=get_macroeconomic_stress(country_code=country_code),
        energy_commodity=get_energy_commodity_exposure(country_code=country_code),
        military_capability=get_military_capability_snapshot(country_code=country_code),
        trade_partners=get_trade_flow_partners(country_code=country_code),
        multilateral=get_multilateral_participation(country_code=country_code),
        capital_flows=get_capital_flows(country_code=country_code),
        elite_institutional=get_elite_institutional(country_code=country_code),
        climate_vulnerability=get_climate_resource_vulnerability(country_code=country_code),
        technology_semiconductor=get_technology_semiconductor(country_code=country_code),
        conflict_imports=get_conflict_event_imports(country_code=country_code, limit=20),
        geospatial=get_geospatial_infrastructure(country_code=country_code),
        # History tab
        fragility_history=fragility_history,
        cdei_history=cdei_history,
        sfi_history=sfi_history,
        gepi_history=gepi_history,
        macro_history=macro_history,
        wiki_history=wiki_history,
        activity_timeline=activity_timeline,
        country_briefing=country_briefing_result,
        executive_one_pager=executive_result,
        desk_terminal_intel=get_desk_terminal_intel_for_country(country_code, limit=10),
        policy_context=_policy_entity_page_context(
            get_effective_user_id(), "country", country_code, country.get("updated_at")
        ),
    )


@bp.route("/country/<country_code>/briefing")
def country_briefing(country_code):
    """Country / Region Operating Picture: LLM-synthesized situation summary + bullets."""
    code_3 = _resolve_country_code(country_code)
    if not code_3:
        return "Country not found", 404
    from app.synthesis import generate_country_operating_picture
    result = generate_country_operating_picture(code_3)
    country = get_integration_country(code_3)
    return render_template(
        "synthesis_country_briefing.html",
        country=country or {"country_name": code_3, "country_code": code_3},
        summary=result.get("summary", ""),
        bullets=result.get("bullets", []),
        data=result.get("data", {}),
        error=result.get("error"),
        generated_at=result.get("generated_at"),
    )


@bp.route("/synthesis")
def synthesis_hub():
    """Redirect to Help #synthesis (synthesis integrated across platform)."""
    return redirect(url_for("main.help_page", _anchor="synthesis"))


@bp.route("/synthesis/thematic")
def synthesis_thematic():
    """Thematic briefing (US-China, Russia-Ukraine, Middle East this week)."""
    topic = (request.args.get("topic") or "US-China").strip()
    from app.synthesis import THEMATIC_TOPICS, generate_thematic_briefing
    if topic not in THEMATIC_TOPICS:
        topic = "US-China"
    result = generate_thematic_briefing(topic)
    return render_template(
        "synthesis_thematic.html",
        topic=topic,
        topics=list(THEMATIC_TOPICS.keys()),
        summary=result.get("summary", ""),
        bullets=result.get("bullets", []),
        data=result.get("data", {}),
        error=result.get("error"),
        generated_at=result.get("generated_at"),
    )


@bp.route("/synthesis/divergences")
def synthesis_divergences():
    """Cross-signal coherence: detect and highlight model mismatches."""
    from app.synthesis import get_signal_divergences
    divergences = get_signal_divergences()
    return render_template(
        "synthesis_divergence.html",
        divergences=divergences,
    )


@bp.route("/synthesis/executive")
def synthesis_executive():
    """Executive one-pager: per country or per region. Risk, drivers, implications, links."""
    country_code = request.args.get("country", "").strip() or None
    region_code = request.args.get("region", "").strip() or None
    if country_code:
        code_3 = _resolve_country_code(country_code)
        if not code_3:
            return "Country not found", 404
        country_code = code_3
    from app.synthesis import generate_executive_one_pager
    result = generate_executive_one_pager(country_code=country_code, region_code=region_code)
    if result.get("error"):
        return result.get("error", "Error"), 404
    data = result
    # Build links dict for template
    links = {}
    if data.get("type") == "country":
        cc = data.get("link_keys", {}).get("country_code")
        if cc:
            links["Country Dashboard"] = url_for("main.country_dashboard", country_code=cc)
            links["Operating Picture"] = url_for("main.country_briefing", country_code=cc)
        links["Risk Engine"] = url_for("main.risk_dashboard")
        links["Supply Chain"] = url_for("main.supply_chain_dashboard")
        links["Integration"] = url_for("main.integration_dashboard")
    else:
        links["Risk Engine"] = url_for("main.risk_dashboard")
        links["Integration"] = url_for("main.integration_dashboard")
        links["Situation Room"] = url_for("main.situation_room")
    data["links"] = links
    return render_template("synthesis_executive.html", data=data)


@bp.route("/country/<country_code>/executive")
def country_executive(country_code):
    """Executive one-pager for a specific country. Redirects to synthesis/executive?country=."""
    code_3 = _resolve_country_code(country_code)
    if not code_3:
        return "Country not found", 404
    return redirect(url_for("main.synthesis_executive", country=code_3))


@bp.route("/synthesis/trajectory")
def synthesis_trajectory():
    """Escalation trajectory: GEPI, risk, article impact time series + trend narrative."""
    days = request.args.get("days", 30, type=int)
    days = min(90, max(7, days))
    from app.synthesis import get_escalation_trajectory
    data = get_escalation_trajectory(days=days)
    return render_template("synthesis_trajectory.html", data=data)


@bp.route("/synthesis/meta-briefing")
def synthesis_meta_briefing():
    """Weekly system state briefing: GEPI, risk, fragility, spike/declining topics, scenario runs, digests, clusters."""
    days = request.args.get("days", 7, type=int)
    days = min(30, max(7, days))
    from app.synthesis import generate_meta_briefing
    data = generate_meta_briefing(days=days)
    return render_template("synthesis_meta_briefing.html", data=data)


@bp.route("/integration/country/<country_code>")
def integration_country(country_code):
    """Redirect to full Country Dashboard (operating sheet)."""
    code_3 = _resolve_country_code(country_code)
    if not code_3:
        return "Country not found", 404
    return redirect(url_for("main.country_dashboard", country_code=code_3))


# --- Diplomacy & Treaty Intelligence ---
DIPLOMACY_PRESETS = [
    ("Trade agreements", {"type": "trade_agreement"}),
    ("WTO RTAs", {"source_wto": "1"}),
    ("In force", {"status": "In Force"}),
    ("With documents", {"has_documents": "1"}),
    ("Defense pacts", {"type": "defense_pact"}),
    ("Goods", {"coverage": "Goods"}),
    ("Services", {"coverage": "Services"}),
    ("Goods & Services", {"coverage": "Goods & Services"}),
    ("Investment treaties", {"type": "investment_treaty"}),
    ("All types", {}),
]


@bp.route("/diplomacy/legislative")
def diplomacy_legislative():
    """Legislative / policy tracker: sanction bills, defense bills, trade amendments, export controls."""
    jurisdiction = request.args.get("jurisdiction", "").strip() or None
    bill_type = request.args.get("bill_type", "").strip() or None
    bills = get_legislative_policy_tracker(jurisdiction=jurisdiction, bill_type=bill_type, limit=100)
    return render_template(
        "diplomacy_legislative.html",
        bills=bills,
        jurisdiction=jurisdiction,
        bill_type=bill_type,
    )


@bp.route("/diplomacy")
def diplomacy_dashboard():
    """Diplomacy & Treaty Intelligence: agreements, sanctions, UN voting & alignment, tools."""
    summary = get_diplomacy_summary()
    treaty_counts_by_type = get_treaty_counts_by_type()
    treaties = get_treaties(limit=20)
    sanctions = get_sanctions(limit=20)
    alignment = get_voting_alignment(limit=30)
    escalation_treaties = get_treaties(escalation_only=True, limit=20)
    recent_treaties = get_treaties(limit=5, order_by="created")
    recent_sanctions = get_sanctions(limit=5)
    un_resolutions = get_un_resolutions(limit=10)
    recent_un_votes = get_un_votes(limit=20)
    legislative_bills = get_legislative_policy_tracker(limit=10)
    multilateral_orgs = ["WTO", "IMF", "NATO", "AIIB", "BRI"]
    multilateral_summary = {org: get_multilateral_summary_by_org(org, limit=10) for org in multilateral_orgs}
    return render_template(
        "diplomacy.html",
        treaties=treaties,
        sanctions=sanctions,
        alignment=alignment,
        un_resolutions=un_resolutions,
        recent_un_votes=recent_un_votes,
        escalation_treaties=escalation_treaties,
        recent_treaties=recent_treaties,
        recent_sanctions=recent_sanctions,
        treaty_types=TREATY_TYPES,
        diplomacy_summary=summary,
        treaty_counts_by_type=treaty_counts_by_type,
        legislative_bills=legislative_bills,
        multilateral_summary=multilateral_summary,
        sanctions_watch_url=url_for("main.diplomacy_sanctions"),
        relationship_mapper_url=url_for("main.relationship_mapper") + "?mode=diplomatic",
        conflict_dashboard_url=url_for("main.conflict_dashboard"),
        diplomacy_treaties_rss_url=url_for("main.diplomacy_treaties_rss"),
        diplomacy_treaties_url=url_for("main.diplomacy_treaties"),
        diplomacy_legislative_url=url_for("main.diplomacy_legislative"),
    )


@bp.route("/diplomacy/treaties")
def diplomacy_treaties():
    treaty_type = request.args.get("type", "").strip() or None
    party = request.args.get("party", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    search = request.args.get("search", "").strip() or None
    source_wto = request.args.get("source_wto") == "1"
    has_documents = request.args.get("has_documents") == "1"
    status = request.args.get("status", "").strip() or None
    provision = request.args.get("provision", "").strip() or None
    region = request.args.get("region", "").strip() or None
    coverage = request.args.get("coverage", "").strip() or None
    order_by = request.args.get("sort", "").strip() or None
    view_mode = request.args.get("view", "table").strip() or "table"
    if view_mode not in ("table", "card", "timeline"):
        view_mode = "table"
    if view_mode == "timeline" and not order_by:
        order_by = "date_desc"
    per_page = 20
    page = max(1, int(request.args.get("page") or 1))
    offset = (page - 1) * per_page
    treaties = get_treaties(
        treaty_type=treaty_type,
        party=party,
        date_from=date_from,
        date_to=date_to,
        search=search,
        source_wto=source_wto or None,
        has_documents=has_documents or None,
        status=status,
        provision=provision,
        region=region,
        coverage=coverage,
        limit=per_page,
        offset=offset,
        order_by=order_by,
    )
    for t in treaties:
        t["_parsed"] = parse_treaty_summary(t.get("summary"))
        t["_has_docs"] = bool(
            (t.get("document_url") and t["document_url"].strip())
            or (t.get("clauses_json") and "agreement_links" in (t["clauses_json"] or "") and "http" in (t["clauses_json"] or ""))
        )
        t["_is_wto"] = bool(
            (t.get("source_url") and "wto.org" in t["source_url"]) or t.get("wto_rta_id")
        )
    total = get_treaties_count(
        treaty_type=treaty_type,
        party=party,
        date_from=date_from,
        date_to=date_to,
        search=search,
        source_wto=source_wto or None,
        has_documents=has_documents or None,
        status=status,
        provision=provision,
        region=region,
        coverage=coverage,
    )
    total_pages = max(1, (total + per_page - 1) // per_page)
    country_name_to_code = _country_name_to_code_map()
    query_base = request.args.to_dict(flat=True) if request.args else {}
    query_prev = dict(query_base) if page > 1 else None
    if query_prev:
        query_prev["page"] = page - 1
    query_next = dict(query_base) if page < total_pages else None
    if query_next:
        query_next["page"] = page + 1
    agreements_stats = get_agreements_page_stats()
    regions = get_treaty_distinct_regions()
    coverages = get_treaty_distinct_coverages()
    treaties_by_year = get_treaties_by_year(limit_years=25)
    query_base_no_page = {k: v for k, v in query_base.items() if k != "page" and k != "view"}
    treaties_by_year_with_urls = []
    for ty in treaties_by_year[:15]:
        params = dict(query_base_no_page)
        params["date_from"] = f"{ty['year']}-01-01"
        params["date_to"] = f"{ty['year']}-12-31"
        params["view"] = "timeline"
        ty["filter_url"] = url_for("main.diplomacy_treaties", **params)
        treaties_by_year_with_urls.append(ty)
    return render_template(
        "diplomacy_treaties.html",
        treaties=treaties,
        treaty_types=TREATY_TYPES,
        treaty_type=treaty_type,
        party=party,
        date_from=date_from,
        date_to=date_to,
        search=search,
        source_wto=source_wto,
        has_documents=has_documents,
        status=status,
        provision=provision,
        region=region,
        coverage=coverage,
        regions=regions,
        coverages=coverages,
        order_by=order_by,
        view_mode=view_mode,
        page=page,
        total_pages=total_pages,
        total=total,
        per_page=per_page,
        country_name_to_code=country_name_to_code,
        query_base=query_base,
        query_prev=query_prev,
        query_next=query_next,
        presets=DIPLOMACY_PRESETS,
        agreements_stats=agreements_stats,
        treaties_by_year=treaties_by_year_with_urls,
        diplomacy_treaties_export_url=url_for("main.diplomacy_treaties_export", **{k: v for k, v in query_base.items() if k != "view"}),
        diplomacy_treaties_export_xlsx_url=url_for("main.diplomacy_treaties_export_xlsx", **{k: v for k, v in query_base.items() if k != "view"}),
        query_base_no_page=query_base_no_page,
    )


@bp.route("/diplomacy/treaties/export")
def diplomacy_treaties_export():
    """Export treaties list as CSV (current filters)."""
    treaty_type = request.args.get("type", "").strip() or None
    party = request.args.get("party", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    search = request.args.get("search", "").strip() or None
    source_wto = request.args.get("source_wto") == "1"
    has_documents = request.args.get("has_documents") == "1"
    status = request.args.get("status", "").strip() or None
    provision = request.args.get("provision", "").strip() or None
    region = request.args.get("region", "").strip() or None
    coverage = request.args.get("coverage", "").strip() or None
    order_by = request.args.get("sort", "").strip() or None
    rows = get_treaties(
        treaty_type=treaty_type,
        party=party,
        date_from=date_from,
        date_to=date_to,
        search=search,
        source_wto=source_wto or None,
        has_documents=has_documents or None,
        status=status,
        provision=provision,
        region=region,
        coverage=coverage,
        limit=2000,
        order_by=order_by,
    )
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow([
        "Name", "Type", "Party A", "Party B", "Signed date", "Status", "Coverage",
        "Escalation clause", "Document URL", "WTO RTA ID", "Source URL",
    ])
    for r in rows:
        parsed = parse_treaty_summary(r.get("summary"))
        w.writerow([
            r.get("name") or "",
            r.get("treaty_type") or "",
            r.get("party_a") or "",
            r.get("party_b") or "",
            (r.get("signed_date") or "")[:10],
            parsed.get("status") or "",
            parsed.get("coverage") or "",
            "Yes" if r.get("has_escalation_clause") else "No",
            r.get("document_url") or "",
            str(r.get("wto_rta_id") or ""),
            r.get("source_url") or "",
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=diplomacy_treaties.csv"},
    )


@bp.route("/diplomacy/treaties/export/xlsx")
def diplomacy_treaties_export_xlsx():
    """Export treaties list as Excel (current filters)."""
    treaty_type = request.args.get("type", "").strip() or None
    party = request.args.get("party", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    search = request.args.get("search", "").strip() or None
    source_wto = request.args.get("source_wto") == "1"
    has_documents = request.args.get("has_documents") == "1"
    status = request.args.get("status", "").strip() or None
    provision = request.args.get("provision", "").strip() or None
    region = request.args.get("region", "").strip() or None
    coverage = request.args.get("coverage", "").strip() or None
    order_by = request.args.get("sort", "").strip() or None
    rows = get_treaties(
        treaty_type=treaty_type,
        party=party,
        date_from=date_from,
        date_to=date_to,
        search=search,
        source_wto=source_wto or None,
        has_documents=has_documents or None,
        status=status,
        provision=provision,
        region=region,
        coverage=coverage,
        limit=5000,
        order_by=order_by,
    )
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Treaties"
        headers = [
            "Name", "Type", "Party A", "Party B", "Signed date", "Status", "Coverage",
            "Escalation clause", "Document URL", "WTO RTA ID", "Source URL",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for row_idx, r in enumerate(rows, 2):
            parsed = parse_treaty_summary(r.get("summary"))
            ws.cell(row=row_idx, column=1, value=r.get("name") or "")
            ws.cell(row=row_idx, column=2, value=r.get("treaty_type") or "")
            ws.cell(row=row_idx, column=3, value=r.get("party_a") or "")
            ws.cell(row=row_idx, column=4, value=r.get("party_b") or "")
            ws.cell(row=row_idx, column=5, value=(r.get("signed_date") or "")[:10])
            ws.cell(row=row_idx, column=6, value=parsed.get("status") or "")
            ws.cell(row=row_idx, column=7, value=parsed.get("coverage") or "")
            ws.cell(row=row_idx, column=8, value="Yes" if r.get("has_escalation_clause") else "No")
            ws.cell(row=row_idx, column=9, value=r.get("document_url") or "")
            ws.cell(row=row_idx, column=10, value=str(r.get("wto_rta_id") or ""))
            ws.cell(row=row_idx, column=11, value=r.get("source_url") or "")
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=diplomacy_treaties.xlsx"},
        )
    except ImportError:
        return redirect(url_for("main.diplomacy_treaties_export", **request.args.to_dict(flat=True)))


@bp.route("/diplomacy/treaties/add", methods=["GET", "POST"])
def diplomacy_treaty_add():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        party_a = (request.form.get("party_a") or "").strip()
        if name and party_a:
            add_treaty(
                treaty_type=request.form.get("treaty_type") or "other",
                name=name,
                party_a=party_a,
                party_b=(request.form.get("party_b") or "").strip(),
                signed_date=(request.form.get("signed_date") or "").strip() or None,
                summary=(request.form.get("summary") or "").strip() or None,
                full_text=(request.form.get("full_text") or "").strip() or None,
                has_escalation_clause=1 if request.form.get("has_escalation_clause") else 0,
                source_url=(request.form.get("source_url") or "").strip() or None,
                document_url=(request.form.get("document_url") or "").strip() or None,
            )
            return redirect(url_for("main.diplomacy_treaties"))
    return render_template("diplomacy_treaty_add.html", treaty_types=TREATY_TYPES)


@bp.route("/diplomacy/treaties/<int:treaty_id>")
def diplomacy_treaty_detail(treaty_id):
    treaty = get_treaty(treaty_id)
    if not treaty:
        return "Treaty not found", 404
    treaty["_parsed"] = parse_treaty_summary(treaty.get("summary"))
    if treaty.get("wto_rta_id") is None and treaty.get("clauses_json"):
        try:
            data = json.loads(treaty["clauses_json"])
            if isinstance(data, dict) and data.get("wto_rta_id"):
                treaty["wto_rta_id"] = data["wto_rta_id"]
        except (TypeError, ValueError):
            pass
    related = get_related_treaties(treaty_id, limit=10)
    clauses_list = []
    wto_data = None
    if treaty.get("clauses_json"):
        try:
            data = json.loads(treaty["clauses_json"])
            if isinstance(data, list):
                clauses_list = data
            elif isinstance(data, dict):
                if "clauses" in data:
                    clauses_list = data["clauses"]
                if data.get("wto_rta_id"):
                    wto_data = data
        except (TypeError, ValueError):
            pass
    escalation_detail = None
    if treaty.get("has_escalation_clause") and treaty.get("full_text"):
        from app.diplomacy import detect_escalation_clause
        escalation_detail = detect_escalation_clause(treaty["full_text"])
    country_name_to_code = _country_name_to_code_map()
    party_a_code = country_name_to_code.get((treaty.get("party_a") or "").strip()) if treaty.get("party_a") else None
    party_b_code = country_name_to_code.get((treaty.get("party_b") or "").strip()) if treaty.get("party_b") else None
    relationship_mapper_url = url_for("main.relationship_mapper") + "?mode=diplomatic"
    if party_a_code:
        relationship_mapper_url += "&focus=" + party_a_code
    return render_template(
        "diplomacy_treaty_detail.html",
        treaty=treaty,
        related_treaties=related,
        clauses_list=clauses_list,
        wto_data=wto_data,
        escalation_detail=escalation_detail,
        relationship_mapper_url=relationship_mapper_url,
        party_a_code=party_a_code,
        party_b_code=party_b_code,
    )


@bp.route("/diplomacy/treaties/<int:treaty_id>/edit", methods=["GET", "POST"])
def diplomacy_treaty_edit(treaty_id):
    treaty = get_treaty(treaty_id)
    if not treaty:
        return "Treaty not found", 404
    if request.method == "POST":
        update_treaty(
            treaty_id,
            treaty_type=request.form.get("treaty_type") or treaty["treaty_type"],
            name=(request.form.get("name") or "").strip() or treaty["name"],
            party_a=(request.form.get("party_a") or "").strip() or treaty["party_a"],
            party_b=(request.form.get("party_b") or "").strip(),
            signed_date=(request.form.get("signed_date") or "").strip() or None,
            summary=(request.form.get("summary") or "").strip(),
            full_text=(request.form.get("full_text") or "").strip(),
            has_escalation_clause=1 if request.form.get("has_escalation_clause") else 0,
            source_url=(request.form.get("source_url") or "").strip(),
            document_url=(request.form.get("document_url") or "").strip(),
        )
        return redirect(url_for("main.diplomacy_treaty_detail", treaty_id=treaty_id))
    return render_template("diplomacy_treaty_edit.html", treaty=treaty, treaty_types=TREATY_TYPES)


@bp.route("/diplomacy/sanctions")
def diplomacy_sanctions():
    """Sanctions registry: OFAC, EU, entity lists, export controls. Merged with sanctions watch."""
    imposing = (request.args.get("imposing") or "").strip() or None
    source = (request.args.get("source") or "").strip() or None
    target = (request.args.get("target") or "").strip() or None
    date_from = (request.args.get("date_from") or "").strip() or None
    date_to = (request.args.get("date_to") or "").strip() or None
    search = (request.args.get("search") or "").strip() or None
    entity_search = (request.args.get("entity_search") or "").strip() or None
    entity_source = (request.args.get("entity_source") or "").strip() or None
    export_search = (request.args.get("export_search") or "").strip() or None
    export_issuer = (request.args.get("export_issuer") or "").strip() or None
    export_type = (request.args.get("export_type") or "").strip() or None

    per_page = 25
    page = max(1, int(request.args.get("page") or 1))
    offset = (page - 1) * per_page

    ofac_sanctions = get_sanctions(source="OFAC", target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search, limit=50)
    eu_sanctions = get_sanctions(source="EU", target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search, limit=50)
    all_sanctions_total = get_sanctions_total_count(source=source, target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search)
    all_sanctions = get_sanctions(source=source, target=target, imposing=imposing, date_from=date_from, date_to=date_to, search=search, limit=per_page, offset=offset)
    total_pages = max(1, (all_sanctions_total + per_page - 1) // per_page)

    if entity_source:
        one_list = get_entity_list_alerts(source=entity_source, search=entity_search, limit=25)
        entity_list_ofac = one_list if entity_source == "OFAC" else []
        entity_list_eu = one_list if entity_source == "EU" else []
        entity_list_us = one_list if entity_source == "US_BIS" else []
        entity_list_china = one_list if entity_source == "China" else []
    else:
        entity_list_ofac = get_entity_list_alerts(source="OFAC", search=entity_search, limit=25)
        entity_list_eu = get_entity_list_alerts(source="EU", search=entity_search, limit=25)
        entity_list_us = get_entity_list_alerts(source="US_BIS", search=entity_search, limit=25)
        entity_list_china = get_entity_list_alerts(source="China", search=entity_search, limit=25)

    export_restrictions = get_export_restrictions(issuer=export_issuer, restriction_type=export_type, search=export_search, limit=50)
    dual_use = get_export_restrictions(restriction_type="dual_use", search=export_search, limit=25)
    china_restrictions = get_export_restrictions(issuer="China", search=export_search, limit=25)

    meta = get_sanctions_watch_meta()
    country_name_to_code = _country_name_to_code_map()

    query_base = request.args.to_dict(flat=True) if request.args else {}
    query_prev = dict(query_base) if page > 1 else None
    if query_prev and page > 1:
        query_prev["page"] = page - 1
    query_next = dict(query_base) if page < total_pages else None
    if query_next and page < total_pages:
        query_next["page"] = page + 1

    return render_template(
        "diplomacy_sanctions.html",
        ofac_sanctions=ofac_sanctions,
        eu_sanctions=eu_sanctions,
        all_sanctions=all_sanctions,
        all_sanctions_total=all_sanctions_total,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        entity_list_ofac=entity_list_ofac,
        entity_list_eu=entity_list_eu,
        entity_list_us=entity_list_us,
        entity_list_china=entity_list_china,
        export_restrictions=export_restrictions,
        dual_use=dual_use,
        china_restrictions=china_restrictions,
        filter_source=source,
        filter_target=target,
        filter_imposing=imposing,
        filter_date_from=date_from,
        filter_date_to=date_to,
        filter_search=search,
        filter_entity_search=entity_search,
        filter_entity_source=entity_source,
        filter_export_search=export_search,
        filter_export_issuer=export_issuer,
        filter_export_type=export_type,
        country_name_to_code=country_name_to_code,
        query_base=query_base,
        query_prev=query_prev,
        query_next=query_next,
        sanctions_export_url=url_for("main.sanctions_watch_export", export_type="sanctions"),
        entity_list_export_url=url_for("main.sanctions_watch_export", export_type="entities"),
        export_restrictions_export_url=url_for("main.sanctions_watch_export", export_type="restrictions"),
        relationship_mapper_url=url_for("main.relationship_mapper") + "?mode=sanctions",
        diplomacy_sanctions_export_url=url_for("main.diplomacy_sanctions_export", **query_base),
        last_updated=meta.get("last_updated"),
        new_sanctions_7d=meta.get("new_sanctions_7d", 0),
        new_entities_7d=meta.get("new_entities_7d", 0),
        new_restrictions_7d=meta.get("new_restrictions_7d", 0),
    )


@bp.route("/diplomacy/sanctions/export")
def diplomacy_sanctions_export():
    """Export diplomacy sanctions registry as CSV (current filters)."""
    imposing = request.args.get("imposing", "").strip() or None
    target = request.args.get("target", "").strip() or None
    source = request.args.get("source", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    search = request.args.get("search", "").strip() or None
    rows = get_sanctions(imposing=imposing, target=target, source=source, date_from=date_from, date_to=date_to, search=search, limit=2000)
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Imposing", "Target", "Source", "Type", "Description", "Start", "End"])
    for r in rows:
        w.writerow([
            r.get("imposing_country") or "",
            r.get("target_country") or "",
            r.get("source") or "",
            r.get("measure_type") or "",
            (r.get("description") or "")[:500],
            (r.get("start_date") or "")[:10],
            (r.get("end_date") or "")[:10],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=diplomacy_sanctions.csv"},
    )


@bp.route("/diplomacy/sanctions/global")
def diplomacy_sanctions_global():
    """Global sanctions registry (sanctions_global): entity/person-level from OFAC, EU, Australia, UN."""
    jurisdiction = request.args.get("jurisdiction", "").strip() or None
    target_type = request.args.get("target_type", "").strip() or None
    country = request.args.get("country", "").strip() or None
    search = request.args.get("search", "").strip() or None
    date_from = request.args.get("date_from", "").strip() or None
    date_to = request.args.get("date_to", "").strip() or None
    per_page = 50
    page = max(1, int(request.args.get("page") or 1))
    offset = (page - 1) * per_page
    sanctions = get_sanctions_global(
        jurisdiction=jurisdiction,
        target_type=target_type,
        country=country,
        search=search,
        date_from=date_from,
        date_to=date_to,
        limit=per_page,
        offset=offset,
    )
    total = get_sanctions_global_count(
        jurisdiction=jurisdiction,
        target_type=target_type,
        country=country,
        search=search,
        date_from=date_from,
        date_to=date_to,
    )
    total_pages = max(1, (total + per_page - 1) // per_page)
    country_name_to_code = _country_name_to_code_map()
    query_base = request.args.to_dict(flat=True) if request.args else {}
    query_prev = dict(query_base) if page > 1 else None
    if query_prev:
        query_prev["page"] = page - 1
    query_next = dict(query_base) if page < total_pages else None
    if query_next:
        query_next["page"] = page + 1
    return render_template(
        "diplomacy_sanctions_global.html",
        sanctions=sanctions,
        jurisdiction=jurisdiction,
        target_type=target_type,
        country=country,
        search=search,
        date_from=date_from,
        date_to=date_to,
        page=page,
        total_pages=total_pages,
        total=total,
        per_page=per_page,
        country_name_to_code=country_name_to_code,
        query_base=query_base,
        query_prev=query_prev,
        query_next=query_next,
        sanctions_registry_url=url_for("main.diplomacy_sanctions"),
    )


@bp.route("/diplomacy/sanctions/add", methods=["GET", "POST"])
def diplomacy_sanction_add():
    if request.method == "POST":
        imposing = (request.form.get("imposing_country") or "").strip()
        target = (request.form.get("target_country") or "").strip()
        if imposing and target:
            add_sanction(
                imposing_country=imposing,
                target_country=target,
                measure_type=(request.form.get("measure_type") or "").strip() or None,
                description=(request.form.get("description") or "").strip() or None,
                start_date=(request.form.get("start_date") or "").strip() or None,
                end_date=(request.form.get("end_date") or "").strip() or None,
                source_url=(request.form.get("source_url") or "").strip() or None,
                source=(request.form.get("source") or "").strip() or None,
            )
            flash("Sanction added successfully.")
            return redirect(url_for("main.diplomacy_sanctions", added=1))
    return redirect(url_for("main.diplomacy_sanctions"))


@bp.route("/diplomacy/alignment")
def diplomacy_alignment():
    """UN voting & alignment: votes, resolutions, which countries vote together. Defunct countries in archive."""
    country_raw = request.args.get("country", "").strip() or None
    country = None
    if country_raw:
        try:
            from app.un_votes.country_map import normalize_country_to_iso3
            country = normalize_country_to_iso3(country_raw) or (country_raw if len(country_raw) == 3 and country_raw.isupper() else None)
        except Exception:
            country = country_raw if len(country_raw) == 3 and country_raw.isupper() else None
    min_votes = request.args.get("min_votes", type=int)
    sort = (request.args.get("sort") or "score_desc").strip()
    bloc = request.args.get("bloc", "").strip() or None
    archive = request.args.get("archive", "").strip() in ("1", "true", "yes")
    if min_votes is None or min_votes < 0:
        min_votes = None
    alignment = get_voting_alignment(
        country=country,
        min_votes=min_votes,
        limit=50,
        include_defunct=archive,
        sort=sort,
        bloc=bloc,
    )
    resolutions = []
    resolutions_for_country = []
    if country:
        votes = get_un_votes(country_code=country, limit=500)
        by_res = {v["resolution_id"]: v for v in votes}
        resolutions_for_country = list(by_res.values())[:50]
    un_polarization = []
    un_shocks = []
    un_bloc_cohesion = []
    try:
        from app.un_votes.readers import get_global_polarization, get_alignment_shocks, get_bloc_cohesion
        un_polarization = get_global_polarization(limit=12)
        un_shocks = get_alignment_shocks(limit=8, shock_only=True)
        un_bloc_cohesion = get_bloc_cohesion(limit=8)
    except Exception:
        pass
    blocs_for_filter = []
    try:
        from app.un_votes.blocs import BLOCS, P5, BLOC_DISPLAY_NAMES
        for key in list(BLOCS.keys()) + ["P5"]:
            blocs_for_filter.append((key, BLOC_DISPLAY_NAMES.get(key, key)))
    except ImportError:
        blocs_for_filter = [("EU", "EU"), ("ASEAN", "ASEAN"), ("G7", "G7"), ("BRICS", "BRICS"), ("GCC", "GCC"), ("OAS", "OAS"), ("AU", "AU")]
    query_base = request.args.to_dict(flat=True) if request.args else {}
    return render_template(
        "diplomacy_alignment.html",
        alignment=alignment,
        resolutions=resolutions,
        resolutions_for_country=resolutions_for_country,
        selected_country=country,
        min_votes=min_votes,
        sort=sort,
        bloc=bloc,
        archive=archive,
        blocs_for_filter=blocs_for_filter,
        query_base=query_base,
        diplomacy_alignment_export_url=url_for("main.diplomacy_alignment_export", **query_base),
        un_polarization=un_polarization,
        un_shocks=un_shocks,
        un_bloc_cohesion=un_bloc_cohesion,
    )


@bp.route("/diplomacy/alignment/export")
def diplomacy_alignment_export():
    """Export voting alignment as CSV. Use ?archive=1 for historical countries."""
    country = request.args.get("country", "").strip() or None
    min_votes = request.args.get("min_votes", type=int)
    sort = (request.args.get("sort") or "score_desc").strip()
    bloc = request.args.get("bloc", "").strip() or None
    archive = request.args.get("archive", "").strip() in ("1", "true", "yes")
    alignment = get_voting_alignment(
        country=country,
        min_votes=min_votes,
        limit=500,
        include_defunct=archive,
        sort=sort,
        bloc=bloc,
    )
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Country A", "Country B", "Alignment %", "Votes agreed", "Votes total"])
    for r in alignment:
        w.writerow([
            r.get("country_a_name") or r.get("country_a") or "",
            r.get("country_b_name") or r.get("country_b") or "",
            "%.1f" % (r.get("alignment_score") or 0),
            r.get("votes_agreed") or 0,
            r.get("votes_total") or 0,
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=diplomacy_alignment.csv"},
    )


@bp.route("/diplomacy/alignment/recompute", methods=["POST"])
def diplomacy_alignment_recompute():
    from app.models import compute_voting_alignment
    compute_voting_alignment(min_votes=5)
    flash("Alignment recomputed.")
    return redirect(url_for("main.diplomacy_alignment"))


@bp.route("/diplomacy/alignment/normalize", methods=["POST"])
def diplomacy_alignment_normalize():
    """Normalize un_votes to ISO3, clear alignment, recompute. Fixes duplicate same-country pairs."""
    from app.models import normalize_un_votes_to_iso3, compute_voting_alignment, clear_resolutions_cache
    migrated, skipped = normalize_un_votes_to_iso3()
    pairs = compute_voting_alignment(min_votes=5)
    clear_resolutions_cache()
    flash(f"Normalized {migrated} votes to ISO3 (skipped {skipped}); recomputed {pairs} alignment pairs.")
    return redirect(url_for("main.diplomacy_alignment"))


@bp.route("/diplomacy/alignment/sync-unige", methods=["POST"])
def diplomacy_alignment_sync_unige():
    """Sync UN votes from UNIGE UNGA-DM (MariaDB or CSV fallback)."""
    from app.jobs.un_votes_sync import sync_un_votes_from_unige
    from app.models import compute_voting_alignment, clear_resolutions_cache
    limit = request.form.get("limit", "").strip()
    limit = int(limit) if limit.isdigit() else None
    count, err = sync_un_votes_from_unige(limit=limit)
    if err:
        flash(f"Sync failed: {err}", "error")
        return redirect(url_for("main.diplomacy_alignment"))
    clear_resolutions_cache()
    pairs = compute_voting_alignment(min_votes=5)
    flash(f"Synced {count} votes from UNIGE; computed {pairs} alignment pairs.")
    return redirect(url_for("main.diplomacy_alignment"))


@bp.route("/diplomacy/alignment/import-csv", methods=["POST"])
def diplomacy_alignment_import_csv():
    """Import UN votes from uploaded CSV (UNIGE All_Votes format)."""
    import io
    from app.jobs.un_votes_sync import sync_un_votes_from_csv
    from app.models import compute_voting_alignment, clear_resolutions_cache
    f = request.files.get("csv_file")
    if not f or not f.filename or not f.filename.lower().endswith((".csv", ".txt")):
        flash("Please upload a CSV file (e.g. All_Votes from UNIGE UNGA-DM).", "error")
        return redirect(url_for("main.diplomacy_alignment"))
    limit = request.form.get("limit", "").strip()
    limit = int(limit) if limit.isdigit() else None
    try:
        # Wrap bytes stream in text-mode for csv.DictReader
        text_stream = io.TextIOWrapper(f.stream, encoding="utf-8", errors="replace")
        count, err = sync_un_votes_from_csv(text_stream, limit=limit)
    except Exception as e:
        flash(f"Import failed: {e}", "error")
        return redirect(url_for("main.diplomacy_alignment"))
    if err:
        flash(f"Import failed: {err}", "error")
        return redirect(url_for("main.diplomacy_alignment"))
    clear_resolutions_cache()
    pairs = compute_voting_alignment(min_votes=5)
    flash(f"Imported {count} votes; computed {pairs} alignment pairs.")
    return redirect(url_for("main.diplomacy_alignment"))


@bp.route("/diplomacy/compare", methods=["GET", "POST"])
def diplomacy_compare():
    """Treaty clause comparator (AI-powered). Prefill from ?treaty_id_a= & treaty_id_b=."""
    result = None
    treaty_id_a = request.args.get("treaty_id_a") or (request.form.get("treaty_id_a") if request.method == "POST" else None)
    treaty_id_b = request.args.get("treaty_id_b") or (request.form.get("treaty_id_b") if request.method == "POST" else None)
    treaty_a = get_treaty(int(treaty_id_a)) if treaty_id_a and str(treaty_id_a).isdigit() else None
    treaty_b = get_treaty(int(treaty_id_b)) if treaty_id_b and str(treaty_id_b).isdigit() else None
    if request.method == "POST":
        from app.diplomacy import compare_treaty_clauses
        text_a = request.form.get("text_a") or ""
        text_b = request.form.get("text_b") or ""
        if treaty_a and treaty_a.get("full_text"):
            text_a = treaty_a["full_text"]
        if treaty_b and treaty_b.get("full_text"):
            text_b = treaty_b["full_text"]
        result = compare_treaty_clauses(text_a, text_b)
        if result and treaty_a:
            result["treaty_a_name"] = treaty_a.get("name")
        if result and treaty_b:
            result["treaty_b_name"] = treaty_b.get("name")
    treaties = get_treaties(limit=50)
    return render_template(
        "diplomacy_compare.html",
        result=result,
        treaties=treaties,
        preselected_treaty_id_a=treaty_id_a,
        preselected_treaty_id_b=treaty_id_b,
        treaty_a=treaty_a,
        treaty_b=treaty_b,
    )


@bp.route("/diplomacy/escalation")
def diplomacy_escalation():
    """Escalation clause monitor: treaties with escalation/retaliation clauses."""
    treaty_type = request.args.get("type", "").strip() or None
    treaties = get_treaties(escalation_only=True, treaty_type=treaty_type, limit=100)
    query_base = request.args.to_dict(flat=True) if request.args else {}
    return render_template(
        "diplomacy_escalation.html",
        treaties=treaties,
        treaty_types=TREATY_TYPES,
        treaty_type=treaty_type,
        query_base=query_base,
        diplomacy_escalation_export_url=url_for("main.diplomacy_escalation_export", **query_base),
    )


@bp.route("/diplomacy/escalation/export")
def diplomacy_escalation_export():
    """Export escalation treaties as CSV."""
    treaty_type = request.args.get("type", "").strip() or None
    rows = get_treaties(escalation_only=True, treaty_type=treaty_type, limit=500)
    from io import StringIO
    import csv
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["Name", "Type", "Party A", "Party B", "Signed date"])
    for r in rows:
        w.writerow([
            r.get("name") or "",
            r.get("treaty_type") or "",
            r.get("party_a") or "",
            r.get("party_b") or "",
            (r.get("signed_date") or "")[:10],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=diplomacy_escalation_treaties.csv"},
    )


@bp.route("/diplomacy/treaties.rss")
def diplomacy_treaties_rss():
    """RSS feed of recently added/updated treaties (optional type filter)."""
    treaty_type = request.args.get("type", "").strip() or None
    treaties = get_treaties(treaty_type=treaty_type, limit=30, order_by="created")
    from flask import make_response
    from email.utils import formatdate
    import time
    now_rfc = formatdate(timeval=time.time(), localtime=False)
    xml = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        "<rss version=\"2.0\">",
        "<channel>",
        "<title>Treaties – Geopolitical Terminal</title>",
        "<link>" + request.host_url.rstrip("/") + url_for("main.diplomacy_treaties") + "</link>",
        "<description>Bilateral agreements, defense pacts, trade and investment treaties.</description>",
        "<lastBuildDate>" + now_rfc + "</lastBuildDate>",
    ]
    for t in treaties:
        title = (t.get("name") or "Treaty")[:100].replace("&", "&amp;").replace("<", "&lt;")
        desc = (t.get("summary") or "")[:500].replace("&", "&amp;").replace("<", "&lt;")
        date_str = (t.get("signed_date") or t.get("created_at") or "")[:10]
        link = request.host_url.rstrip("/") + url_for("main.diplomacy_treaty_detail", treaty_id=t.get("id"))
        try:
            from datetime import datetime
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            pub_rfc = formatdate(timeval=dt.timestamp(), localtime=False)
        except (ValueError, TypeError):
            pub_rfc = now_rfc
        guid = link.replace("&", "&amp;")
        xml.append("<item>")
        xml.append("<title>" + title + "</title>")
        xml.append("<description>" + desc + "</description>")
        xml.append("<pubDate>" + pub_rfc + "</pubDate>")
        xml.append("<link>" + guid + "</link>")
        xml.append("<guid isPermaLink=\"true\">" + guid + "</guid>")
        xml.append("</item>")
    xml.append("</channel></rss>")
    resp = make_response("\n".join(xml))
    resp.headers["Content-Type"] = "application/rss+xml; charset=utf-8"
    return resp


@bp.route("/diplomacy/treaties/<int:treaty_id>/scan-escalation", methods=["POST"])
def diplomacy_treaty_scan_escalation(treaty_id):
    treaty = get_treaty(treaty_id)
    if not treaty:
        return "Treaty not found", 404
    from app.diplomacy import detect_escalation_clause
    full_text = treaty.get("full_text") or ""
    detection = detect_escalation_clause(full_text)
    if detection.get("has_escalation"):
        update_treaty_escalation(treaty_id, 1)
    else:
        update_treaty_escalation(treaty_id, 0)
    return redirect(url_for("main.diplomacy_treaty_detail", treaty_id=treaty_id))


# --- Export PDF (print-friendly page; user can Print to PDF) ---
@bp.route("/export/digest/<int:digest_id>/pdf")
def export_digest_pdf(digest_id):
    d = get_digest(digest_id)
    if not d:
        return "Digest not found", 404
    try:
        d["items"] = json.loads(d["content"]).get("items", [])
    except (json.JSONDecodeError, TypeError):
        d["items"] = []
    return render_template("export_digest_pdf.html", digest=d)


@bp.route("/reports")
def reports_hub():
    """Reports hub: links to briefing builder, 10-year outlook, digests, scenario exports."""
    recent_briefings = get_saved_briefings(user_id=get_effective_user_id())[:5]
    return render_template("reports_hub.html", recent_briefings=recent_briefings)


def _briefing_group_articles(articles, grouping):
    """Group articles by topic, event_type, or date. Returns list of (group_label, [articles])."""
    if not grouping or grouping == "none" or not articles:
        return [("", articles)]
    from collections import defaultdict
    groups = defaultdict(list)
    for a in articles:
        if grouping == "event_type":
            key = (a.get("event_type") or "").strip() or "Other"
        elif grouping == "topic":
            topics = (a.get("topics") or "").strip()
            key = (topics.split(",")[0].strip() if topics else "") or "Other"
        elif grouping == "date":
            dt = a.get("published_utc") or a.get("scraped_at") or ""
            key = dt[:10] if dt else "Unknown"
        else:
            key = ""
        groups[key].append(a)
    if grouping == "date":
        return sorted(groups.items(), key=lambda x: x[0], reverse=True)
    if grouping == "event_type":
        order = ["Sanctions", "Military", "Diplomacy", "Election", "Trade", "Other"]
        return sorted(groups.items(), key=lambda x: (order.index(x[0]) if x[0] in order else 99, x[0]))
    return sorted(groups.items(), key=lambda x: (-len(x[1]), x[0]))


def _briefing_executive_summary(articles):
    """Generate 2–3 sentence executive summary from article counts by event_type and topics."""
    from collections import Counter
    event_counts = Counter()
    topic_counts = Counter()
    high_impact = 0
    for a in articles:
        et = (a.get("event_type") or "").strip() or "Other"
        event_counts[et] += 1
        imp = a.get("impact_score")
        if imp is not None and int(imp) >= 7:
            high_impact += 1
        for t in (a.get("topics") or "").split(","):
            t = t.strip()
            if t:
                topic_counts[t] += 1
    parts = [f"This briefing covers {len(articles)} article(s)."]
    if event_counts:
        top_events = event_counts.most_common(4)
        event_str = ", ".join(f"{n} {e}" for e, n in top_events)
        parts.append(f"By event type: {event_str}.")
    if high_impact:
        parts.append(f"{high_impact} high-impact item(s) (7+).")
    if topic_counts:
        top_topics = topic_counts.most_common(3)
        parts.append(f"Recurring themes: {', '.join(t[0] for t in top_topics)}.")
    return " ".join(parts)


@bp.route("/export/briefing", methods=["GET", "POST"])
def export_briefing():
    if request.method == "POST":
        title = (request.form.get("title") or "Briefing").strip()
        intro = (request.form.get("intro") or "").strip()
        format_type = (request.form.get("format") or "summary_link").strip() or "summary_link"
        valid_formats = ("minimal", "summary_only", "summary_link", "with_takeaways", "executive", "full")
        if format_type not in valid_formats:
            format_type = "summary_link"
        grouping = (request.form.get("grouping") or "none").strip() or "none"
        if grouping not in ("none", "topic", "event_type", "date"):
            grouping = "none"
        sort_by = (request.form.get("sort_by") or "order").strip() or "order"
        cover_page = (request.form.get("cover_page") or "").strip() == "1"
        cover_line = (request.form.get("cover_line") or "").strip()
        ids_raw = request.form.get("article_ids") or request.form.getlist("article_id") or []
        if isinstance(ids_raw, str):
            ids_raw = [x.strip() for x in ids_raw.replace(",", "\n").split() if x.strip()]
        seen = set()
        article_ids = []
        invalid_ids = []
        for x in ids_raw:
            if not isinstance(x, str) or not x.isdigit():
                continue
            aid = int(x)
            if aid in seen:
                continue
            seen.add(aid)
            if len(article_ids) >= 50:
                break
            article_ids.append(aid)
        articles = []
        for aid in article_ids:
            a = get_article(aid)
            if a:
                articles.append(a)
            else:
                invalid_ids.append(aid)
        if sort_by == "impact":
            articles = sorted(articles, key=lambda x: (-(x.get("impact_score") or 0), (x.get("published_utc") or "")[:10] or ""))
        grouped = _briefing_group_articles(articles, grouping)
        report_date = (request.form.get("report_date") or "").strip()
        if not report_date:
            report_date = datetime.utcnow().strftime("%Y-%m-%d")
        generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        sensitivity_tier = _parse_sensitivity_tier(request.form.get("sensitivity_tier"))
        legal_review_required = request.form.get("legal_review_required") == "1"
        from collections import Counter
        event_counts = Counter((a.get("event_type") or "Other") for a in articles)
        date_min = min((a.get("published_utc") or a.get("scraped_at") or "")[:10] for a in articles) if articles else ""
        date_max = max((a.get("published_utc") or a.get("scraped_at") or "")[:10] for a in articles) if articles else ""
        return render_template(
            "export_briefing_pdf.html",
            title=title,
            intro=intro,
            articles=articles,
            grouped=grouped,
            format_type=format_type,
            report_date=report_date,
            generated_at=generated_at,
            invalid_ids=invalid_ids,
            valid_count=len(articles),
            cover_line=cover_line,
            cover_page=cover_page,
            executive_summary=_briefing_executive_summary(articles),
            event_counts=dict(event_counts),
            date_range=(date_min, date_max) if date_min and date_max else None,
            sensitivity_tier=sensitivity_tier,
            legal_review_required=legal_review_required,
        )
    # GET: optional prefill from from_feed or load_id
    title = (request.args.get("title") or "Daily briefing").strip()
    intro = (request.args.get("intro") or "").strip()
    article_ids_prefill = request.args.get("article_ids", "").strip()
    uid_briefing = get_effective_user_id()
    load_id = request.args.get("load", type=int)
    loaded_saved_for_sensitivity = None
    if load_id:
        saved = get_saved_briefing(load_id, user_id=uid_briefing)
        if saved:
            title = saved.get("title") or title
            intro = saved.get("intro") or intro
            article_ids_prefill = ",".join(str(i) for i in (saved.get("article_ids") or []))
            loaded_saved_for_sensitivity = saved
    from_feed = request.args.get("from_feed", "").strip()
    if from_feed:
        days = min(30, max(1, int(request.args.get("days", 7) or 7)))
        impact = _parse_impact_param(request.args.get("impact"))
        feed_articles = get_articles(limit=50, days=days, min_impact=impact)
        article_ids_prefill = ",".join(str(a["id"]) for a in feed_articles)
    from_view_id = request.args.get("from_view", type=int)
    if from_view_id:
        view = get_saved_view(from_view_id, user_id=uid_briefing)
        if view and view.get("params"):
            p = view["params"]
            feed_articles = get_articles(
                limit=50,
                source=p.get("source"),
                topic=p.get("topic"),
                topics_list=p.get("topics_list"),
                days=p.get("days"),
                date_from=p.get("date_from"),
                date_to=p.get("date_to"),
                min_impact=p.get("min_impact"),
                country=p.get("country"),
                countries_list=p.get("countries_list"),
                risk_category=p.get("risk_category"),
                risk_categories_list=p.get("risk_categories_list"),
            )
            article_ids_prefill = ",".join(str(a["id"]) for a in feed_articles)
    from_watchlist_id = request.args.get("from_watchlist", type=int)
    if from_watchlist_id:
        wl_articles = get_articles_for_watchlist(from_watchlist_id, limit=50)
        article_ids_prefill = ",".join(str(a["id"]) for a in wl_articles)
    from_alert_id = request.args.get("from_alert", type=int)
    if from_alert_id:
        alert_articles = get_alert_matches(from_alert_id, days=7, limit=50)
        article_ids_prefill = ",".join(str(a["id"]) for a in alert_articles)
    recent_articles = get_articles(limit=20)
    saved_briefings = get_saved_briefings(user_id=uid_briefing)
    saved_views = get_saved_views(user_id=uid_briefing)
    watchlists = get_watchlists()
    alerts = get_alerts(user_id=uid_briefing)
    edit_briefing_id = request.args.get("edit", type=int)
    edit_briefing = None
    if edit_briefing_id:
        edit_briefing = get_saved_briefing(edit_briefing_id, user_id=uid_briefing)
        if not edit_briefing:
            edit_briefing_id = None
        else:
            edit_briefing["name"] = edit_briefing.get("name") or "Briefing"
    urow = get_user_by_id(uid_briefing) if uid_briefing else None
    briefing_sensitivity_default = "internal"
    briefing_legal_default = False
    if edit_briefing:
        briefing_sensitivity_default = edit_briefing.get("sensitivity_tier") or "internal"
        briefing_legal_default = bool(edit_briefing.get("legal_review_required"))
    elif loaded_saved_for_sensitivity:
        briefing_sensitivity_default = loaded_saved_for_sensitivity.get("sensitivity_tier") or "internal"
        briefing_legal_default = bool(loaded_saved_for_sensitivity.get("legal_review_required"))
    elif urow:
        briefing_sensitivity_default = urow.get("default_sensitivity_tier") or "internal"
        briefing_legal_default = bool(urow.get("default_legal_review"))
    return render_template(
        "briefing_builder.html",
        title=title,
        intro=intro,
        article_ids_prefill=article_ids_prefill,
        recent_articles=recent_articles,
        saved_briefings=saved_briefings,
        saved_views=saved_views,
        watchlists=watchlists,
        alerts=alerts,
        edit_briefing_id=edit_briefing_id,
        edit_briefing=edit_briefing,
        briefing_sensitivity_default=briefing_sensitivity_default,
        briefing_legal_default=briefing_legal_default,
    )


@bp.route("/export/briefing/save", methods=["POST"])
@login_required
def export_briefing_save():
    """Save current briefing config (title, intro, article_ids) and redirect to builder."""
    name = (request.form.get("name") or "Briefing").strip() or "Briefing"
    title = (request.form.get("title") or "Briefing").strip()
    intro = (request.form.get("intro") or "").strip()
    ids_raw = request.form.get("article_ids") or ""
    article_ids = []
    for x in ids_raw.replace(",", "\n").split():
        x = x.strip()
        if x.isdigit():
            article_ids.append(int(x))
    st = _parse_sensitivity_tier(request.form.get("sensitivity_tier"))
    lr = request.form.get("legal_review_required") == "1"
    add_saved_briefing(
        name, title, intro, article_ids[:50], user_id=current_user.id, sensitivity_tier=st, legal_review_required=lr
    )
    return redirect(url_for("main.export_briefing"))


@bp.route("/export/briefing/load/<int:briefing_id>")
def export_briefing_load(briefing_id):
    """Load a saved briefing and redirect to builder with prefilled form."""
    return redirect(url_for("main.export_briefing", load=briefing_id))


@bp.route("/export/briefing/edit/<int:briefing_id>")
def export_briefing_edit(briefing_id):
    """Load a saved briefing into the builder in edit mode (Update button)."""
    return redirect(url_for("main.export_briefing", load=briefing_id, edit=briefing_id))


@bp.route("/export/briefing/update/<int:briefing_id>", methods=["POST"])
@login_required
def export_briefing_update(briefing_id):
    """Update an existing saved briefing and redirect to builder."""
    name = (request.form.get("name") or "Briefing").strip() or "Briefing"
    title = (request.form.get("title") or "Briefing").strip()
    intro = (request.form.get("intro") or "").strip()
    ids_raw = request.form.get("article_ids") or ""
    article_ids = []
    for x in ids_raw.replace(",", "\n").split():
        x = x.strip()
        if x.isdigit():
            article_ids.append(int(x))
    st = _parse_sensitivity_tier(request.form.get("sensitivity_tier"))
    lr = request.form.get("legal_review_required") == "1"
    update_saved_briefing(
        briefing_id,
        name,
        title,
        intro,
        article_ids[:50],
        user_id=current_user.id,
        sensitivity_tier=st,
        legal_review_required=lr,
    )
    return redirect(url_for("main.export_briefing"))


@bp.route("/export/briefing/duplicate/<int:briefing_id>")
@login_required
def export_briefing_duplicate(briefing_id):
    """Duplicate a saved briefing as 'Copy of …' and open in builder."""
    new_id = duplicate_saved_briefing(briefing_id, user_id=current_user.id)
    if new_id:
        return redirect(url_for("main.export_briefing_load", briefing_id=new_id))
    return redirect(url_for("main.export_briefing"))


@bp.route("/export/briefing/delete/<int:briefing_id>", methods=["POST"])
@login_required
def export_briefing_delete(briefing_id):
    delete_saved_briefing(briefing_id, user_id=current_user.id)
    return redirect(url_for("main.export_briefing"))


@bp.route("/export/briefing/docx", methods=["POST"])
def export_briefing_docx():
    """Generate and download briefing as DOCX (same format options as PDF)."""
    title = (request.form.get("title") or "Briefing").strip()
    intro = (request.form.get("intro") or "").strip()
    format_type = (request.form.get("format") or "summary_link").strip() or "summary_link"
    valid_formats = ("minimal", "summary_only", "summary_link", "with_takeaways", "executive", "full")
    if format_type not in valid_formats:
        format_type = "summary_link"
    grouping = (request.form.get("grouping") or "none").strip() or "none"
    if grouping not in ("none", "topic", "event_type", "date"):
        grouping = "none"
    sort_by = (request.form.get("sort_by") or "order").strip() or "order"
    report_date = (request.form.get("report_date") or "").strip()
    if not report_date:
        report_date = datetime.utcnow().strftime("%Y-%m-%d")
    ids_raw = request.form.get("article_ids") or ""
    article_ids = []
    for x in ids_raw.replace(",", "\n").split():
        x = x.strip()
        if x.isdigit():
            article_ids.append(int(x))
    articles = []
    for aid in article_ids[:50]:
        a = get_article(aid)
        if a:
            articles.append(a)
    if sort_by == "impact":
        articles = sorted(articles, key=lambda x: (-(x.get("impact_score") or 0), (x.get("published_utc") or "")[:10] or ""))
    grouped = _briefing_group_articles(articles, grouping)
    executive_summary = _briefing_executive_summary(articles)
    sensitivity_tier = _parse_sensitivity_tier(request.form.get("sensitivity_tier"))
    legal_review_required = request.form.get("legal_review_required") == "1"
    from app.export_docs import build_briefing_docx
    content = build_briefing_docx(
        title, intro, articles,
        format_type=format_type,
        grouped=grouped,
        executive_summary=executive_summary,
        report_date=report_date,
        sensitivity_tier=sensitivity_tier,
        legal_review_required=legal_review_required,
    )
    from flask import send_file
    from io import BytesIO
    filename = (title or "briefing").replace(" ", "-")[:50] + ".docx"
    return send_file(
        BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        as_attachment=True,
        download_name=filename,
    )
